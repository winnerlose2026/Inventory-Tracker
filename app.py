#!/usr/bin/env python3
"""Inventory Tracker - Flask Web GUI"""

import io
import os
import secrets
from datetime import datetime, timedelta
from flask import (
    Flask, jsonify, request, render_template, send_file, make_response,
    session, redirect, url_for,
)
from inventory_tracker import (
    load_inventory, save_inventory, load_usage, save_usage,
    add_item, update_item, record_usage, restock, remove_item,
    reverse_usage,
)

app = Flask(__name__)

# Session config — 30-day signed cookies. SECRET_KEY should come from Render's
# environment (otherwise sessions reset on every redeploy when the random
# fallback regenerates).
app.config["SECRET_KEY"] = (
    os.environ.get("FLASK_SECRET_KEY")
    or os.environ.get("SECRET_KEY")
    or secrets.token_hex(32)
)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# HTTPS-only cookie in production; on localhost dev we'd lock ourselves out.
app.config["SESSION_COOKIE_SECURE"] = (
    os.environ.get("FLASK_ENV", "").lower() != "development"
)
app.permanent_session_lifetime = timedelta(days=30)


# ---------------------------------------------------------------------------
# CORS — lets a Shopify-hosted page (or any allowed origin) call /api/*.
# Set ALLOWED_ORIGINS to a comma-separated list, e.g.
#   ALLOWED_ORIGINS=https://your-store.myshopify.com,https://your-store.com
# Leave unset during local development.
# ---------------------------------------------------------------------------

def _allowed_origins() -> list[str]:
    raw = os.environ.get("ALLOWED_ORIGINS", "").strip()
    return [o.strip() for o in raw.split(",") if o.strip()]


def _origin_allowed(origin: str) -> bool:
    if not origin:
        return False
    allowed = _allowed_origins()
    if not allowed:
        return False
    if "*" in allowed:
        return True
    return origin in allowed


@app.after_request
def _apply_cors(response):
    origin = request.headers.get("Origin", "")
    if _origin_allowed(origin):
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Vary"] = "Origin"
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Inventory-Token"
        response.headers["Access-Control-Max-Age"] = "600"
    return response


@app.route("/api/<path:_any>", methods=["OPTIONS"])
def _cors_preflight(_any):
    # Short-circuit the preflight; _apply_cors adds the headers.
    return make_response("", 204)


# ---------------------------------------------------------------------------
# Auth — session-based login for the browser, INVENTORY_API_TOKEN header for
# cron jobs and scripts. A write request is authorised if EITHER:
#   1. The browser session has session["user"] set (humans), OR
#   2. The request carries the right X-Inventory-Token header (cron/scripts).
# Reads on /api/* stay open so embedded widgets (Shopify storefront, etc.)
# can keep loading data without a session cookie.
# ---------------------------------------------------------------------------

def _user_logged_in() -> bool:
    return bool(session.get("user"))


def _has_valid_api_token() -> bool:
    expected = os.environ.get("INVENTORY_API_TOKEN", "").strip()
    if not expected:
        return False
    got = (request.headers.get("X-Inventory-Token") or "").strip()
    return bool(got) and secrets.compare_digest(got, expected)


def _is_authenticated() -> bool:
    return _user_logged_in() or _has_valid_api_token()


@app.before_request
def _gate_writes():
    # OPTIONS preflight, login flow, and static files are always open.
    if request.method == "OPTIONS":
        return
    if request.endpoint in ("login", "logout", "static"):
        return
    # Writes to /api/* require either a session or the API token.
    if request.method in ("POST", "PUT", "DELETE") and request.path.startswith("/api/"):
        if not _is_authenticated():
            return jsonify({"ok": False, "error": "unauthorized"}), 401


@app.route("/api/auth/check")
def api_auth_check():
    """Tell the page who's logged in (or whether the API token is valid)."""
    if _user_logged_in():
        return jsonify({
            "required": True,
            "authorized": True,
            "user": session.get("user"),
            "auth_type": "session",
        })
    if _has_valid_api_token():
        return jsonify({
            "required": True,
            "authorized": True,
            "user": None,
            "auth_type": "token",
        })
    return jsonify({"required": True, "authorized": False})


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    next_url = request.args.get("next") or request.form.get("next") or "/"
    if not next_url.startswith("/") or next_url.startswith("//"):
        next_url = "/"

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        # Allowed users come from INVENTORY_USERNAMES (comma-separated) with
        # backward-compat fallback to the legacy single INVENTORY_USERNAME.
        # All comparisons are case-insensitive so "jd" / "JD" / "Jd" all match.
        users_raw = (os.environ.get("INVENTORY_USERNAMES")
                     or os.environ.get("INVENTORY_USERNAME", ""))
        allowed = {u.strip().lower()
                   for u in users_raw.split(",") if u.strip()}
        expected_pass = os.environ.get("INVENTORY_PASSWORD", "")

        # Both the user list and password must be set for login to be
        # possible. If either is missing, the operator hasn't finished
        # configuration — fail closed.
        if allowed and expected_pass \
                and username.lower() in allowed \
                and secrets.compare_digest(password, expected_pass):
            session.permanent = True
            # Preserve the casing the user typed so the header chip reads
            # "Jay" / "JD" the way they signed in, not the env-var spelling.
            session["user"] = username
            return redirect(next_url)
        error = "Invalid username or password."

    return render_template("login.html", error=error, next_url=next_url)


@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

@app.route("/favicon.ico")
def favicon_ico():
    """Direct /favicon.ico requests bypass the template's <link> tag.
    Send the static file straight from /static/favicon.ico so the tab
    icon shows up even when the browser doesn't read the head first."""
    return app.send_static_file("favicon.ico")


@app.route("/")
def index():
    # Humans must be logged in to see the dashboard. The API token is for
    # scripts hitting /api/*, not for serving the HTML page.
    if not _user_logged_in():
        return redirect(url_for("login", next=request.full_path or "/"))
    return render_template("index.html", current_user=session.get("user", ""))


# ---------------------------------------------------------------------------
# API – Inventory
# ---------------------------------------------------------------------------

def _enrich_on_order(item: dict) -> dict:
    """Add convenience fields summarising pending on_order entries."""
    pending = item.get("on_order") or []
    total = round(sum(float(p.get("qty") or 0) for p in pending), 2)
    etas = [p.get("eta", "") for p in pending if p.get("eta")]
    next_eta = min(etas) if etas else ""
    item["on_order_qty"] = total
    item["on_order_next_eta"] = next_eta
    return item


@app.route("/api/inventory")
def api_inventory():
    inv = load_inventory()
    return jsonify([_enrich_on_order(dict(v)) for v in inv.values()])


@app.route("/api/inventory", methods=["POST"])
def api_add():
    d = request.json
    add_item(
        name=d["name"],
        quantity=float(d["quantity"]),
        unit=d["unit"],
        category=d.get("category", "general"),
        low_stock_threshold=float(d.get("low_stock_threshold", 5)),
        price=float(d.get("price", 0)),
        distributor=d.get("distributor", ""),
        warehouse=d.get("warehouse", ""),
        case_cost=float(d.get("case_cost", 0)),
        case_size=int(d.get("case_size", 0) or 0),
        weekly_usage=float(d.get("weekly_usage", 0)),
    )
    return jsonify({"ok": True})


@app.route("/api/inventory/<path:name>", methods=["PUT"])
def api_update(name):
    d = request.json
    update_item(
        name,
        quantity=float(d["quantity"]) if "quantity" in d else None,
        unit=d.get("unit"),
        category=d.get("category"),
        low_stock_threshold=float(d["low_stock_threshold"]) if "low_stock_threshold" in d else None,
        price=float(d["price"]) if "price" in d else None,
        distributor=d.get("distributor"),
        warehouse=d.get("warehouse"),
        case_cost=float(d["case_cost"]) if "case_cost" in d else None,
        case_size=int(d["case_size"]) if "case_size" in d else None,
        weekly_usage=float(d["weekly_usage"]) if "weekly_usage" in d else None,
    )
    return jsonify({"ok": True})


@app.route("/api/inventory/<path:name>", methods=["DELETE"])
def api_remove(name):
    remove_item(name)
    return jsonify({"ok": True})


@app.route("/api/on-order/ship-date", methods=["POST"])
def api_on_order_ship_date():
    """Set (or clear) a ship_date on all pending on_order entries for a PO.

    Body:
      po_number   (required)
      ship_date   ISO date or empty/null to clear
      item_key    (optional) limit the update to a single SKU; default
                  is to update every SKU's on_order list that carries
                  this PO. Per-PO is the common case because a PO ships
                  as a whole — every line item arrives together.

    Behavior:
      - Stores ship_date on each matching entry.
      - Computes arrival_date = ship_date + 7 days and stores it.
      - Clearing ship_date (empty string / null) also clears
        arrival_date, returning the entry to the default 30-day-from-
        ordered_at rollover.
      - inventory.load_inventory()'s rollover pass picks up the new
        arrival_date on its next call, so a ship_date that's already in
        the past promotes the entry immediately (after the +7).
    """
    from datetime import datetime, timedelta
    from inventory_tracker import load_inventory, save_inventory

    body = request.json or {}
    po_number = (body.get("po_number") or "").strip()
    if not po_number:
        return jsonify({"ok": False, "error": "po_number required"}), 400
    item_key = (body.get("item_key") or "").strip().lower() or None
    ship_raw = body.get("ship_date")

    # Empty value clears.
    if ship_raw is None or (isinstance(ship_raw, str) and not ship_raw.strip()):
        ship_iso = ""
        arrival_iso = ""
    else:
        try:
            ship_dt = datetime.fromisoformat(str(ship_raw).strip())
        except ValueError:
            return jsonify({
                "ok": False,
                "error": "ship_date must be ISO 8601 (YYYY-MM-DD or full datetime)",
            }), 400
        ship_iso = ship_dt.isoformat()
        arrival_iso = (ship_dt + timedelta(days=7)).isoformat()

    inv = load_inventory()
    updated = 0
    touched_items = []
    for key, item in inv.items():
        if item_key and key != item_key:
            continue
        pending = item.get("on_order") or []
        for entry in pending:
            if (entry.get("po_number") or "") != po_number:
                continue
            entry["ship_date"] = ship_iso
            entry["arrival_date"] = arrival_iso
            updated += 1
            if item.get("name") not in touched_items:
                touched_items.append(item.get("name") or key)
    save_inventory(inv)
    return jsonify({
        "ok": True,
        "po_number": po_number,
        "ship_date": ship_iso,
        "arrival_date": arrival_iso,
        "entries_updated": updated,
        "items": touched_items,
    })


@app.route("/api/admin/po-order-date", methods=["POST"])
def api_admin_po_order_date():
    """Set (or clear) ``ordered_at`` on every on_order entry matching a
    PO number. Use when a PO was ingested before po_order_date was
    plumbed through the scanner — the on_order entry inherited the
    ingestion timestamp instead of the PDF's printed "ORDER DATE",
    which threw off the Pending POs view and the 30-day rollover ETA.

    Body:
      po_number       required.
      order_date      ISO date (YYYY-MM-DD) or empty/null to clear.
      recompute_eta   bool; default true. If true and order_date is
                      set, also reset eta = order_date + lead_days
                      (entry-by-entry, using each entry's lead_days).

    Returns {ok, po_number, order_date, entries_updated, items}.
    """
    from datetime import datetime, timedelta
    from inventory_tracker import load_inventory, save_inventory

    body = request.json or {}
    po_number = (body.get("po_number") or "").strip()
    if not po_number:
        return jsonify({"ok": False, "error": "po_number required"}), 400
    order_raw = body.get("order_date")
    recompute_eta = bool(body.get("recompute_eta", True))

    if order_raw is None or (isinstance(order_raw, str)
                             and not order_raw.strip()):
        order_iso = ""
        order_dt = None
    else:
        try:
            order_dt = datetime.fromisoformat(str(order_raw).strip())
        except ValueError:
            return jsonify({
                "ok": False,
                "error": "order_date must be ISO 8601 (YYYY-MM-DD or full datetime)",
            }), 400
        order_iso = order_dt.isoformat()

    inv = load_inventory()
    updated = 0
    touched_items: list[str] = []
    for key, item in inv.items():
        pending = item.get("on_order") or []
        for entry in pending:
            if (entry.get("po_number") or "") != po_number:
                continue
            entry["ordered_at"] = order_iso
            if recompute_eta and order_dt is not None:
                lead = int(entry.get("lead_days") or 0)
                if lead > 0:
                    entry["eta"] = (order_dt + timedelta(days=lead)).isoformat()
            updated += 1
            name = item.get("name") or key
            if name not in touched_items:
                touched_items.append(name)
    save_inventory(inv)
    return jsonify({
        "ok": True,
        "po_number": po_number,
        "order_date": order_iso,
        "entries_updated": updated,
        "items": touched_items,
    })


@app.route("/api/admin/remove-po", methods=["POST"])
def api_admin_remove_po():
    """Drop all pending on_order entries matching a po_number.

    Used to manually retire a PO whose source email has been deleted or
    archived (so the auto-supersede via re-scan can't reach it). The
    operation does NOT touch already-rolled-over quantity or usage
    entries — only items still sitting in item["on_order"].
    """
    body = request.json or {}
    po_number = (body.get("po_number") or "").strip()
    reason = (body.get("reason") or "canceled by distributor").strip()
    if not po_number:
        return jsonify({"ok": False, "error": "po_number required"}), 400
    from inventory_tracker import (
        load_inventory, save_inventory,
        load_canceled_pos, save_canceled_pos,
    )
    inv = load_inventory()
    removed = 0
    affected = []
    for key, item in inv.items():
        pending = item.get("on_order") or []
        kept = [e for e in pending if (e.get("po_number") or "") != po_number]
        if len(kept) != len(pending):
            removed += len(pending) - len(kept)
            affected.append(item.get("name", key))
        item["on_order"] = kept
    save_inventory(inv)
    # Record the PO in the ignore list so the email scanner won't
    # re-ingest it from a still-sitting source email.
    canceled = load_canceled_pos()
    canceled[po_number] = {
        "canceled_at": datetime.now().isoformat(timespec="seconds"),
        "reason":      reason,
        "removed_entries": removed,
        "affected_items":  affected,
    }
    save_canceled_pos(canceled)
    return jsonify({
        "ok": True,
        "po_number": po_number,
        "removed_entries": removed,
        "affected_items": affected,
        "added_to_ignore_list": True,
    })


@app.route("/api/admin/uncancel-po", methods=["POST"])
def api_admin_uncancel_po():
    """Remove a po_number from the canceled-POs ignore list.

    Inverse of /api/admin/remove-po. Use when a PO was canceled in error
    (or when an operator needs to allow a re-ingest of a PO whose stored
    entries were wiped). Does NOT restore the previously-removed
    on_order entries -- those have to come back via a fresh ingest from
    the source email or a manual POST to /api/email/ingest-events.
    """
    body = request.json or {}
    po_number = (body.get("po_number") or "").strip()
    if not po_number:
        return jsonify({"ok": False, "error": "po_number required"}), 400
    from inventory_tracker import load_canceled_pos, save_canceled_pos
    canceled = load_canceled_pos()
    prior = canceled.pop(po_number, None)
    save_canceled_pos(canceled)
    return jsonify({
        "ok": True,
        "po_number": po_number,
        "was_canceled": prior is not None,
        "prior_entry": prior,
    })


# ---------------------------------------------------------------------------
# API – Chefs Warehouse POs
# ---------------------------------------------------------------------------
# CW POs live in their own JSON file (data/chefs_warehouse_pos.json) so
# they never touch inventory.json. The Pending POs tab merges them in
# for display via /api/chefs-warehouse/pos; the Inventory tab never
# shows them.

def _cw_po_summary(record: dict) -> dict:
    """Shape a stored CW PO record into the same group dict the
    Pending POs frontend uses for on_order groups (so the merge in
    loadPendingPOs is shape-compatible)."""
    lines = record.get("lines") or []
    def _ln(L):
        return {
            "variety": L.get("variety") or "",
            "qty":     float(L.get("qty") or 0),
            "unit":    L.get("unit") or "cs",
            "name":    (L.get("description") or "").title(),
            "sliced":  bool(L.get("sliced")),
            "cw_item": L.get("cw_item") or "",
        }
    total_cs = float(record.get("total_cs")
                     or sum(float(L.get("qty") or 0) for L in lines))
    return {
        "po_number":    record.get("po_number") or "",
        "po_revision":  record.get("po_revision") or "",
        "distributor":  "Chefs Warehouse",
        "warehouse":    record.get("warehouse") or "",
        "dc_code":      record.get("dc_code") or "",
        "ordered_at":   record.get("ordered_at") or "",
        "eta":          record.get("eta") or "",
        "ship_date":    record.get("ship_date") or "",
        "arrival_date": record.get("arrival_date") or "",
        "buyer_name":   record.get("buyer_name") or "",
        "ship_to_name": record.get("ship_to_name") or "",
        "total_cs":     round(total_cs, 2),
        "total_usd":    record.get("total_usd"),
        "lines":        [_ln(L) for L in lines],
        "source":       record.get("source") or "",
        "source_subject": record.get("source_subject") or "",
    }


@app.route("/api/chefs-warehouse/pos")
def api_chefs_warehouse_pos():
    """List all Chefs Warehouse POs (active + arrived).

    Query params:
      ``status``  filter to "pending" (default), "arrived", "canceled",
                  or "all".

    A CW PO is "pending" until an operator sets a ship_date or
    arrival_date that's in the past, or marks it canceled. The Pending
    POs tab fetches the default (pending only).
    """
    from inventory_tracker import load_chefs_warehouse_pos, load_canceled_pos
    records = load_chefs_warehouse_pos()
    canceled = load_canceled_pos()

    status_filter = (request.args.get("status") or "pending").lower()
    freight_idx = _freight_ship_date_index()

    out = []
    now = datetime.now()
    for r in records:
        po_num = (r.get("po_number") or "").strip()
        if r.get("canceled") or po_num in canceled:
            status = "canceled"
        else:
            # Auto-ETA rule (2026-05-27): CW POs only auto-flip to
            # "arrived" off an OPERATOR-set arrival_date. The parser-set
            # eta (from the PDF's printed delivery date or a 30-day
            # fallback) is ignored for status — the vendor's promise of
            # when they'll deliver is not the same as confirmation that
            # the truck showed up. Operator types ship_date -> the
            # ship-date endpoint stores arrival_date = ship_date + 7d,
            # and that's what flips us to "arrived" past its date.
            arrival_str = (r.get("arrival_date") or "").strip()
            arrival_dt = None
            if arrival_str:
                try:
                    arrival_dt = datetime.fromisoformat(arrival_str)
                except ValueError:
                    arrival_dt = None
            if arrival_dt and arrival_dt <= now:
                status = "arrived"
            else:
                status = "pending"

        if status_filter != "all" and status != status_filter:
            continue

        item = _cw_po_summary(r)
        item["status"] = status
        # Backfill a missing ship date on ARRIVED CW POs from the freight
        # invoice index (links by PO #). Display-only; never persisted and
        # never recomputes arrival, so the computed status is untouched.
        if status == "arrived" and not (item.get("ship_date") or "").strip():
            sd = freight_idx.get(_norm_po_key(po_num))
            if sd:
                item["ship_date"] = sd
                item["ship_date_source"] = "freight"
        out.append(item)

    out.sort(key=lambda x: (x.get("ordered_at") or "", x.get("po_number") or ""))
    return jsonify({"ok": True, "count": len(out), "pos": out})


@app.route("/api/chefs-warehouse/ingest-pos", methods=["POST"])
def api_chefs_warehouse_ingest_pos():
    """Accept externally-parsed CW PO records and apply them.

    Used by the Cowork scheduled routine that fetches Graph mail
    directly: it pulls CW PDFs, parses them with the same parser, and
    POSTs the dict-form records here. Mirrors /api/email/ingest-events
    but for the CW-only channel.

    Request body:
        {
          "dry_run": false,
          "source":  "cowork-routine",
          "cw_pos":  [ <ChefsWarehousePO as dict>, ... ]
        }
    """
    import traceback as _tb
    try:
        from sync_inventory import _apply_cw_pos
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False,
                        "error": f"import failed: {type(exc).__name__}: {exc}"}), 500
    body = request.json or {}
    dry_run = bool(body.get("dry_run", False))
    source = str(body.get("source") or "external").strip() or "external"
    cw_pos_raw = body.get("cw_pos") or []
    if not isinstance(cw_pos_raw, list):
        return jsonify({"ok": False, "error": "cw_pos must be a list"}), 400
    try:
        report = _apply_cw_pos(cw_pos_raw, dry_run=dry_run, source=source)
    except Exception as exc:  # noqa: BLE001
        return jsonify({
            "ok": False,
            "error": f"{type(exc).__name__}: {exc}",
            "traceback": _tb.format_exc()[-2000:],
        }), 500
    return jsonify({"ok": True, "dry_run": dry_run, "report": report})


# ---------------------------------------------------------------------------
# API – Freight (Lineage outbound shipping invoices)
# ---------------------------------------------------------------------------
# Lineage Freight Management LLC invoices arrive in the same mailbox as
# distributor POs (sender: noreply@tms.blujaysolutions.net, subject:
# "Billable Invoice(s) from LINEAGE FREIGHT MANAGEMENT LLC"). The
# attachment is a .zip containing one or more PDF invoices, one per
# shipment H&H sent to a DC. We parse the PDFs in the cron-side scan
# script (scripts/cowork_graph_scan.py) and POST the dict-form records
# here, so the web service doesn't need outbound network access.
#
# Records live in data/freight_invoices.json and never touch inventory
# state. The Freight Costs tab reads /api/freight/invoices for display.


@app.route("/api/freight/invoices")
def api_freight_invoices():
    """Return all Lineage freight invoices, sorted by ship_date desc.

    Query params:
        ``dc``        filter to a single destination DC, e.g. "Manassas, VA"
        ``distributor`` filter to "US Foods" | "Cheney Brothers" | "Chefs Warehouse"
        ``since``     ISO date (YYYY-MM-DD); only invoices on or after
        ``until``     ISO date; only invoices on or before
    """
    from inventory_tracker import load_freight_invoices
    records = load_freight_invoices()
    dc = (request.args.get("dc") or "").strip()
    dist = (request.args.get("distributor") or "").strip()
    since = (request.args.get("since") or "").strip()
    until = (request.args.get("until") or "").strip()
    out = []
    for r in records:
        if dc and r.get("dest_dc") != dc:
            continue
        if dist and r.get("distributor") != dist:
            continue
        sd = r.get("ship_date") or ""
        if since and sd and sd < since:
            continue
        if until and sd and sd > until:
            continue
        out.append(r)
    out.sort(key=lambda x: (x.get("ship_date") or "", x.get("invoice_number") or ""),
             reverse=True)
    # Also derive a few summary stats so the dashboard doesn't have to
    # recompute them client-side.
    total_cost = sum(float(r.get("total_due") or 0) for r in out)
    total_pallets = sum(int(r.get("pallets") or 0) for r in out)
    total_cases = sum(int(r.get("cases") or 0) for r in out)
    # Aggregate by line-item category (Basis Item / Fuel Surcharge /
    # Lumper / Detention / ...). Each Lineage invoice breaks the total
    # into one or more line items, and operators want to see the
    # categories rolled up across the visible filter window.
    by_category: dict = {}
    by_dc: dict = {}
    for r in out:
        dc = r.get("dest_dc") or "Unknown"
        bd = by_dc.setdefault(dc, {"count": 0, "cost": 0.0,
                                   "pallets": 0, "cases": 0})
        bd["count"]   += 1
        bd["cost"]    += float(r.get("total_due") or 0)
        bd["pallets"] += int(r.get("pallets") or 0)
        bd["cases"]   += int(r.get("cases") or 0)
        for li in (r.get("line_items") or []):
            desc = (li.get("description") or "Other").strip()
            bc = by_category.setdefault(desc, {"count": 0, "total": 0.0})
            bc["count"] += 1
            bc["total"] += float(li.get("total") or 0)
    # Round for clean transit
    for v in by_category.values():
        v["total"] = round(v["total"], 2)
    for v in by_dc.values():
        v["cost"] = round(v["cost"], 2)
        v["per_pallet"] = round(v["cost"] / v["pallets"], 2) if v["pallets"] else 0
        v["per_case"]   = round(v["cost"] / v["cases"], 4)   if v["cases"]   else 0

    summary = {
        "count":          len(out),
        "total_cost":     round(total_cost, 2),
        "total_pallets":  total_pallets,
        "total_cases":    total_cases,
        "avg_cost_per_pallet": round(total_cost / total_pallets, 2) if total_pallets else 0,
        "avg_cost_per_case":   round(total_cost / total_cases, 4) if total_cases else 0,
        "by_category":    by_category,
        "by_dc":          by_dc,
    }
    return jsonify({"invoices": out, "summary": summary})


@app.route("/api/freight/ingest", methods=["POST"])
def api_freight_ingest():
    """Accept externally-parsed Lineage freight invoice records.

    Called by the 6h cron mailbox scan after it has fetched
    Lineage emails, unzipped the attachment, and parsed each PDF.

    Request body:
        {
          "dry_run": false,
          "source":  "cowork-routine/lineage",
          "invoices": [ <FreightInvoice as dict>, ... ]
        }

    Dedup is by invoice_number — re-ingesting the same invoice replaces
    the prior record. Records without invoice_number are rejected.
    """
    import traceback as _tb
    from datetime import datetime
    try:
        from inventory_tracker import (
            load_freight_invoices, save_freight_invoices,
        )
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False,
                        "error": f"import failed: {type(exc).__name__}: {exc}"}), 500
    body = request.json or {}
    dry_run = bool(body.get("dry_run", False))
    source = str(body.get("source") or "external").strip() or "external"
    raw = body.get("invoices") or []
    if not isinstance(raw, list):
        return jsonify({"ok": False, "error": "invoices must be a list"}), 400

    existing = load_freight_invoices()
    by_inv = {str(r.get("invoice_number") or ""): r for r in existing if r.get("invoice_number")}

    added = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

    for idx, rec in enumerate(raw):
        if not isinstance(rec, dict):
            errors.append(f"invoices[{idx}]: not an object")
            skipped += 1
            continue
        inv_no = str(rec.get("invoice_number") or "").strip()
        if not inv_no:
            errors.append(f"invoices[{idx}]: missing invoice_number")
            skipped += 1
            continue
        # Re-derive cost_per_pallet / cost_per_case server-side so we
        # never trust client math. Clients may omit them entirely.
        try:
            total = float(rec.get("total_due") or 0)
            pallets = int(rec.get("pallets") or 0)
            cases = int(rec.get("cases") or 0)
        except (TypeError, ValueError):
            total = pallets = cases = 0
        rec["cost_per_pallet"] = round(total / pallets, 2) if pallets else 0
        rec["cost_per_case"]   = round(total / cases, 4) if cases else 0
        # Stamp ingest source + time
        rec.setdefault("source", source)
        rec["ingested_at"] = rec.get("ingested_at") or now

        if inv_no in by_inv:
            by_inv[inv_no] = rec
            updated += 1
        else:
            by_inv[inv_no] = rec
            added += 1

    if not dry_run:
        merged = sorted(by_inv.values(),
                        key=lambda x: (x.get("ship_date") or "",
                                       x.get("invoice_number") or ""))
        try:
            save_freight_invoices(merged)
        except Exception as exc:  # noqa: BLE001
            return jsonify({
                "ok": False,
                "error": f"save failed: {type(exc).__name__}: {exc}",
                "traceback": _tb.format_exc()[-2000:],
            }), 500

    return jsonify({
        "ok": True,
        "dry_run": dry_run,
        "report": {
            "source": source,
            "added": added,
            "updated": updated,
            "skipped": skipped,
            "total_after": len(by_inv),
            "errors": errors,
        },
    })


@app.route("/api/freight/scan", methods=["POST"])
def api_freight_scan():
    """Deep-sweep the configured MS365 mailboxes for Lineage Freight invoices.

    Same Graph credentials the legacy /api/email/scan path uses (web
    service has MS365_TENANT_ID / MS365_CLIENT_ID / MS365_CLIENT_SECRET
    / MS365_USER) but with a Lineage-only sender filter so we can sweep
    a very wide window without burning budget on every PO PDF in the
    inbox.

    Request body (all optional):
        {
          "dry_run":        false,
          "lookback_days":  365,    # default 730 (~2 years)
          "mailboxes":      "JD@ms.hhbagels.com,info@ms.hhbagels.com",
                                    # default MS365_USER env var
          "max_messages":   500     # default 500, hard cap 5000
        }

    Returns:
        {"ok": true, "report": {"messages_seen": N, "invoices_added": A,
         "invoices_updated": U, "errors": [...]}}
    """
    import io
    import json as _json
    import os as _os
    import re as _re
    import sys as _sys
    import traceback as _tb
    import urllib.error
    import urllib.parse
    import urllib.request
    import zipfile
    from dataclasses import asdict
    from datetime import datetime, timezone, timedelta

    body = request.json or {}
    dry_run = bool(body.get("dry_run", False))
    # Optional explicit ISO date window (YYYY-MM-DD). When set, supersedes
    # lookback_days. Lets the caller batch a multi-year backfill into
    # bite-sized 6-month windows so each request fits within gunicorn's
    # 180s timeout.
    since_date = (body.get("since_date") or "").strip()
    until_date = (body.get("until_date") or "").strip()
    try:
        lookback_days = int(body.get("lookback_days") or 730)
    except (TypeError, ValueError):
        lookback_days = 730
    lookback_days = max(1, min(lookback_days, 3650))
    try:
        max_messages = int(body.get("max_messages") or 500)
    except (TypeError, ValueError):
        max_messages = 500
    max_messages = max(1, min(max_messages, 5000))
    mailboxes_raw = (body.get("mailboxes")
                     or _os.environ.get("MS365_USER", "")).strip()
    mailboxes = [m.strip() for m in mailboxes_raw.split(",") if m.strip()]

    tenant = _os.environ.get("MS365_TENANT_ID")
    client_id = _os.environ.get("MS365_CLIENT_ID")
    client_secret = _os.environ.get("MS365_CLIENT_SECRET")
    if not all([tenant, client_id, client_secret, mailboxes]):
        return jsonify({
            "ok": False,
            "error": "MS365 credentials or mailboxes not configured",
        }), 200

    try:
        from integrations.lineage_freight_parser import parse_freight_pdf
        from inventory_tracker import (
            load_freight_invoices, save_freight_invoices,
        )
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False,
                        "error": f"import failed: {type(exc).__name__}: {exc}"}), 500

    GRAPH_BASE = "https://graph.microsoft.com/v1.0"

    def _token():
        url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
        data = urllib.parse.urlencode({
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        }).encode("utf-8")
        req = urllib.request.Request(url, data=data,
                                     headers={"Content-Type": "application/x-www-form-urlencoded"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            payload = _json.loads(resp.read().decode("utf-8"))
        return payload["access_token"]

    def _graph_get(token, path):
        url = path if path.startswith("http") else f"{GRAPH_BASE}{path}"
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            return _json.loads(resp.read().decode("utf-8"))

    def _graph_get_bytes(token, path):
        url = path if path.startswith("http") else f"{GRAPH_BASE}{path}"
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
        with urllib.request.urlopen(req, timeout=120) as resp:
            return resp.read()

    errors: list[str] = []
    seen = 0
    parsed_invoices: list[dict] = []
    diag_atts_seen = 0
    lineage_attachment_samples: list = []
    diag_zips = 0
    diag_pdfs_in_zip = 0
    diag_parse_none = 0
    try:
        token = _token()
        # Two ways to specify the window:
        #   - lookback_days (relative): now() - N days .. now()
        #   - since_date / until_date (absolute): explicit YYYY-MM-DD pair,
        #     useful for batching a multi-year backfill into 6-month chunks
        if since_date or until_date:
            try:
                since = datetime.fromisoformat(since_date) if since_date else (
                    datetime.now(timezone.utc) - timedelta(days=3650))
                since = since.replace(tzinfo=timezone.utc)
            except ValueError:
                return jsonify({"ok": False,
                                "error": f"bad since_date {since_date!r}"}), 200
            try:
                until = (datetime.fromisoformat(until_date)
                         if until_date else datetime.now(timezone.utc))
                until = until.replace(tzinfo=timezone.utc)
            except ValueError:
                return jsonify({"ok": False,
                                "error": f"bad until_date {until_date!r}"}), 200
            since_iso = since.replace(microsecond=0).isoformat().replace("+00:00", "Z")
            until_iso = until.replace(microsecond=0).isoformat().replace("+00:00", "Z")
            flt = ("hasAttachments eq true "
                   f"and receivedDateTime ge {since_iso} "
                   f"and receivedDateTime lt {until_iso}")
        else:
            since = datetime.now(timezone.utc) - timedelta(days=lookback_days)
            since_iso = since.replace(microsecond=0).isoformat().replace("+00:00", "Z")
            flt = ("hasAttachments eq true "
                   f"and receivedDateTime ge {since_iso}")
        LINEAGE_DOMAINS = ("tms.blujaysolutions.net", "blujaysolutions.net",
                           "tms.e2open.com", "e2open.com",
                           "lineagelogistics.com", "onelineage.com")

        # Track sender domains we see vs match, for diagnostics
        domain_seen: dict = {}
        lineage_subjects: list = []
        for mb in mailboxes:
            user = urllib.parse.quote(mb)
            # ALL_MAIL search ALL folders (not just Inbox) because Lineage
            # invoices may have been auto-filed via Outlook rules or
            # manually moved. Use AllItems via the standard "messages"
            # endpoint (no /mailFolders/Inbox prefix).
            list_url = (f"{GRAPH_BASE}/users/{user}/messages"
                        f"?$top=100&$filter={urllib.parse.quote(flt)}"
                        f"&$select=id,subject,from")
            fetched = 0
            while list_url and fetched < max_messages:
                page = _graph_get(token, list_url)
                msgs = page.get("value", [])
                if not msgs:
                    break
                for m in msgs:
                    seen += 1
                    if fetched >= max_messages:
                        break
                    # Post-filter by sender domain or subject keyword
                    sender = (((m.get("from") or {}).get("emailAddress") or {})
                              .get("address") or "").lower()
                    subj_check = (m.get("subject") or "").upper()
                    dom = sender.split("@", 1)[-1] if "@" in sender else ""
                    domain_seen[dom] = domain_seen.get(dom, 0) + 1
                    looks_lineage = (
                        any(dom == d or dom.endswith("." + d) for d in LINEAGE_DOMAINS)
                        or ("LINEAGE FREIGHT" in subj_check
                            and "BILLABLE INVOICE" in subj_check)
                    )
                    if not looks_lineage:
                        continue
                    fetched += 1
                    if len(lineage_subjects) < 5:
                        lineage_subjects.append((sender, m.get("subject") or ""))
                    mid = m.get("id") or ""
                    subj = m.get("subject") or ""
                    # List + fetch zip attachments
                    try:
                        atts_resp = _graph_get(token,
                            f"/users/{user}/messages/{urllib.parse.quote(mid)}/attachments"
                            "?$select=id,name,contentType,size")
                    except urllib.error.HTTPError as exc:
                        errors.append(f"{mid[:12]}.. list-att: HTTP {exc.code}")
                        continue
                    for a in (atts_resp.get("value") or []):
                        diag_atts_seen += 1
                        aname_raw = a.get("name") or ""
                        aname = aname_raw.lower()
                        actype = (a.get("contentType") or "").lower()
                        if len(lineage_attachment_samples) < 10:
                            lineage_attachment_samples.append({
                                "name": aname_raw,
                                "ctype": actype,
                                "size": a.get("size"),
                            })
                        if not (aname.endswith(".zip") or aname.endswith(".pdf")
                                or actype in ("application/zip",
                                              "application/x-zip-compressed",
                                              "application/pdf")):
                            if len(errors) < 5:
                                errors.append(f"skip att type={actype!r} name={aname!r}")
                            continue
                        try:
                            ab = _graph_get_bytes(token,
                                f"/users/{user}/messages/{urllib.parse.quote(mid)}"
                                f"/attachments/{urllib.parse.quote(a.get('id') or '')}/$value")
                        except urllib.error.HTTPError as exc:
                            errors.append(f"{mid[:12]}.. fetch-att: HTTP {exc.code}")
                            continue
                        pdfs: list[tuple[str, bytes]] = []
                        if aname.endswith(".zip") or actype in ("application/zip",
                                                                "application/x-zip-compressed"):
                            diag_zips += 1
                            try:
                                zf = zipfile.ZipFile(io.BytesIO(ab))
                                for info in zf.infolist():
                                    if info.filename.lower().endswith(".pdf"):
                                        pdfs.append((info.filename, zf.read(info.filename)))
                                        diag_pdfs_in_zip += 1
                            except zipfile.BadZipFile as exc:
                                errors.append(f"{mid[:12]}.. bad-zip: {exc}")
                                continue
                        else:
                            pdfs = [(a.get("name") or "invoice.pdf", ab)]
                        for fname, pb in pdfs:
                            try:
                                inv = parse_freight_pdf(
                                    pb, pdf_filename=fname,
                                    source_message_id=mid, source_subject=subj)
                            except Exception as exc:  # noqa: BLE001
                                errors.append(f"{mid[:12]}.. parse[{fname}]: {exc}")
                                continue
                            if inv is None:
                                diag_parse_none += 1
                                if len(errors) < 10:
                                    # Capture the first chars of extracted
                                    # PDF text so we can see what pypdf is
                                    # actually returning on the server.
                                    try:
                                        from pypdf import PdfReader as _PR
                                        import io as _io2
                                        _r = _PR(_io2.BytesIO(pb))
                                        _t = (_r.pages[0].extract_text() or "")[:160]
                                    except Exception as _e:
                                        _t = f"<text extract failed: {_e}>"
                                    errors.append(
                                        f"{mid[:12]}.. parse[{fname}] None "
                                        f"({len(pb)}b) txt={_t!r}")
                                continue
                            parsed_invoices.append(asdict(inv))
                # Next page
                list_url = page.get("@odata.nextLink")
    except urllib.error.HTTPError as exc:
        return jsonify({
            "ok": False,
            "error": f"Graph HTTP {exc.code}: {exc.reason}",
            "errors": errors,
        }), 200
    except Exception as exc:  # noqa: BLE001
        return jsonify({
            "ok": False,
            "error": f"{type(exc).__name__}: {exc}",
            "traceback": _tb.format_exc()[-1500:],
            "errors": errors,
        }), 200

    # Merge into freight_invoices.json
    existing = load_freight_invoices()
    by_inv = {str(r.get("invoice_number") or ""): r
              for r in existing if r.get("invoice_number")}
    added = updated = 0
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    for rec in parsed_invoices:
        ino = str(rec.get("invoice_number") or "").strip()
        if not ino:
            continue
        try:
            t = float(rec.get("total_due") or 0)
            pl = int(rec.get("pallets") or 0)
            cs = int(rec.get("cases") or 0)
        except (TypeError, ValueError):
            t = pl = cs = 0
        rec["cost_per_pallet"] = round(t / pl, 2) if pl else 0
        rec["cost_per_case"]   = round(t / cs, 4) if cs else 0
        rec["ingested_at"]     = rec.get("ingested_at") or now
        if ino in by_inv:
            by_inv[ino] = rec; updated += 1
        else:
            by_inv[ino] = rec; added += 1
    if not dry_run:
        merged = sorted(by_inv.values(),
                        key=lambda x: (x.get("ship_date") or "",
                                       x.get("invoice_number") or ""))
        save_freight_invoices(merged)

    # Trim domain_seen to top 25 for log readability
    top_domains = dict(sorted(domain_seen.items(), key=lambda x: -x[1])[:25])
    return jsonify({
        "ok": True,
        "dry_run": dry_run,
        "report": {
            "mailboxes":        mailboxes,
            "lookback_days":    lookback_days,
            "messages_seen":    seen,
            "invoices_parsed":  len(parsed_invoices),
            "invoices_added":   added,
            "invoices_updated": updated,
            "total_after":      len(by_inv),
            "errors":           errors[:50],
            "error_count":      len(errors),
            "top_sender_domains": top_domains,
            "sample_lineage_matches": lineage_subjects,
            "lineage_attachment_samples": lineage_attachment_samples,
            "diag": {
                "pypdf_version":        (lambda:
                    __import__("pypdf").__version__ if hasattr(__import__("pypdf"), "__version__") else "?")(),
                "attachments_seen":     diag_atts_seen,
                "zip_attachments":      diag_zips,
                "pdfs_extracted_from_zip": diag_pdfs_in_zip,
                "parse_returned_none":  diag_parse_none,
            },
        },
    })


@app.route("/api/chefs-warehouse/ship-date", methods=["POST"])
def api_chefs_warehouse_ship_date():
    """Set / clear ship_date (and the derived arrival_date) on a CW PO.

    Body: ``{"po_number": "...", "ship_date": "YYYY-MM-DD" | ""}``.
    arrival_date is set to ship_date + 7 days (CW transit lead) to
    match the on_order convention; clearing ship_date also clears it.
    """
    from datetime import timedelta
    from inventory_tracker import (
        load_chefs_warehouse_pos, save_chefs_warehouse_pos,
    )
    body = request.json or {}
    po_number = (body.get("po_number") or "").strip()
    ship_iso = (body.get("ship_date") or "").strip()
    if not po_number:
        return jsonify({"ok": False, "error": "po_number required"}), 400

    if ship_iso:
        try:
            ship_dt = datetime.fromisoformat(ship_iso)
        except ValueError:
            return jsonify({"ok": False,
                            "error": "ship_date must be YYYY-MM-DD"}), 400
        arrival_iso = (ship_dt + timedelta(days=7)).isoformat()
    else:
        ship_iso = ""
        arrival_iso = ""

    records = load_chefs_warehouse_pos()
    found = False
    for r in records:
        if (r.get("po_number") or "").strip() != po_number:
            continue
        r["ship_date"]    = ship_iso
        r["arrival_date"] = arrival_iso
        r["updated_at"]   = datetime.now().isoformat()
        found = True
    if not found:
        return jsonify({"ok": False,
                        "error": f"PO {po_number} not found in CW POs"}), 404
    save_chefs_warehouse_pos(records)
    return jsonify({
        "ok": True,
        "po_number": po_number,
        "ship_date": ship_iso,
        "arrival_date": arrival_iso,
    })


@app.route("/api/chefs-warehouse/cancel", methods=["POST"])
def api_chefs_warehouse_cancel():
    """Mark a CW PO canceled. Removes it from the default Pending POs
    list and adds the PO# to the shared canceled-POs ignore list so a
    re-scan of the same source email doesn't re-add it.
    """
    from inventory_tracker import (
        load_chefs_warehouse_pos, save_chefs_warehouse_pos,
        load_canceled_pos, save_canceled_pos,
    )
    body = request.json or {}
    po_number = (body.get("po_number") or "").strip()
    reason = (body.get("reason") or "canceled by distributor").strip()
    if not po_number:
        return jsonify({"ok": False, "error": "po_number required"}), 400

    records = load_chefs_warehouse_pos()
    found = False
    for r in records:
        if (r.get("po_number") or "").strip() != po_number:
            continue
        r["canceled"]        = True
        r["canceled_at"]     = datetime.now().isoformat(timespec="seconds")
        r["canceled_reason"] = reason
        found = True
    save_chefs_warehouse_pos(records)

    canceled = load_canceled_pos()
    canceled[po_number] = {
        "canceled_at": datetime.now().isoformat(timespec="seconds"),
        "reason":      reason,
        "distributor": "Chefs Warehouse",
    }
    save_canceled_pos(canceled)

    return jsonify({
        "ok": True,
        "po_number": po_number,
        "found_in_cw_file": found,
        "added_to_ignore_list": True,
    })


# ---------------------------------------------------------------------------
# API – Arrived (rolled-over) inventory POs
# ---------------------------------------------------------------------------
# When a USF/Cheney on_order entry's trigger date passes,
# inventory_tracker._rollover_on_order() drops it from item["on_order"]
# and _append_rollover_usage() writes a usage row tagged
# source="on_order_rollover". Those usage rows are the ONLY persistent
# record that an inventory-side PO arrived (the on_order entry is gone),
# so the Pending POs tab can't show arrived USF/Cheney POs from
# /api/inventory alone — that's why "Arrived" historically showed only
# Chefs Warehouse POs (which keep their own record). This endpoint
# reconstructs arrived inventory POs by grouping those usage rows by
# po_number so the frontend can merge them into the Arrived view.

# ---------------------------------------------------------------------------
# Freight -> PO ship-date linkage
# ---------------------------------------------------------------------------
# Lineage freight invoices carry the H&H PO number (plus order_number /
# shipper_ref) and the real ship_date. Arrived POs frequently have no
# ship_date — inventory POs lose it when they roll over, and some POs never
# had one entered — so we link by PO number to backfill the ship date for
# display in the Pending POs "Arrived" view.

def _norm_po_key(s: str) -> str:
    """Normalize a PO / reference token for cross-source matching.

    Uppercases, trims, drops a leading ``HHB-`` / ``HHB `` shipper prefix,
    and strips trailing punctuation. Leading zeros are preserved on
    purpose — Cheney PO numbers like ``014511...`` vs ``054511...`` differ
    only in those digits (the 2nd digit encodes the destination DC).
    """
    t = str(s or "").strip().upper()
    if t.startswith("HHB-") or t.startswith("HHB "):
        t = t[4:]
    return t.strip().strip(".").strip()


def _freight_ship_date_index() -> dict:
    """Map normalized PO/order/shipper-ref key -> earliest freight ship_date.

    Built from data/freight_invoices.json. When several invoices reference
    the same PO (split shipments), the earliest non-empty ship_date wins —
    that's when the order actually left the origin DC. Freight is auxiliary,
    so any load failure yields an empty index rather than an error.
    """
    try:
        from inventory_tracker import load_freight_invoices
        invoices = load_freight_invoices()
    except Exception:  # noqa: BLE001
        return {}
    idx: dict = {}
    for inv in invoices or []:
        sd = (inv.get("ship_date") or "").strip()
        if not sd:
            continue
        for fld in ("po_number", "order_number", "shipper_ref"):
            key = _norm_po_key(inv.get(fld) or "")
            if not key:
                continue
            prev = idx.get(key)
            if prev is None or sd < prev:
                idx[key] = sd
    return idx


@app.route("/api/arrived-pos")
def api_arrived_pos():
    """List arrived inventory-side POs reconstructed from the usage log.

    Groups usage rows with source=="on_order_rollover" by po_number and
    enriches each group with the distributor / warehouse pulled from the
    SKU's current inventory record. Chefs Warehouse POs are intentionally
    excluded here — they're served (arrived + active) by
    /api/chefs-warehouse/pos.

    Each group matches the shape the Pending POs frontend expects:
      po_number, po_revision, distributor, warehouse, ordered_at (best
      effort), eta, ship_date, arrival_date, total_cs, lines[], status.
    """
    import re as _re
    from inventory_tracker import load_inventory, load_usage

    inv = load_inventory()
    meta = {}
    for key, item in inv.items():
        meta[key] = {
            "distributor": item.get("distributor") or "",
            "warehouse":   item.get("warehouse") or "",
            "name":        item.get("name") or key,
        }

    groups: dict = {}
    for e in (load_usage() or []):
        if (e.get("source") or "") != "on_order_rollover":
            continue
        po = (e.get("po_number") or "").strip()
        if not po:
            continue
        m = meta.get(e.get("item_key") or "", {})
        g = groups.get(po)
        if g is None:
            g = groups[po] = {
                "po_number":    po,
                "po_revision":  e.get("po_revision") or "",
                "distributor":  m.get("distributor") or "",
                "warehouse":    m.get("warehouse") or "",
                "ordered_at":   "",
                "eta":          "",
                "ship_date":    "",
                "arrival_date": e.get("timestamp") or "",
                "total_cs":     0.0,
                "lines":        [],
                "status":       "arrived",
            }
        qty = abs(float(e.get("amount") or 0))
        g["total_cs"] += qty
        name = m.get("name") or e.get("item_name") or ""
        variety = name.split(" Bagel")[0] if " Bagel" in name else name
        g["lines"].append({
            "variety": variety,
            "name":    name,
            "qty":     qty,
            "unit":    e.get("unit") or "cs",
        })
        # arrival_date = the latest rollover timestamp across the PO's lines.
        ts = e.get("timestamp") or ""
        if ts > (g["arrival_date"] or ""):
            g["arrival_date"] = ts
        # First non-empty distributor / warehouse wins (lines may map to
        # SKUs that lost their metadata; keep the first useful one).
        if not g["distributor"] and m.get("distributor"):
            g["distributor"] = m["distributor"]
        if not g["warehouse"] and m.get("warehouse"):
            g["warehouse"] = m["warehouse"]
        # The rollover note carries "(ETA YYYY-MM-DD)" — surface it.
        if not g["eta"]:
            mm = _re.search(r"ETA (\d{4}-\d{2}-\d{2})", e.get("note") or "")
            if mm:
                g["eta"] = mm.group(1)

    out = list(groups.values())
    for g in out:
        g["total_cs"] = round(g["total_cs"], 2)

    # Backfill missing ship dates from the Lineage freight invoices,
    # linked by PO number. Display-only enrichment — nothing is persisted.
    freight_idx = _freight_ship_date_index()
    backfilled = 0
    for g in out:
        if (g.get("ship_date") or "").strip():
            continue
        sd = freight_idx.get(_norm_po_key(g.get("po_number") or ""))
        if sd:
            g["ship_date"] = sd
            g["ship_date_source"] = "freight"
            backfilled += 1

    # Newest arrivals first.
    out.sort(key=lambda x: (x.get("arrival_date") or "", x.get("po_number") or ""),
             reverse=True)
    return jsonify({"ok": True, "count": len(out),
                    "ship_date_backfilled": backfilled, "pos": out})


# ---------------------------------------------------------------------------
# API – Daily Production
# ---------------------------------------------------------------------------
# Production sheets are separate from inventory. Each record describes
# what was baked for a particular PO on a particular day, parsed from
# the Daily Production Sheet PDF that the production team emails out.

@app.route("/api/production")
def api_production_list():
    """List production records. Optional query params: distributor,
    warehouse, since (ISO date). Returns newest-first."""
    from inventory_tracker import load_production
    records = load_production()
    dist = (request.args.get("distributor") or "").strip()
    wh   = (request.args.get("warehouse") or "").strip()
    since = (request.args.get("since") or "").strip()
    out = []
    for r in records:
        if dist and (r.get("distributor") or "") != dist:
            continue
        if wh and (r.get("warehouse") or "") != wh:
            continue
        if since and (r.get("production_date") or "") < since:
            continue
        out.append(r)
    out.sort(key=lambda r: (r.get("production_date") or "", r.get("po_number") or ""),
             reverse=True)
    return jsonify(out)


@app.route("/api/production/lots-by-pair")
def api_production_lots_by_pair():
    """Lot-level breakdown grouped by (warehouse, variety) with FIFO state.

    Returns a dict keyed by ``"<warehouse>|<variety>"`` whose values are
    **oldest-first** lists of:

        {lot, cs, cs_produced, cs_consumed, cs_remaining,
         is_active, is_next_out,
         production_date, po_number, received_at}

    The FIFO rule (consume earliest lot first) is computed centrally in
    ``compute_lot_fifo_state``; this endpoint just enumerates the pairs and
    delegates so the front-end and any future server-side consumers see the
    same numbers. ``cs`` is preserved as an alias for ``cs_produced`` for
    backward compatibility with older JS revisions.
    """
    from inventory_tracker import (
        load_production, load_usage, load_inventory, compute_lot_fifo_state,
    )
    records = load_production()
    usage = load_usage()
    inv_snapshot = load_inventory()

    # Discover all distinct (warehouse, variety) pairs we have lots for.
    pairs: set = set()
    for r in records:
        wh = r.get("warehouse") or ""
        if not wh:
            continue
        for L in r.get("lines", []):
            lot = (L.get("lot_number") or "").strip()
            if not lot:
                continue
            variety = L.get("variety") or ""
            if not variety:
                continue
            pairs.add((wh, variety))

    out: dict = {}
    for (wh, variety) in pairs:
        lots = compute_lot_fifo_state(
            wh, variety,
            production_records=records,
            usage_records=usage,
            inventory_snapshot=inv_snapshot,
        )
        # Surface cs as an alias for cs_produced so any older client that
        # still reads L.cs keeps working until it is redeployed.
        for L in lots:
            L["cs"] = L["cs_produced"]
        out[f"{wh}|{variety}"] = lots
    return jsonify(out)


@app.route("/api/sales/ingest", methods=["POST"])
def api_sales_ingest():
    """Accept Toast product-mix rows pushed by an external aggregator.

    Body: {
      "rows": [
        {"restaurant_guid": "...", "location": "UES",
         "business_date": "2026-05-12",
         "item_guid": "...", "item": "Plain Bagel",
         "menu_group": "Bagels",          # optional
         "qty": 124, "gross": 372.00, "net": 369.50}
      ],
      "replace_dates": false   # if true, wipe existing rows for the
                               # (location, date) pairs in this batch
                               # before appending. Useful to refresh a
                               # day without orphaning items removed
                               # from the catalog.
    }

    Dedupe key: (restaurant_guid, business_date, item_guid). Posting the
    same item for the same day overwrites the prior row.
    """
    from inventory_tracker import load_sales, save_sales
    body = request.json or {}
    rows = body.get("rows") or []
    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "rows must be a list"}), 400
    replace_dates = bool(body.get("replace_dates", False))

    existing = load_sales()
    # Key existing rows.
    by_key = {(r.get("restaurant_guid"), r.get("business_date"), r.get("item_guid")): i
              for i, r in enumerate(existing)}

    if replace_dates:
        seed = set((r.get("restaurant_guid"), r.get("business_date")) for r in rows)
        existing = [r for r in existing
                    if (r.get("restaurant_guid"), r.get("business_date")) not in seed]
        by_key = {(r.get("restaurant_guid"), r.get("business_date"), r.get("item_guid")): i
                  for i, r in enumerate(existing)}

    added, updated = 0, 0
    for r in rows:
        if not r.get("restaurant_guid") or not r.get("business_date") or not r.get("item_guid"):
            continue
        key = (r.get("restaurant_guid"), r.get("business_date"), r.get("item_guid"))
        if key in by_key:
            existing[by_key[key]] = r
            updated += 1
        else:
            existing.append(r)
            by_key[key] = len(existing) - 1
            added += 1
    save_sales(existing)
    return jsonify({"ok": True, "added": added, "updated": updated,
                    "total_rows": len(existing)})


@app.route("/api/report/toast-sales")
def api_report_toast_sales():
    """Top-selling items aggregated by week or month, optionally by
    location. Output is the data behind the Report -> Top Consumed
    section.

    Query params:
      period        'week' | 'month' (default 'week')
      buckets       int, default 8 (weeks) / 6 (months)
      location      restaurant_guid OR location string, optional
      top_n         int, default 10
    """
    from inventory_tracker import load_sales
    from datetime import datetime as _dt, timedelta as _td

    args = request.args
    period = (args.get("period") or "week").lower()
    if period not in ("week", "month"):
        return jsonify({"ok": False, "error": "period must be week|month"}), 400
    try:
        buckets = int(args.get("buckets") or (8 if period == "week" else 6))
    except (TypeError, ValueError):
        buckets = 8 if period == "week" else 6
    try:
        top_n = int(args.get("top_n") or 10)
    except (TypeError, ValueError):
        top_n = 10
    location_q = (args.get("location") or "").strip()
    # Anchor date — the Report page's Prev/Next/Today buttons and the
    # date picker all set this to a single YYYY-MM-DD. When present we
    # return the SINGLE week/month bucket containing that date so the
    # user sees the period they picked (matching the Restock card's
    # one-window-at-a-time semantics). When absent we fall back to
    # returning the most recent N buckets that have data.
    end_date_q = (args.get("end_date") or "").strip()
    # When True (default) a cache miss for the picked bucket triggers a
    # live Toast pull. The client can pass live=0 to force cache-only
    # behavior (e.g. for fast retries while a fetch is in flight).
    live_q = (args.get("live") or "1").strip() not in ("0", "false", "no")

    # Bucket each row.
    def _bucket_key(date_iso: str) -> str:
        try:
            d = _dt.strptime(date_iso, "%Y-%m-%d")
        except ValueError:
            return ""
        if period == "week":
            # ISO week, Monday-start. Anchor on the Monday of that week.
            mon = d - _td(days=d.weekday())
            return mon.strftime("%Y-%m-%d")
        return d.strftime("%Y-%m")

    def _bucket_range(anchor_iso: str):
        """Return (start_date_iso, end_date_iso) inclusive for the
        bucket containing anchor_iso. Caps end at today — we never
        ask Toast for future dates."""
        ed = _dt.strptime(anchor_iso, "%Y-%m-%d")
        if period == "week":
            start = ed - _td(days=ed.weekday())
            end   = start + _td(days=6)
        else:
            start = ed.replace(day=1)
            if ed.month == 12:
                end = ed.replace(day=31)
            else:
                end = ed.replace(month=ed.month + 1, day=1) - _td(days=1)
        today = _dt.now().date()
        if end.date() > today:
            end = _dt.combine(today, _dt.min.time())
        return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")

    def _target_locations(loc_q: str):
        """Return [(restaurant_guid, location_name), ...] for the live
        fetch. If a specific location is selected, returns just that
        one. Otherwise returns every active retail location."""
        from inventory_tracker import TOAST_RETAIL_LOCATIONS
        if loc_q:
            if "-" in loc_q or len(loc_q) == 36:
                for L in TOAST_RETAIL_LOCATIONS:
                    if L["restaurant_guid"] == loc_q:
                        return [(L["restaurant_guid"], L["location"])]
                return [(loc_q, "")]
            q = loc_q.lower()
            return [(L["restaurant_guid"], L["location"])
                    for L in TOAST_RETAIL_LOCATIONS
                    if q in (L.get("location") or "").lower()]
        return [(L["restaurant_guid"], L["location"])
                for L in TOAST_RETAIL_LOCATIONS
                if (L.get("status") or "active") == "active"]

    def _ensure_bucket_cached(all_rows: list, anchor_iso: str, loc_q: str):
        """If the bucket containing anchor_iso has no rows for the
        target locations, pull from Toast and append to sales.json.
        Returns (updated_rows, meta dict) where meta carries fetch info
        for the client (rows_fetched, fetch_error, fetched_at)."""
        meta = {"rows_fetched": 0, "fetch_error": None, "fetched_at": None}
        if not live_q:
            return all_rows, meta
        try:
            start_iso, end_iso = _bucket_range(anchor_iso)
        except ValueError:
            return all_rows, meta
        targets = _target_locations(loc_q)
        if not targets:
            return all_rows, meta
        # Which (guid, date) pairs already have at least 1 row?
        cached_keys = set()
        for r in all_rows:
            d = (r.get("business_date") or "").strip()
            if start_iso <= d <= end_iso:
                cached_keys.add((r.get("restaurant_guid") or "", d))
        # Enumerate every (target_location, date) pair in the bucket
        # range; anything not in cached_keys is a fetch candidate.
        missing = []
        s = _dt.strptime(start_iso, "%Y-%m-%d")
        e = _dt.strptime(end_iso,   "%Y-%m-%d")
        d_cursor = s
        while d_cursor <= e:
            d_iso = d_cursor.strftime("%Y-%m-%d")
            for (guid, name) in targets:
                if (guid, d_iso) not in cached_keys:
                    missing.append((guid, d_iso, name))
            d_cursor += _td(days=1)
        if not missing:
            return all_rows, meta
        try:
            from integrations import toast_api
            if not toast_api.is_configured():
                meta["fetch_error"] = "toast_not_configured"
                return all_rows, meta
            new_rows = toast_api.fetch_product_mix_batch(
                missing, max_workers=6, timeout_s=28.0)
        except Exception as ex:
            meta["fetch_error"] = f"{type(ex).__name__}: {ex}"
            return all_rows, meta
        if not new_rows:
            # No orders for any of the missing (guid, date) pairs is a
            # legitimate response — still record the attempt so the UI
            # knows a fetch ran.
            meta["fetched_at"] = _dt.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
            return all_rows, meta
        from inventory_tracker import save_sales
        # Dedupe on (restaurant_guid, business_date, item_guid) — if a
        # date was partially cached, prefer the fresh row.
        by_key: dict = {}
        for r in all_rows:
            k = (r.get("restaurant_guid"), r.get("business_date"),
                 r.get("item_guid"))
            by_key[k] = r
        for r in new_rows:
            k = (r.get("restaurant_guid"), r.get("business_date"),
                 r.get("item_guid"))
            by_key[k] = r
        merged = list(by_key.values())
        save_sales(merged)
        meta["rows_fetched"] = len(new_rows)
        meta["fetched_at"]   = _dt.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        return merged, meta

    # Load the full row set ONCE. If an anchor date is set we may
    # mutate this with a Toast fetch before bucketing.
    rows = load_sales() or []
    fetch_meta = {"rows_fetched": 0, "fetch_error": None, "fetched_at": None}
    if end_date_q:
        try:
            rows, fetch_meta = _ensure_bucket_cached(rows, end_date_q,
                                                    location_q)
        except Exception as ex:
            fetch_meta["fetch_error"] = f"{type(ex).__name__}: {ex}"

    # Apply the location filter (after the live fetch so newly-pulled
    # rows are visible).
    if location_q:
        if "-" in location_q or len(location_q) == 36:
            rows = [r for r in rows if r.get("restaurant_guid") == location_q]
        else:
            q = location_q.lower()
            rows = [r for r in rows if q in (r.get("location") or "").lower()]

    if not rows and not end_date_q:
        return jsonify({"ok": True, "period": period, "buckets": [],
                        "total_rows": 0, "fetch": fetch_meta})

    # Aggregate by NORMALIZED display name (+ menu group) rather than
    # item_guid. Toast issues a fresh item_guid every time a menu item
    # is re-published (price change, name tweak, modifier shuffle), so
    # the same user-visible product can appear under several guids in
    # a single week. Keying on guid surfaces these as duplicate rows;
    # keying on name collapses them. The underlying sales.json still
    # keeps the raw guid for traceability.
    def _norm(s: str) -> str:
        return " ".join((s or "").lower().split())

    buckets_map: dict = {}
    for r in rows:
        bk = _bucket_key(r.get("business_date") or "")
        if not bk:
            continue
        slot = buckets_map.setdefault(bk, {})
        name      = (r.get("item") or "").strip()
        menu_grp  = (r.get("menu_group") or "").strip()
        if not name and not r.get("item_guid"):
            continue
        item_key  = (_norm(name), _norm(menu_grp))
        agg = slot.setdefault(item_key, {
            "item":       name,
            "menu_group": menu_grp,
            "qty":        0,
            "gross":      0.0,
            "net":        0.0,
            "guid_count": 0,
        })
        agg["qty"]        += int(r.get("qty") or 0)
        agg["gross"]      += float(r.get("gross") or 0)
        agg["net"]        += float(r.get("net") or 0)
        agg["guid_count"] += 1  # informational; not used in render
        # Prefer the longest/non-empty observed values for display.
        if name and len(name) > len(agg["item"]):
            agg["item"] = name
        if menu_grp and not agg["menu_group"]:
            agg["menu_group"] = menu_grp

    def _label_for(bk: str) -> str:
        if period == "week":
            d = _dt.strptime(bk, "%Y-%m-%d")
            return f"{d.strftime('%b %-d')} \u2013 {(d + _td(days=6)).strftime('%b %-d, %Y')}"
        d = _dt.strptime(bk, "%Y-%m")
        return d.strftime("%B %Y")

    def _serialize(bk: str) -> dict:
        items = list(buckets_map.get(bk, {}).values())
        items.sort(key=lambda x: x["gross"], reverse=True)
        total = sum(x["gross"] for x in items) or 1
        for it in items:
            it["mix_pct"] = round(100 * it["gross"] / total, 2)
            it["gross"]   = round(it["gross"], 2)
            it["net"]     = round(it["net"], 2)
        return {
            "key":         bk,
            "label":       _label_for(bk),
            "total_gross": round(total, 2) if items else 0.0,
            "items":       items[:top_n],
            "item_count":  len(items),
        }

    # When a specific date is picked, return ONLY the bucket containing
    # that date \u2014 even if it has no rows, so the UI can render a clear
    # "no data for week of X" state instead of falling silently empty.
    if end_date_q:
        try:
            ed = _dt.strptime(end_date_q, "%Y-%m-%d")
            if period == "week":
                anchor = (ed - _td(days=ed.weekday())).strftime("%Y-%m-%d")
            else:
                anchor = ed.strftime("%Y-%m")
            return jsonify({"ok": True, "period": period,
                            "buckets":     [_serialize(anchor)],
                            "anchor":      anchor,
                            "anchor_date": end_date_q,
                            "total_rows":  len(rows),
                            "fetch":       fetch_meta})
        except ValueError:
            pass

    # No anchor date \u2014 fall back to the most recent N buckets with data.
    bucket_keys = sorted(buckets_map.keys(), reverse=True)
    out_buckets = [_serialize(bk) for bk in bucket_keys[:buckets]]

    return jsonify({"ok": True, "period": period,
                    "buckets": out_buckets,
                    "total_rows": len(rows)})


@app.route("/api/sales/locations")
def api_sales_locations():
    """List every retail Toast location in the registry, with row
    counts and date ranges from the sales store. Locations with no
    ingested data return rows=0 — they still show in the dropdown so
    the user can pick any location for comparison.
    """
    from inventory_tracker import load_sales, TOAST_RETAIL_LOCATIONS
    rows = load_sales() or []
    # Start with the registry — every retail location guaranteed, with
    # state abbreviation so the UI can render "Penn Station, NY" without
    # any per-location lookup.
    by_loc: dict = {}
    for L in TOAST_RETAIL_LOCATIONS:
        by_loc[L["restaurant_guid"]] = {
            "restaurant_guid": L["restaurant_guid"],
            "location":        L["location"],
            "state":           L.get("state") or "",
            "status":          L.get("status") or "active",
            "rows":            0,
            "min_date":        "",
            "max_date":        "",
        }
    # Layer in any ingested data — increment counts + date range.
    for r in rows:
        guid = r.get("restaurant_guid") or ""
        if not guid:
            continue
        slot = by_loc.setdefault(guid, {
            "restaurant_guid": guid,
            "location":        r.get("location") or "",
            "state":           "",
            "rows":            0,
            "min_date":        r.get("business_date") or "",
            "max_date":        r.get("business_date") or "",
        })
        slot["rows"] += 1
        d = r.get("business_date") or ""
        if d:
            if not slot["min_date"] or d < slot["min_date"]:
                slot["min_date"] = d
            if not slot["max_date"] or d > slot["max_date"]:
                slot["max_date"] = d
    return jsonify({"ok": True,
                    "locations": sorted(by_loc.values(),
                                        key=lambda x: x["location"])})


@app.route("/api/traceability/search")
def api_traceability_search():
    """Cross-source traceability search.

    Query params:
      ``from``        ISO date — only include records on/after this date.
      ``to``          ISO date — only include records on/before this date.
      ``lot``         Substring match on lot_number (case-insensitive).
      ``mfg_code``    4-digit item code (matches the first 4 digits of
                      any lot OR the variety via HH_MFG_CODE_TO_VARIETY).
      ``warehouse``   Exact warehouse name.
      ``distributor`` Exact distributor name.

    Returns a list of matching production records, each with these
    downstream fields populated for the requesting line:
      ``po_status``      pending | arrived | canceled | unknown
      ``ship_date``      from the on_order entry, if pending.
      ``arrival_date``   from the on_order entry, if pending.
      ``eta``            from the on_order entry, if pending.
      ``lines[].on_hand_now``  current cs on hand for that variety at
                                that warehouse (live inventory).
      ``lines[].usage_total_cs`` total cs moved on usage log for this
                                PO + variety (from usage.json).
      ``lines[].usage_event_count``  count of usage events.
    """
    from inventory_tracker import (
        load_production, load_inventory, load_usage,
        load_canceled_pos,
    )
    try:
        from integrations.hh_mfg_codes import HH_MFG_CODE_TO_VARIETY
    except Exception:
        HH_MFG_CODE_TO_VARIETY = {}

    args = request.args
    q_from = (args.get("from") or "").strip()
    q_to   = (args.get("to") or "").strip()
    q_lot  = (args.get("lot") or "").strip().lower()
    q_code = (args.get("mfg_code") or "").strip()
    q_wh   = (args.get("warehouse") or "").strip()
    q_dist = (args.get("distributor") or "").strip()

    records = load_production()

    # Apply filters.
    out = []
    code_variety = HH_MFG_CODE_TO_VARIETY.get(q_code) if q_code else None
    for r in records:
        if q_from and (r.get("production_date") or "") < q_from:
            continue
        if q_to and (r.get("production_date") or "") > q_to:
            continue
        if q_wh and (r.get("warehouse") or "") != q_wh:
            continue
        if q_dist and (r.get("distributor") or "") != q_dist:
            continue
        # Lot / mfg-code filter only keeps the record if at least one
        # line matches. We do NOT prune the other lines from the record
        # — operators usually want to see the full sheet for context.
        if q_lot or q_code:
            matched_any = False
            for L in r.get("lines", []):
                lot = (L.get("lot_number") or "")
                if q_lot and q_lot in lot.lower():
                    matched_any = True
                    break
                if q_code:
                    # Match either the first 4 digits of the lot, or the
                    # variety mapped from the code.
                    if lot.startswith(q_code):
                        matched_any = True
                        break
                    if code_variety and (L.get("variety") or "") == code_variety:
                        matched_any = True
                        break
            if not matched_any:
                continue
        out.append(r)

    # Build downstream lookups.
    inv = load_inventory()
    # {(warehouse, variety) -> on_hand_qty} for currently-on-hand snapshot
    on_hand = {}
    for item in inv.values():
        wh = item.get("warehouse") or ""
        name = item.get("name") or ""
        variety = name.split(" Bagel")[0] if " Bagel" in name else name
        on_hand[(wh, variety)] = (on_hand.get((wh, variety), 0)
                                  + float(item.get("quantity") or 0))

    # Pending-PO lookup keyed by po_number — first non-empty wins.
    # Inventory on_order is checked first (USF + Cheney); then Chefs
    # Warehouse POs (which live in their own file because they don't
    # touch inventory). A Daily Production sheet stamped with a CW PO#
    # therefore still resolves to "pending" with the right ship/arrival
    # data.
    po_pending = {}
    for item in inv.values():
        for e in item.get("on_order") or []:
            po = (e.get("po_number") or "").strip()
            if not po:
                continue
            slot = po_pending.setdefault(po, {})
            for k in ("ship_date", "arrival_date", "eta"):
                if not slot.get(k) and e.get(k):
                    slot[k] = e[k]
    try:
        from inventory_tracker import load_chefs_warehouse_pos
        for cw in load_chefs_warehouse_pos():
            po = (cw.get("po_number") or "").strip()
            if not po or cw.get("canceled"):
                continue
            slot = po_pending.setdefault(po, {})
            for k in ("ship_date", "arrival_date", "eta"):
                if not slot.get(k) and cw.get(k):
                    slot[k] = cw[k]
    except Exception:  # noqa: BLE001
        # CW POs are auxiliary; failure here must not break traceability.
        pass

    try:
        canceled = load_canceled_pos()
    except Exception:
        canceled = {}

    # Usage rollup keyed by (po_number, variety).
    usage_log = load_usage() or []
    usage_roll: dict = {}  # {(po, variety): {cs, n}}
    for e in usage_log:
        po = (e.get("po_number") or "").strip()
        if not po:
            continue
        name = e.get("item_name") or ""
        variety = name.split(" Bagel")[0] if " Bagel" in name else name
        slot = usage_roll.setdefault((po, variety), {"cs": 0.0, "n": 0})
        slot["cs"] += abs(float(e.get("amount") or 0))
        slot["n"] += 1

    # Enrich each record.
    enriched = []
    for r in out:
        rcopy = dict(r)
        po = (r.get("po_number") or "").strip()
        wh = r.get("warehouse") or ""
        pend = po_pending.get(po) if po else None
        if po and po in canceled:
            rcopy["po_status"] = "canceled"
        elif pend:
            rcopy["po_status"]    = "pending"
            rcopy["ship_date"]    = pend.get("ship_date") or ""
            rcopy["arrival_date"] = pend.get("arrival_date") or ""
            rcopy["eta"]          = pend.get("eta") or ""
        elif po:
            # PO existed but isn't in on_order anymore -> rolled over.
            rcopy["po_status"] = "arrived"
        else:
            rcopy["po_status"] = "unknown"

        enriched_lines = []
        for L in r.get("lines") or []:
            v = L.get("variety") or ""
            lcopy = dict(L)
            lcopy["on_hand_now"] = round(on_hand.get((wh, v), 0), 2)
            ur = usage_roll.get((po, v)) if po else None
            lcopy["usage_total_cs"]    = round(ur["cs"], 2) if ur else 0
            lcopy["usage_event_count"] = ur["n"] if ur else 0
            enriched_lines.append(lcopy)
        rcopy["lines"] = enriched_lines
        enriched.append(rcopy)

    # Sort newest-first by production date for stable grouping.
    enriched.sort(key=lambda r: (r.get("production_date") or "",
                                  r.get("po_number") or ""),
                  reverse=True)
    return jsonify({
        "ok": True,
        "filters": {
            "from": q_from, "to": q_to, "lot": q_lot, "mfg_code": q_code,
            "warehouse": q_wh, "distributor": q_dist,
        },
        "count": len(enriched),
        "records": enriched,
        "mfg_code_map": HH_MFG_CODE_TO_VARIETY,
    })


@app.route("/api/production/summary")
def api_production_summary():
    """Roll-up across production records.

    Query params:
      period  = "day" | "week" | "month"  (default "week")

    Returns:
      {
        "period": "week",
        "buckets": [
          { "key": "2026-W19", "start": "2026-05-04", "end": "2026-05-10",
            "total_cs": 1232,
            "by_distributor": {"Cheney Brothers": 224, "US Foods": 1008, ...},
            "by_variety": {"Plain": 176, "Everything": 184, ...}
          }, ...
        ]
      }
    """
    from datetime import datetime, timedelta
    from inventory_tracker import load_production
    records = load_production()
    period = (request.args.get("period") or "week").lower()

    def bucket_key(iso_date: str) -> tuple:
        try:
            dt = datetime.fromisoformat(iso_date)
        except (TypeError, ValueError):
            return ("", "", "")
        if period == "day":
            return (iso_date, iso_date, iso_date)
        if period == "month":
            start = dt.replace(day=1)
            # Last day of month: jump to next month then back one day
            if start.month == 12:
                nxt = start.replace(year=start.year + 1, month=1)
            else:
                nxt = start.replace(month=start.month + 1)
            end = nxt - timedelta(days=1)
            return (f"{start.year:04d}-{start.month:02d}",
                    start.date().isoformat(), end.date().isoformat())
        # default "week" — ISO weeks (Monday-start)
        iso = dt.isocalendar()
        start = dt - timedelta(days=dt.weekday())
        end = start + timedelta(days=6)
        return (f"{iso.year:04d}-W{iso.week:02d}",
                start.date().isoformat(), end.date().isoformat())

    buckets: dict = {}
    for r in records:
        key, start, end = bucket_key(r.get("production_date") or "")
        if not key:
            continue
        b = buckets.setdefault(key, {
            "key": key, "start": start, "end": end,
            "total_cs": 0, "by_distributor": {}, "by_variety": {},
        })
        b["total_cs"] += int(r.get("total_cases") or 0)
        d = r.get("distributor") or "Unassigned"
        b["by_distributor"][d] = b["by_distributor"].get(d, 0) + int(r.get("total_cases") or 0)
        for line in (r.get("lines") or []):
            v = line.get("variety") or "Unknown"
            b["by_variety"][v] = b["by_variety"].get(v, 0) + int(line.get("cs_count") or 0)
    return jsonify({
        "period": period,
        "buckets": sorted(buckets.values(), key=lambda x: x["start"], reverse=True),
    })


@app.route("/api/production/ingest", methods=["POST"])
def api_production_ingest():
    """Parse a Daily Production Sheet PDF and store as a record.

    Body:
      pdf_b64       base64-encoded PDF bytes (required)
      subject       email subject (optional, used as a fallback PO source)
      sender        email From address (optional, kept for audit)
      message_id    Graph internetMessageId or RFC822 Message-Id (required
                    for idempotency — re-posting the same message is a no-op)
      received_at   ISO timestamp (optional)

    Returns 200 with {ok, status: "ingested"|"duplicate"|"parse_error",
    record, error}. parse_error is surfaced for image-only scan PDFs so
    the operator can re-request a text PDF.
    """
    import base64
    from datetime import datetime
    from inventory_tracker import load_production, save_production
    from integrations.production_pdf_parser import parse_production_pdf

    body = request.json or {}
    pdf_b64 = body.get("pdf_b64") or ""
    if not pdf_b64:
        return jsonify({"ok": False, "error": "pdf_b64 required"}), 400
    subject    = (body.get("subject") or "").strip()
    sender     = (body.get("sender") or "").strip()
    message_id = (body.get("message_id") or "").strip()
    received_at = (body.get("received_at") or "").strip()

    records = load_production()
    if message_id:
        for existing in records:
            if existing.get("source_message_id") == message_id:
                return jsonify({
                    "ok": True,
                    "status": "duplicate",
                    "record": existing,
                })

    try:
        pdf_bytes = base64.b64decode(pdf_b64)
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"bad pdf_b64: {exc}"}), 400

    sheet = parse_production_pdf(pdf_bytes, subject=subject)

    # Content-signature dedup. Same email can land twice with two different
    # IDs (RFC <...@outlook.com> Message-ID vs. Microsoft Graph
    # internetMessageId) when more than one scanner path picks it up — see
    # the 2026-05-19 incident where five Daily Production records got stored
    # twice. The message_id check above only catches *exact* re-sends; this
    # block catches the cross-ID dup by matching on (warehouse, po_number,
    # production_date) + identical line signature. Legitimate amendments
    # (different lines/cs counts) still pass through.
    if (sheet.po_number and sheet.warehouse and sheet.production_date
            and sheet.lines):
        incoming_sig = tuple(sorted(
            (L.variety or "", int(L.cs_count or 0), L.lot_number or "")
            for L in sheet.lines))
        for existing in records:
            if (existing.get("warehouse") != sheet.warehouse
                    or existing.get("po_number") != sheet.po_number
                    or existing.get("production_date") != sheet.production_date):
                continue
            existing_sig = tuple(sorted(
                (L.get("variety") or "", int(L.get("cs_count") or 0),
                 L.get("lot_number") or "")
                for L in (existing.get("lines") or [])))
            if existing_sig == incoming_sig:
                return jsonify({
                    "ok": True,
                    "status": "duplicate",
                    "record": existing,
                    "dedup_reason": "content_signature_match",
                })

    if sheet.error and not sheet.lines:
        # Persist a stub for the dashboard so the operator can see the
        # email arrived but needs manual entry, then surface the error.
        stub = {
            "production_date": "",
            "warehouse": "",
            "warehouse_raw": "",
            "distributor": "",
            "po_number": "",
            "lines": [],
            "total_cases": 0,
            "unmapped_varieties": [],
            "source_message_id": message_id,
            "source_subject": subject,
            "source_sender": sender,
            "received_at": received_at or datetime.now().isoformat(),
            "ingested_at": datetime.now().isoformat(),
            "parse_error": sheet.error,
        }
        records.append(stub)
        save_production(records)
        return jsonify({
            "ok": True,
            "status": "parse_error",
            "record": stub,
            "error": sheet.error,
        })

    record = {
        "production_date":    sheet.production_date,
        "warehouse":          sheet.warehouse,
        "warehouse_raw":      sheet.warehouse_raw,
        "distributor":        sheet.distributor,
        "po_number":          sheet.po_number,
        "lines": [
            {"variety": L.variety, "raw_variety": L.raw_variety,
             "cs_count": L.cs_count, "lot_number": L.lot_number}
            for L in sheet.lines
        ],
        "total_cases":        sheet.total_cases,
        "unmapped_varieties": sheet.unmapped_varieties,
        "source_message_id":  message_id,
        "source_subject":     subject,
        "source_sender":      sender,
        "received_at":        received_at or datetime.now().isoformat(),
        "ingested_at":        datetime.now().isoformat(),
        "parse_error":        "",
    }
    records.append(record)
    save_production(records)
    return jsonify({"ok": True, "status": "ingested", "record": record})


@app.route("/api/production/scan", methods=["POST"])
def api_production_scan():
    """Wide-lookback scan of mailbox(es) for Daily Production sheet emails.

    Mirrors the existing /api/email/scan flow but routes recognized
    Daily Production emails (sender = *@hhbagels.com, subject contains
    "Daily Production") through the production PDF parser into
    data/production.json instead of the on_order pipeline.

    Body:
      lookback_days  int   how far back to look (defaults to 365)
      max_messages   int   per-mailbox cap on qualified messages
                           (defaults to 200, hard cap 2000)
      dry_run        bool  parse but don't persist

    Returns {ok, scanned, ingested, parse_errors, records:[brief]}.
    """
    import base64
    import traceback as _tb
    from datetime import datetime, timezone, timedelta
    try:
        from integrations.email_scanner import (
            EmailInboxClient, _distributor_from_sender, GRAPH_BASE,
        )
        from integrations.production_pdf_parser import parse_production_pdf
        from inventory_tracker import load_production, save_production
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"import failed: {exc}",
                        "traceback": _tb.format_exc()[-1500:]}), 500

    body = request.json or {}
    try:
        lookback_days = int(body.get("lookback_days") or 365)
    except (TypeError, ValueError):
        lookback_days = 365
    try:
        until_days = int(body.get("until_days") or 0)
    except (TypeError, ValueError):
        until_days = 0
    try:
        max_messages = int(body.get("max_messages") or 200)
    except (TypeError, ValueError):
        max_messages = 200
    max_messages = max(1, min(max_messages, 2000))
    dry_run = bool(body.get("dry_run", False))
    refresh = bool(body.get("refresh", False))

    since = datetime.now(timezone.utc) - timedelta(days=lookback_days)
    since_iso = since.replace(microsecond=0).isoformat().replace("+00:00", "Z")
    until_iso = None
    if until_days > 0:
        until = datetime.now(timezone.utc) - timedelta(days=until_days)
        until_iso = until.replace(microsecond=0).isoformat().replace("+00:00", "Z")

    client = EmailInboxClient()
    try:
        token = client._ms365_token()
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"ms365 token failed: {exc}"}), 500

    import os, json as _json, urllib.parse, urllib.request, urllib.error
    users = [u.strip() for u in os.environ.get("MS365_USER", "").split(",") if u.strip()]
    folder = urllib.parse.quote(os.environ.get("MS365_FOLDER", "Inbox"))

    scanned = 0
    qualifying_count = 0
    ingested = 0
    parse_errors = []
    brief_records = []
    existing = load_production()
    seen_msg_ids = {r.get("source_message_id") for r in existing if r.get("source_message_id")}
    touched_msg_ids = set()  # records modified or added in THIS run

    for upn in users:
        user = urllib.parse.quote(upn)
        # Build a date filter; narrow to a window if `until_days` was given.
        date_filter = f"receivedDateTime ge {since_iso}"
        if until_iso:
            date_filter += f" and receivedDateTime lt {until_iso}"
        qs = urllib.parse.urlencode({
            "$top": "100",
            "$select": ("id,subject,from,toRecipients,receivedDateTime,"
                        "hasAttachments,internetMessageId"),
            "$filter": f"hasAttachments eq true and {date_filter}",
        })
        next_url = f"{GRAPH_BASE}/users/{user}/mailFolders/{folder}/messages?{qs}"
        pages = 0
        per_mailbox = 0
        while next_url and pages < 40 and per_mailbox < max_messages:
            pages += 1
            try:
                req = urllib.request.Request(next_url, method="GET")
                req.add_header("Authorization", f"Bearer {token}")
                req.add_header("Accept", "application/json")
                req.add_header("ConsistencyLevel", "eventual")
                with urllib.request.urlopen(req, timeout=30) as resp:
                    page = _json.loads(resp.read().decode("utf-8"))
            except Exception as exc:  # noqa: BLE001
                parse_errors.append(f"{upn} page {pages}: {exc}")
                break
            for m in page.get("value", []):
                scanned += 1
                if per_mailbox >= max_messages:
                    break
                subject = m.get("subject") or ""
                sender = (((m.get("from") or {}).get("emailAddress") or {})
                          .get("address") or "")
                # Production emails: sender on hhbagels.com + subject says
                # "Daily Production" (case-flex)
                dom = sender.split("@")[-1].lower() if "@" in sender else ""
                if not (dom == "hhbagels.com" or dom.endswith(".hhbagels.com")):
                    continue
                if "daily production" not in subject.lower():
                    continue
                qualifying_count += 1
                msg_id = m.get("internetMessageId") or m.get("id") or ""
                if msg_id and msg_id in seen_msg_ids and not refresh:
                    continue  # already ingested
                # Pull the attachments list
                att_url = f"{GRAPH_BASE}/users/{user}/messages/{m.get('id')}/attachments"
                try:
                    araw, _ = client._graph_get(att_url, token)
                    apage = _json.loads(araw.decode("utf-8"))
                except Exception as exc:  # noqa: BLE001
                    parse_errors.append(f"{subject[:60]!r}: list-att failed: {exc}")
                    continue
                pdf_atts = [a for a in apage.get("value", [])
                            if (a.get("name") or "").lower().endswith(".pdf")
                            or (a.get("contentType") or "").lower() == "application/pdf"]
                if not pdf_atts:
                    continue
                # Fetch + parse the first PDF attachment
                a = pdf_atts[0]
                acid = a.get("id")
                fetch_url = f"{GRAPH_BASE}/users/{user}/messages/{m.get('id')}/attachments/{acid}"
                try:
                    fraw, _ = client._graph_get(fetch_url, token)
                    apayload = _json.loads(fraw.decode("utf-8"))
                    pdf_bytes = base64.b64decode(apayload.get("contentBytes") or "")
                except Exception as exc:  # noqa: BLE001
                    parse_errors.append(f"{subject[:60]!r}: fetch-att failed: {exc}")
                    continue
                sheet = parse_production_pdf(pdf_bytes, subject=subject)
                if sheet.error and not sheet.lines:
                    parse_errors.append(f"{subject[:60]!r}: {sheet.error}")
                    if dry_run:
                        per_mailbox += 1
                        continue
                    # Persist a stub so the operator sees it in the UI
                    record = {
                        "production_date": "", "warehouse": "", "warehouse_raw": "",
                        "distributor": "", "po_number": "", "lines": [],
                        "total_cases": 0, "unmapped_varieties": [],
                        "source_message_id": msg_id,
                        "source_subject": subject, "source_sender": sender,
                        "received_at": m.get("receivedDateTime") or "",
                        "ingested_at": datetime.now().isoformat(),
                        "parse_error": sheet.error,
                    }
                    existing.append(record)
                    seen_msg_ids.add(msg_id)
                    if msg_id:
                        touched_msg_ids.add(msg_id)
                    brief_records.append({"subject": subject, "po_number": "",
                                          "warehouse": "", "total_cases": 0,
                                          "parse_error": sheet.error})
                    per_mailbox += 1
                    continue
                if dry_run:
                    per_mailbox += 1
                    continue
                record = {
                    "production_date":    sheet.production_date,
                    "warehouse":          sheet.warehouse,
                    "warehouse_raw":      sheet.warehouse_raw,
                    "distributor":        sheet.distributor,
                    "po_number":          sheet.po_number,
                    "lines": [
                        {"variety": L.variety, "raw_variety": L.raw_variety,
                         "cs_count": L.cs_count, "lot_number": L.lot_number}
                        for L in sheet.lines
                    ],
                    "total_cases":        sheet.total_cases,
                    "unmapped_varieties": sheet.unmapped_varieties,
                    "source_message_id":  msg_id,
                    "source_subject":     subject,
                    "source_sender":      sender,
                    "received_at":        m.get("receivedDateTime") or "",
                    "ingested_at":        datetime.now().isoformat(),
                    "parse_error":        "",
                }
                if refresh and msg_id and msg_id in seen_msg_ids:
                    # Update in place rather than append a duplicate.
                    for i, r in enumerate(existing):
                        if r.get("source_message_id") == msg_id:
                            # Preserve original received_at if already set;
                            # overwrite parsed-from-PDF fields with the
                            # fresh values.
                            record["received_at"] = r.get("received_at") or record["received_at"]
                            existing[i] = record
                            break
                else:
                    existing.append(record)
                    seen_msg_ids.add(msg_id)
                if msg_id:
                    touched_msg_ids.add(msg_id)
                ingested += 1
                brief_records.append({
                    "subject": subject, "po_number": sheet.po_number,
                    "warehouse": sheet.warehouse, "production_date": sheet.production_date,
                    "total_cases": sheet.total_cases,
                })
                per_mailbox += 1
            next_url = page.get("@odata.nextLink")

    if not dry_run:
        # Re-load from disk and merge by source_message_id to avoid a
        # read-modify-write race when two scan calls overlap (a
        # symptom: the second call would clobber the first's appended
        # rows). The in-memory `existing` was loaded at the start of
        # THIS call and may now be stale.
        from inventory_tracker import load_production
        current = load_production()
        by_msg = {r.get("source_message_id"): r for r in current
                  if r.get("source_message_id")}
        # Records this call touched (added OR refreshed in place) live in
        # `existing` and are keyed by source_message_id in `touched_msg_ids`.
        # For msg_ids we touched, our in-memory version wins over the disk
        # version (e.g. refresh=true rewrote the lines with lot codes).
        # For msg_ids we did NOT touch but are present in BOTH `current` and
        # `existing`, the disk version wins (it may have been refreshed by a
        # concurrent scan). For msg_ids only in `existing`, those are new
        # additions — keep them.
        existing_by_msg = {r.get("source_message_id"): r for r in existing
                           if r.get("source_message_id")}
        for mid in touched_msg_ids:
            if mid in existing_by_msg:
                by_msg[mid] = existing_by_msg[mid]
        for r in existing:
            mid = r.get("source_message_id")
            if mid and mid not in by_msg:
                by_msg[mid] = r
            elif not mid:
                current.append(r)
        merged = [r for r in current if not r.get("source_message_id")] \
                 + list(by_msg.values())
        save_production(merged)

    return jsonify({
        "ok": True, "dry_run": dry_run,
        "lookback_days": lookback_days, "mailboxes": users,
        "messages_scanned": scanned,
        "messages_qualifying": qualifying_count,
        "ingested": ingested,
        "parse_errors": parse_errors,
        "records": brief_records,
    })


@app.route("/api/admin/production/renormalize-varieties", methods=["POST"])
def api_admin_production_renormalize_varieties():
    """Re-run _normalize_variety on every stored production line.

    Useful after extending _VARIETY_ALIASES or _VARIETY_SHORTHAND so
    historical records (which kept their original best-effort variety
    label) get aggregated into the canonical buckets. Lines whose
    canonical changes also have their lot_number re-paired against
    the updated variety so the Distributors / Traceability tabs stay
    consistent.

    Body: { dry_run: bool }
    """
    from inventory_tracker import load_production, save_production
    from integrations.production_pdf_parser import _normalize_variety
    body = request.json or {}
    dry_run = bool(body.get("dry_run", False))
    records = load_production()
    changed_lines = 0
    changed_records = 0
    samples = []
    for r in records:
        any_changed = False
        for L in r.get("lines") or []:
            raw_v = L.get("raw_variety") or L.get("variety") or ""
            old_can = L.get("variety") or ""
            new_can, _recognized = _normalize_variety(raw_v)
            if new_can != old_can:
                if len(samples) < 25:
                    samples.append({
                        "raw": raw_v,
                        "old": old_can,
                        "new": new_can,
                        "po":  r.get("po_number") or "",
                        "date": r.get("production_date") or "",
                    })
                L["variety"] = new_can
                changed_lines += 1
                any_changed = True
        if any_changed:
            changed_records += 1
    if not dry_run:
        save_production(records)
    return jsonify({
        "ok": True,
        "dry_run": dry_run,
        "records_scanned": len(records),
        "records_changed": changed_records,
        "lines_changed": changed_lines,
        "samples": samples,
    })


@app.route("/api/admin/production/reclassify", methods=["POST"])
def api_admin_production_reclassify():
    """Re-run the warehouse classifier on every stored production record.

    Useful after we add new entries to _WAREHOUSE_TO_CANONICAL or extend
    _classify_warehouse — historical records ingested under the old
    rules keep their warehouse/distributor unchanged unless we ask the
    parser to look at them again. This walks data/production.json,
    re-classifies each row using its existing warehouse_raw (or empty
    -> "In-House Inventory" for older sheets without a header), and
    persists the updates.

    Body: { dry_run: bool }
    """
    from inventory_tracker import load_production, save_production
    from integrations.production_pdf_parser import _classify_warehouse

    body = request.json or {}
    dry_run = bool(body.get("dry_run", False))
    records = load_production()
    changed_count = 0
    samples = []
    for r in records:
        raw = r.get("warehouse_raw") or ""
        new_wh, new_dist = _classify_warehouse(raw)
        old_wh   = r.get("warehouse") or ""
        old_dist = r.get("distributor") or ""
        if new_wh == old_wh and new_dist == old_dist:
            continue
        changed_count += 1
        if len(samples) < 25:
            samples.append({
                "warehouse_raw": raw,
                "old_warehouse": old_wh,
                "new_warehouse": new_wh,
                "old_distributor": old_dist,
                "new_distributor": new_dist,
            })
        if not dry_run:
            r["warehouse"] = new_wh
            r["distributor"] = new_dist
    if not dry_run:
        save_production(records)
    return jsonify({
        "ok": True, "dry_run": dry_run,
        "total_records": len(records),
        "changed": changed_count,
        "samples": samples,
    })


@app.route("/api/production/<source_message_id>", methods=["DELETE"])
def api_production_delete(source_message_id):
    """Drop a production record by its source_message_id."""
    from inventory_tracker import load_production, save_production
    records = load_production()
    kept = [r for r in records if r.get("source_message_id") != source_message_id]
    removed = len(records) - len(kept)
    if removed:
        save_production(kept)
    return jsonify({"ok": True, "removed": removed})


# ---------------------------------------------------------------------------
# API – $PLH report (production revenue per labor hour at the bakery)
# ---------------------------------------------------------------------------
# Per-case sell prices (revenue side):
#   US Foods           $27.00
#   Cheney Brothers    $26.50
#   anything else      $29.50   (Chefs Warehouse, unassigned, etc.)
# Default labor rate $17/hr used to back-fill `dollars` on a labor entry
# that only carries `hours`. PLH = revenue / labor_hours.

CASE_PRICE_BY_DISTRIBUTOR = {
    "US Foods":        27.00,
    "Cheney Brothers": 26.50,
}
CASE_PRICE_DEFAULT = 29.50
LABOR_RATE_DEFAULT = 17.00


def _case_price_for(distributor: str) -> float:
    return CASE_PRICE_BY_DISTRIBUTOR.get((distributor or "").strip(),
                                         CASE_PRICE_DEFAULT)


def _plh_bucket_keys(grain: str, offset: int = 0):
    """Return [(key, start_iso, end_iso, label)] for the buckets in scope.

    grain = "week"     -> 4 ISO weeks ending this week (Mon-Sun)
    grain = "month"    -> 4 trailing calendar months ending this month
    grain = "quarter"  -> the 4 quarters of a calendar year

    offset shifts the window backward by:
      week   -> 1 week per step  (4-week window slides one week at a time)
      month  -> 1 month per step (4-month window slides one month at a time)
      quarter-> 1 year per step  (4-quarter window jumps a full year)
    offset = 0 is the current window. Negative offsets are clamped to 0
    (no "future" windows past now).
    """
    from datetime import datetime, timedelta
    today = datetime.now().date()
    if offset < 0:
        offset = 0

    if grain == "month":
        # 4 trailing calendar months ending with the current month. Each
        # prev/next click steps the window back by 1 month (overlapping
        # navigation rather than jumping the whole window at once).
        cur_abs = today.year * 12 + (today.month - 1)
        # The newest month in this window:
        newest_abs = cur_abs - offset
        # Build oldest -> newest so the chart reads left to right naturally.
        out = []
        for i in range(3, -1, -1):
            abs_i = newest_abs - i
            yy = abs_i // 12
            mm = (abs_i % 12) + 1
            start = datetime(yy, mm, 1).date()
            nxt_abs = abs_i + 1
            n_yy = nxt_abs // 12
            n_mm = (nxt_abs % 12) + 1
            end = (datetime(n_yy, n_mm, 1).date() - timedelta(days=1))
            key   = start.strftime("%Y-%m")
            label = start.strftime("%b %Y")
            out.append((key, start.isoformat(), end.isoformat(), label))
        return out

    if grain == "quarter":
        anchor_year = today.year - offset
        out = []
        for q in range(1, 5):
            start_mm = (q - 1) * 3 + 1
            start = datetime(anchor_year, start_mm, 1).date()
            end_mm = start_mm + 3
            end_yy = anchor_year
            if end_mm > 12:
                end_mm = 1
                end_yy = anchor_year + 1
            end = (datetime(end_yy, end_mm, 1).date() - timedelta(days=1))
            key   = f"{anchor_year}-Q{q}"
            label = f"Q{q} {anchor_year}"
            out.append((key, start.isoformat(), end.isoformat(), label))
        return out

    # default: weekly. 4 ISO weeks anchored on the Monday `offset` weeks
    # back from this week (each prev/next click shifts the window by 1 week,
    # so navigation is smooth and overlapping rather than jumping a full
    # 4-week window at a time).
    day = today.weekday()        # Mon=0..Sun=6
    this_monday = today - timedelta(days=day)
    anchor_monday = this_monday - timedelta(days=7 * offset)
    out = []
    for i in range(3, -1, -1):    # 3 windows ago -> latest
        wk_start = anchor_monday - timedelta(days=7 * i)
        wk_end   = wk_start + timedelta(days=6)
        iso = wk_start.isocalendar()
        key   = f"{iso.year:04d}-W{iso.week:02d}"
        label = f"Week of {wk_start.strftime('%b %d')}"
        out.append((key, wk_start.isoformat(), wk_end.isoformat(), label))
    return out


def _plh_window_label(grain: str, buckets: list) -> str:
    """Human label for the chart header — e.g. 'Apr 20 – May 17, 2026',
    'Apr – Jun 2026', or 'Q1 – Q4 2026'."""
    if not buckets:
        return ""
    first_start = buckets[0][1]
    last_end    = buckets[-1][2]
    from datetime import date
    fs = date.fromisoformat(first_start)
    le = date.fromisoformat(last_end)
    if grain == "quarter":
        # Always within one year for this view.
        return f"Q1 – Q4 {fs.year}"
    if grain == "month":
        if fs.year == le.year:
            return f"{fs.strftime('%b')} – {le.strftime('%b')} {fs.year}"
        return f"{fs.strftime('%b %Y')} – {le.strftime('%b %Y')}"
    # week
    if fs.year == le.year:
        return f"{fs.strftime('%b %d')} – {le.strftime('%b %d, %Y')}"
    return f"{fs.strftime('%b %d, %Y')} – {le.strftime('%b %d, %Y')}"


def _date_in_range(d: str, start: str, end: str) -> bool:
    if not d:
        return False
    d = d[:10]
    return start <= d <= end


def _bucket_is_empty(start: str, end: str, prod, labor, bsales) -> bool:
    """Cheap check: does any production / labor / bakery-sales row fall in
    the [start, end] window? Used to skip 'we have not gotten there yet'
    periods when anchoring the chart's default view."""
    for r in prod:
        if _date_in_range(r.get("production_date") or "", start, end):
            return False
    for e in labor:
        if _date_in_range(e.get("date") or "", start, end):
            if (e.get("hours") or 0) > 0 or (e.get("dollars") or 0) > 0:
                return False
    for w in bsales:
        if _date_in_range(w.get("week_start") or "", start, end):
            if (w.get("total") or 0) > 0:
                return False
    return True


@app.route("/api/report/plh")
def api_report_plh():
    """Production revenue, labor hours, and $PLH per time bucket.

    Query params:
      grain   "week" (default) | "month" | "quarter"
      offset  int >= 0. 0 = current window. Each step shifts back one
              full window: 4 weeks / 1 quarter / 1 year.
    """
    from inventory_tracker import load_production, load_labor, load_bakery_sales
    grain = (request.args.get("grain") or "week").lower()
    try:
        offset = max(0, int(request.args.get("offset") or 0))
    except (TypeError, ValueError):
        offset = 0
    prod  = load_production()
    labor = load_labor()
    bsales = load_bakery_sales()

    # For week/month grain, anchor the view at the most recent POPULATED
    # period instead of "today." If the current week or month hasn't been
    # filled in yet (the bakery model spreadsheet is updated weekly), the
    # auto_shift slides the window back so we still show four populated
    # bars instead of three + a blank.
    auto_shift = 0
    if grain in ("week", "month"):
        # Probe up to 4 periods back; if everything is empty, stop shifting
        # and just render the empties so the user can see the gap.
        for probe in range(0, 5):
            pb = _plh_bucket_keys(grain, offset=probe)
            ls, le = pb[-1][1], pb[-1][2]
            if not _bucket_is_empty(ls, le, prod, labor, bsales):
                auto_shift = probe
                break
    effective_offset = offset + auto_shift
    buckets = _plh_bucket_keys(grain, offset=effective_offset)

    out = []
    for key, start, end, label in buckets:
        bucket = {
            "key": key, "label": label, "start": start, "end": end,
            "total_cs": 0,
            "revenue_dollars": 0.0,            # cs * case price (production-side)
            "by_distributor": {},
            "labor_hours": 0.0,
            "labor_dollars": 0.0,
            "bakery_sales_dollars": 0.0,        # true sales from bakery_sales.json
            "bakery_sales_by_channel": {},
            "bakery_sales_weeks": 0,            # how many weekly rows landed here
            "splh": None,                       # sales per labor hour
            "labor_pct_of_sales": None,         # labor $ / sales $
            "plh": None,                        # legacy: revenue / hours
        }
        for r in prod:
            if not _date_in_range(r.get("production_date") or "", start, end):
                continue
            dist = r.get("distributor") or ""
            price = _case_price_for(dist)
            for line in (r.get("lines") or []):
                cs = int(line.get("cs_count") or 0)
                bucket["total_cs"] += cs
                bucket["revenue_dollars"] += cs * price
                bdist = bucket["by_distributor"].setdefault(
                    dist or "Other", {"cs": 0, "revenue_dollars": 0.0})
                bdist["cs"] += cs
                bdist["revenue_dollars"] += cs * price
        # Track bakery-xlsx labor separately so the sales-vs-labor metrics
        # (SPLH, labor % of sales) reconcile to the workbook's I11/O10
        # numbers exactly. The "all labor" pool still drives total labor
        # hours/cost and the production-side $PLH, since those reflect the
        # full picture of what the bakery actually paid out.
        bx_hours = 0.0
        bx_dollars = 0.0
        for e in labor:
            if not _date_in_range(e.get("date") or "", start, end):
                continue
            hrs = float(e.get("hours") or 0)
            dol = float(e.get("dollars") or 0)
            bucket["labor_hours"]   += hrs
            bucket["labor_dollars"] += dol
            if (e.get("source") or "") == "bakery-xlsx":
                bx_hours   += hrs
                bx_dollars += dol
        bucket["bakery_xlsx_labor_hours"]   = round(bx_hours, 4)
        bucket["bakery_xlsx_labor_dollars"] = round(bx_dollars, 2)
        # Aggregate bakery weekly sales whose week_start falls in this
        # bucket. A week is counted toward whichever bucket its Monday
        # lands in -- accepting the small straddle effect at month/
        # quarter boundaries in exchange for clean, deterministic math.
        for w in bsales:
            ws = (w.get("week_start") or "")
            if not _date_in_range(ws, start, end):
                continue
            bucket["bakery_sales_dollars"] += float(w.get("total") or 0)
            bucket["bakery_sales_weeks"]   += 1
            for ch, amt in (w.get("channels") or {}).items():
                bucket["bakery_sales_by_channel"][ch] = (
                    bucket["bakery_sales_by_channel"].get(ch, 0.0)
                    + float(amt or 0)
                )
        # If we have hours but no dollars on any entry, impute via the default
        # rate so the cost picture is at least directionally right.
        if bucket["labor_hours"] and not bucket["labor_dollars"]:
            bucket["labor_dollars"] = bucket["labor_hours"] * LABOR_RATE_DEFAULT
            bucket["labor_dollars_imputed"] = True
        if bucket["labor_hours"]:
            bucket["plh"] = bucket["revenue_dollars"] / bucket["labor_hours"]
        # Sales-vs-labor metrics ALWAYS use the bakery-xlsx labor pool so
        # the chart matches the workbook the user updates each week.
        if bx_hours and bucket["bakery_sales_dollars"]:
            bucket["splh"] = bucket["bakery_sales_dollars"] / bx_hours
        if bx_dollars and bucket["bakery_sales_dollars"]:
            bucket["labor_pct_of_sales"] = (
                bx_dollars / bucket["bakery_sales_dollars"]
            )
        out.append(bucket)

    return jsonify({
        "grain":            grain,
        "offset":           offset,
        "effective_offset": effective_offset,
        "auto_shift":       auto_shift,
        "window_label":     _plh_window_label(grain, buckets),
        "case_prices": {
            "US Foods": 27.00, "Cheney Brothers": 26.50, "default": 29.50,
        },
        "labor_rate_default": LABOR_RATE_DEFAULT,
        "buckets":          out,
    })


@app.route("/api/admin/labor/ingest", methods=["POST"])
def api_admin_labor_ingest():
    """Append or replace bakery labor entries.

    Body:
      entries:  [{date: YYYY-MM-DD, hours: float, dollars?: float, source?: str}]
      replace:  bool — when true, REPLACE all entries from the same `source`
                with the new set; when false (default), append + dedupe by
                (date, source) keeping the new entry.

    No Toast call here — this is the "feed me labor data from any source"
    endpoint. A separate script can pull from Toast and POST. CSV upload
    or manual entry can use this same endpoint.
    """
    from inventory_tracker import load_labor, save_labor
    body = request.json or {}
    incoming = body.get("entries") or []
    replace = bool(body.get("replace", False))
    if not incoming:
        return jsonify({"ok": False, "error": "entries required"}), 400

    existing = load_labor()
    sources_in_incoming = {(e.get("source") or "manual") for e in incoming}
    if replace:
        existing = [e for e in existing if (e.get("source") or "manual")
                    not in sources_in_incoming]

    # Dedupe by (date, source). Newer wins.
    by_key = {}
    for e in existing:
        by_key[(e.get("date"), e.get("source") or "manual")] = e
    for e in incoming:
        d = (e.get("date") or "").strip()
        if not d:
            continue
        rec = {
            "date":    d,
            "hours":   float(e.get("hours") or 0),
            "dollars": float(e.get("dollars") or 0),
            "source":  e.get("source") or "manual",
            "ingested_at": datetime_now_iso(),
        }
        by_key[(rec["date"], rec["source"])] = rec
    merged = list(by_key.values())
    save_labor(merged)
    return jsonify({"ok": True, "total_entries": len(merged),
                    "added_or_updated": len(incoming)})


# ---------------------------------------------------------------------------
# API -- Bakery weekly sales (used while Toast is not connected for the
# production bakery). Source is the "Bakery Model - Sales v. Labor"
# spreadsheet JD updates weekly.
# ---------------------------------------------------------------------------

@app.route("/api/admin/bakery-sales/ingest", methods=["POST"])
def api_admin_bakery_sales_ingest():
    """Append or replace bakery weekly sales entries.

    Body:
      entries:  [
        {
          week_start:  "YYYY-MM-DD",   # Monday -- required, used as the key
          week_end:    "YYYY-MM-DD",
          channels:    {channel: dollars, ...},
          total:       float,
          splh:        float | null,
          source:      str,            # default "bakery-xlsx"
        }, ...
      ]
      replace:  bool -- when true, REPLACE all entries that share a source
                with the incoming set; when false (default), dedupe by
                (week_start, source) keeping the new entry.
    """
    from inventory_tracker import load_bakery_sales, save_bakery_sales
    body = request.json or {}
    incoming = body.get("entries") or []
    replace = bool(body.get("replace", False))
    if not incoming:
        return jsonify({"ok": False, "error": "entries required"}), 400

    existing = load_bakery_sales()
    sources_in_incoming = {(e.get("source") or "bakery-xlsx") for e in incoming}
    if replace:
        existing = [e for e in existing
                    if (e.get("source") or "bakery-xlsx")
                    not in sources_in_incoming]

    by_key = {}
    for e in existing:
        by_key[(e.get("week_start"), e.get("source") or "bakery-xlsx")] = e
    added = 0
    for e in incoming:
        ws_ = (e.get("week_start") or "").strip()
        if not ws_:
            continue
        rec = {
            "week_start": ws_,
            "week_end":   (e.get("week_end") or "").strip() or None,
            "location":   e.get("location") or "Bakery",
            "channels":   {k: float(v or 0) for k, v in
                           (e.get("channels") or {}).items()},
            "total":      float(e.get("total") or 0),
            "splh":       (float(e["splh"]) if e.get("splh") not in
                           (None, "", "#DIV/0!") else None),
            "source":     e.get("source") or "bakery-xlsx",
            "ingested_at": datetime_now_iso(),
        }
        by_key[(rec["week_start"], rec["source"])] = rec
        added += 1
    merged = sorted(by_key.values(), key=lambda r: r.get("week_start") or "")
    save_bakery_sales(merged)
    return jsonify({"ok": True, "total_entries": len(merged),
                    "added_or_updated": added})


@app.route("/api/report/bakery-sales")
def api_report_bakery_sales():
    """Return bakery weekly sales rows, optionally filtered by date range.

    Query params:
      start  YYYY-MM-DD  (inclusive -- compared against week_start)
      end    YYYY-MM-DD  (inclusive -- compared against week_start)
      limit  int         (default 26 -- most recent N weeks)
    """
    from inventory_tracker import load_bakery_sales
    rows = load_bakery_sales() or []
    args = request.args
    start = (args.get("start") or "").strip()
    end   = (args.get("end") or "").strip()
    if start:
        rows = [r for r in rows if (r.get("week_start") or "") >= start]
    if end:
        rows = [r for r in rows if (r.get("week_start") or "") <= end]
    try:
        limit = int(args.get("limit") or 26)
    except (TypeError, ValueError):
        limit = 26
    rows = sorted(rows, key=lambda r: r.get("week_start") or "")
    if limit and limit > 0:
        rows = rows[-limit:]
    channel_totals: dict[str, float] = {}
    grand_total = 0.0
    for r in rows:
        grand_total += float(r.get("total") or 0)
        for ch, v in (r.get("channels") or {}).items():
            channel_totals[ch] = channel_totals.get(ch, 0.0) + float(v or 0)
    return jsonify({
        "rows": rows,
        "summary": {
            "weeks":          len(rows),
            "total":          grand_total,
            "channel_totals": channel_totals,
        },
    })


def datetime_now_iso() -> str:
    from datetime import datetime
    return datetime.now().isoformat()



# ---------------------------------------------------------------------------
# API – Forecast bridge (weekly_usage -> usage ledger, so lots burn down)
# ---------------------------------------------------------------------------
# Background: every SKU carries a `weekly_usage` rate (cs/wk) but nothing in
# the system converted that rate into actual usage events, so production
# lots never showed cs_consumed > 0 at the DCs. The user's request: bridge
# the rate into the usage ledger so FIFO lot consumption works, with a
# weekly true-up that reconciles forecasts against vendor on-hand snapshots.
#
# Design:
#   - /api/forecast/decrement-daily   Idempotent per (sku, UTC date).
#                                     Posts +weekly_usage/7 as a positive
#                                     usage event tagged source="forecast-daily"
#                                     and decrements item quantity by the
#                                     same amount. Triggered by the daily
#                                     Render cron (bagel-inventory-forecast-daily).
#   - /api/forecast/true-up           Operator posts a vendor on-hand
#                                     snapshot. We reverse every uncovered
#                                     "forecast-daily" entry since the prior
#                                     truth, post the actual usage as
#                                     source="vendor-truth", and set
#                                     quantity = reported_qty.
#
# Both write directly through load_usage/save_usage so they can tag entries
# with a custom `source` (the existing record_usage helper hard-codes the
# absence of one). The FIFO lot consumption logic in
# inventory_tracker.compute_lot_fifo_state already honors `reversed: true`
# and skips `source == "reversal"`, so the moving parts are minimal.

@app.route("/api/forecast/decrement-daily", methods=["POST"])
def api_forecast_decrement_daily():
    """Apply one day's worth of forecast usage to every SKU.

    Idempotent: for each SKU, if a non-reversed entry already exists today
    with source == "forecast-daily", that SKU is skipped. Safe to re-run
    or to call mid-day after the cron.

    Body (all optional):
      dry_run             bool   compute deltas without writing
      date                str    YYYY-MM-DD override for testing (defaults
                                 to today UTC). The idempotency check keys
                                 off this value.
      warehouse_prefix    str    restrict to SKUs whose `warehouse` field
                                 starts with this string (e.g. "Zebulon").
                                 Default: all SKUs with weekly_usage > 0.

    Returns {ok, date, applied: [...], skipped_existing: [...], dry_run}.
    """
    from datetime import datetime, timezone
    from inventory_tracker import (
        load_inventory, save_inventory, load_usage, save_usage,
        _variety_from_name,
    )

    body = request.json or {}
    dry_run = bool(body.get("dry_run"))
    date_iso = (body.get("date") or "").strip()
    wh_prefix = (body.get("warehouse_prefix") or "").strip()
    if not date_iso:
        date_iso = datetime.now(timezone.utc).date().isoformat()

    inv = load_inventory()
    usage = load_usage()

    # Index existing forecast-daily entries by (item_key, YYYY-MM-DD).
    # Reversed entries don't count — the slot is "open" to re-fill.
    have_today: set[tuple[str, str]] = set()
    for e in usage:
        if e.get("source") != "forecast-daily":
            continue
        if e.get("reversed"):
            continue
        ts = (e.get("timestamp") or "")[:10]
        if not ts:
            continue
        have_today.add((e.get("item_key") or "", ts))

    applied: list[dict] = []
    skipped: list[str] = []
    now_iso = datetime.now(timezone.utc).isoformat()
    for key, item in inv.items():
        wkly = float(item.get("weekly_usage") or 0)
        if wkly <= 0:
            continue
        wh = item.get("warehouse") or ""
        if wh_prefix and not wh.startswith(wh_prefix):
            continue
        if (key, date_iso) in have_today:
            skipped.append(item.get("name") or key)
            continue
        daily_amount = round(wkly / 7.0, 4)
        if daily_amount <= 0:
            continue
        cur_qty = float(item.get("quantity") or 0)
        new_qty = round(cur_qty - daily_amount, 4)
        # Floor at 0 — we don't model negative inventory; the operator
        # will see the SKU pinned at 0 and reorder.
        if new_qty < 0:
            new_qty = 0.0
            daily_amount = round(cur_qty, 4)
            if daily_amount <= 0:
                # Already at zero; nothing to decrement, but record
                # nothing so the SKU stays eligible tomorrow.
                continue
        applied.append({
            "name": item.get("name"),
            "warehouse": wh,
            "weekly_usage": wkly,
            "daily_amount": daily_amount,
            "old_quantity": cur_qty,
            "new_quantity": new_qty,
        })
        if not dry_run:
            entry = {
                "item_key": key,
                "item_name": item.get("name") or "",
                "amount": daily_amount,
                "unit": item.get("unit") or "",
                "note": f"Forecast burn (weekly_usage/7) for {date_iso}",
                "timestamp": now_iso,
                "source": "forecast-daily",
                "forecast_date": date_iso,
            }
            if wh:
                entry["warehouse"] = wh
            v = _variety_from_name(item.get("name") or "")
            if v:
                entry["variety"] = v
            usage.append(entry)
            item["quantity"] = new_qty
            item["updated"] = now_iso

    if not dry_run and applied:
        save_usage(usage)
        save_inventory(inv)

    return jsonify({
        "ok": True,
        "date": date_iso,
        "dry_run": dry_run,
        "applied": applied,
        "skipped_existing": skipped,
    })


@app.route("/api/forecast/backfill-historical", methods=["POST"])
def api_forecast_backfill_historical():
    """One-shot opening reconciliation between lot production and on-hand.

    Why this exists:
        Production lots get recorded as they arrive (each PO bakes new
        lots into data/production.json), but FIFO consumption only ever
        knew about *future* burns through the usage ledger. There was no
        equivalent for the "between PO arrival and the first vendor
        snapshot" gap — so the lots-by-pair view sat at cs_consumed=0
        forever, even on SKUs whose on-hand was clearly much smaller
        than total cases received.

    What this does:
        For each (warehouse, variety) pair on an inventory item:

            target_consumed = max(0, sum(lots_produced) - quantity)
            existing_consumed = sum(positive non-reversed usage entries)
            delta = target_consumed - existing_consumed

        If delta > 0, append one positive usage event of `delta` tagged
        `source = "historical-backfill"` so FIFO drains oldest-first
        until the lot remaining sums match the on-hand quantity. The
        endpoint does NOT touch item.quantity — quantity is already
        ground truth from the vendor; this purely seeds the consumption
        side of the ledger.

    Body (all optional):
        warehouse   str    limit to one warehouse (e.g. "Zebulon, NC").
                           Default: every warehouse.
        dry_run     bool   compute deltas without writing.

    Returns {ok, applied: [...], no_op: [...], dry_run}.
    """
    from datetime import datetime, timezone
    from inventory_tracker import (
        load_inventory, load_production, load_usage, save_usage,
        _variety_from_name,
    )

    body = request.json or {}
    warehouse = (body.get("warehouse") or "").strip()
    dry_run = bool(body.get("dry_run"))

    inv = load_inventory()
    prod = load_production()
    usage = load_usage()

    # Sum produced cases per (warehouse, variety).
    produced_by_pair: dict[tuple[str, str], float] = {}
    for r in prod:
        wh = r.get("warehouse") or ""
        for L in r.get("lines") or []:
            v = (L.get("variety") or "").strip()
            if not wh or not v:
                continue
            produced_by_pair[(wh, v)] = (
                produced_by_pair.get((wh, v), 0.0)
                + float(L.get("cs_count") or 0)
            )

    # Existing positive (non-reversed) consumption per (warehouse, variety).
    inv_for_lookup = inv
    consumed_by_pair: dict[tuple[str, str], float] = {}
    for e in usage:
        if e.get("reversed"):
            continue
        if e.get("source") == "reversal":
            continue
        amt = float(e.get("amount") or 0)
        if amt <= 0:
            continue
        e_wh = e.get("warehouse") or ""
        e_var = e.get("variety") or ""
        if not e_wh or not e_var:
            it = inv_for_lookup.get(e.get("item_key") or "") or {}
            e_wh = e_wh or it.get("warehouse", "")
            e_var = e_var or _variety_from_name(
                it.get("name") or e.get("item_name") or "")
        if not e_wh or not e_var:
            continue
        consumed_by_pair[(e_wh, e_var)] = (
            consumed_by_pair.get((e_wh, e_var), 0.0) + amt
        )

    applied: list[dict] = []
    no_op: list[dict] = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for key, item in inv.items():
        wh = item.get("warehouse") or ""
        if warehouse and wh != warehouse:
            continue
        name = item.get("name") or ""
        v = _variety_from_name(name)
        if not wh or not v:
            continue
        produced = produced_by_pair.get((wh, v), 0.0)
        if produced <= 0:
            continue
        cur_qty = float(item.get("quantity") or 0)
        target_consumed = max(0.0, produced - cur_qty)
        existing = consumed_by_pair.get((wh, v), 0.0)
        delta = round(target_consumed - existing, 4)
        if delta <= 0.001:
            no_op.append({
                "name": name, "warehouse": wh,
                "produced": produced, "quantity": cur_qty,
                "target_consumed": target_consumed,
                "existing_consumed": existing,
            })
            continue
        applied.append({
            "name": name, "warehouse": wh,
            "produced": produced, "quantity": cur_qty,
            "target_consumed": target_consumed,
            "existing_consumed": existing,
            "backfill_amount": delta,
        })
        if not dry_run:
            usage.append({
                "item_key": key,
                "item_name": name,
                "amount": delta,
                "unit": item.get("unit") or "cs",
                "note": (f"Opening reconciliation: align FIFO lots "
                         f"with vendor on-hand ({cur_qty} cs)"),
                "timestamp": now_iso,
                "source": "historical-backfill",
                "warehouse": wh,
                "variety": v,
            })

    if not dry_run and applied:
        save_usage(usage)

    return jsonify({
        "ok": True,
        "warehouse": warehouse or "(all)",
        "dry_run": dry_run,
        "applied": applied,
        "no_op": no_op,
    })


@app.route("/api/forecast/true-up", methods=["POST"])
def api_forecast_true_up():
    """Reconcile against a vendor on-hand snapshot.

    The vendor's number is ground truth. We reverse every uncovered
    "forecast-daily" entry on each named SKU since the last vendor-truth
    (or since the SKU was added if there has never been one), then post a
    single "vendor-truth" usage entry equal to:

        actual_used = prior_truth_qty + arrivals_since - reported_qty

    where:
        prior_truth_qty   = item.quantity (live) + forecast_burned - arrivals_since
                            (i.e. what on-hand WAS at the last truth date)
        arrivals_since    = sum of negative-amount entries since last truth
        forecast_burned   = sum of positive forecast-daily entries since last
                            truth that are NOT already reversed
        reported_qty      = the new vendor snapshot

    If actual_used <= 0 we skip the positive entry (vendor reports more
    than expected — that's a count correction, not a consumption event).

    Quantity is set to reported_qty in inventory.json so the dashboard
    matches the vendor's number.

    Body:
      warehouse     str    label that matches inventory items' `warehouse`
                           field (e.g. "Zebulon, NC"). Required.
      reported_at   str    ISO date/datetime of the snapshot. Defaults to
                           now (UTC).
      items         list   [{name: str, reported_qty: float}, ...]

    Returns {ok, reported_at, reconciled: [...], skipped: [...]}.
    """
    from datetime import datetime, timezone
    from inventory_tracker import (
        load_inventory, save_inventory, load_usage, save_usage,
        _variety_from_name,
    )

    body = request.json or {}
    warehouse = (body.get("warehouse") or "").strip()
    if not warehouse:
        return jsonify({"ok": False, "error": "warehouse required"}), 400
    items_in = body.get("items") or []
    if not isinstance(items_in, list) or not items_in:
        return jsonify({"ok": False, "error": "items must be a non-empty list"}), 400
    reported_at = (body.get("reported_at") or "").strip()
    if not reported_at:
        reported_at = datetime.now(timezone.utc).isoformat()
    dry_run = bool(body.get("dry_run"))

    inv = load_inventory()
    usage = load_usage()

    # Build a lookup of usage entries by item_key for fast filtering.
    by_key: dict[str, list[dict]] = {}
    for idx, e in enumerate(usage):
        k = e.get("item_key") or ""
        if k:
            by_key.setdefault(k, []).append((idx, e))

    reconciled: list[dict] = []
    skipped: list[dict] = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for payload in items_in:
        name = (payload.get("name") or "").strip()
        if not name:
            skipped.append({"name": "", "reason": "missing name"})
            continue
        try:
            reported_qty = float(payload.get("reported_qty"))
        except (TypeError, ValueError):
            skipped.append({"name": name, "reason": "reported_qty not numeric"})
            continue
        key = name.lower()
        item = inv.get(key)
        if not item:
            skipped.append({"name": name, "reason": "SKU not found"})
            continue
        if (item.get("warehouse") or "") != warehouse:
            skipped.append({"name": name, "reason": f"warehouse mismatch ({item.get('warehouse')})"})
            continue

        # Find last vendor-truth timestamp for this SKU (exclusive lower
        # bound for the reconciliation window). Empty string = no prior
        # truth, so window is "all history".
        prior_truth_ts = ""
        for _idx, e in by_key.get(key, []):
            if e.get("source") != "vendor-truth":
                continue
            if e.get("reversed"):
                continue
            ts = e.get("timestamp") or ""
            if ts > prior_truth_ts:
                prior_truth_ts = ts

        arrivals_since = 0.0       # sum of negative amounts (restocks)
        forecast_burned = 0.0      # positive forecast-daily not reversed
        reverse_indices: list[int] = []
        for idx, e in by_key.get(key, []):
            if e.get("reversed"):
                continue
            ts = e.get("timestamp") or ""
            if prior_truth_ts and ts <= prior_truth_ts:
                continue
            amt = float(e.get("amount") or 0)
            if amt < 0:
                arrivals_since += -amt
            elif amt > 0 and e.get("source") == "forecast-daily":
                forecast_burned += amt
                reverse_indices.append(idx)
            # Positive non-forecast entries (manual /api/use, prior
            # vendor-truth) are left untouched — they represent real
            # consumption the operator already booked.

        cur_qty = float(item.get("quantity") or 0)
        prior_truth_qty = cur_qty + forecast_burned - arrivals_since
        actual_used = prior_truth_qty + arrivals_since - reported_qty
        actual_used = round(actual_used, 4)

        entry_record = {
            "name": name,
            "warehouse": warehouse,
            "prior_truth_ts": prior_truth_ts or None,
            "prior_truth_qty": round(prior_truth_qty, 4),
            "arrivals_since": round(arrivals_since, 4),
            "forecast_burned_reversed": round(forecast_burned, 4),
            "reported_qty": reported_qty,
            "actual_used": actual_used,
            "old_quantity": cur_qty,
            "new_quantity": reported_qty,
        }
        reconciled.append(entry_record)

        if dry_run:
            continue

        # Reverse the forecast-daily entries that fell in this window.
        for idx in reverse_indices:
            usage[idx]["reversed"] = True
            usage[idx]["reversed_at"] = now_iso
            usage[idx]["reversed_by"] = "vendor-truth"

        # Post the vendor-truth event. Skip when actual_used <= 0
        # (vendor reported MORE than expected — a positive count
        # correction we model as the quantity bump only, no consumption).
        if actual_used > 0:
            truth_entry = {
                "item_key": key,
                "item_name": item.get("name") or name,
                "amount": actual_used,
                "unit": item.get("unit") or "",
                "note": f"Vendor on-hand true-up at {reported_at}",
                "timestamp": now_iso,
                "source": "vendor-truth",
                "reported_at": reported_at,
                "warehouse": warehouse,
            }
            v = _variety_from_name(item.get("name") or name)
            if v:
                truth_entry["variety"] = v
            usage.append(truth_entry)

        item["quantity"] = reported_qty
        item["updated"] = now_iso
        item["last_synced"] = now_iso
        item["last_synced_from"] = "vendor-truth"
        item["last_count_at"] = now_iso

    if not dry_run and reconciled:
        save_usage(usage)
        save_inventory(inv)

    return jsonify({
        "ok": True,
        "warehouse": warehouse,
        "reported_at": reported_at,
        "dry_run": dry_run,
        "reconciled": reconciled,
        "skipped": skipped,
    })


# ---------------------------------------------------------------------------
# API – Usage
# ---------------------------------------------------------------------------

@app.route("/api/use", methods=["POST"])
def api_use():
    d = request.json
    record_usage(d["name"], float(d["amount"]), d.get("note", ""))
    return jsonify({"ok": True})


@app.route("/api/restock", methods=["POST"])
def api_restock():
    d = request.json
    restock(d["name"], float(d["amount"]), d.get("note", ""))
    return jsonify({"ok": True})


@app.route("/api/usage")
def api_usage():
    name = request.args.get("name")
    limit = int(request.args.get("limit", 50))
    usage = load_usage()
    if name:
        key = name.lower().strip()
        usage = [e for e in usage if e["item_key"] == key]
    usage = list(reversed(usage))[:limit]
    return jsonify(usage)


@app.route("/api/usage/reverse", methods=["POST"])
def api_usage_reverse():
    """Undo a single usage/restock entry by its timestamp."""
    d = request.json or {}
    ts = (d.get("timestamp") or "").strip()
    if not ts:
        return jsonify({"ok": False, "error": "timestamp required"}), 400
    result = reverse_usage(ts)
    return jsonify(result)


# ---------------------------------------------------------------------------
# API – Report
# ---------------------------------------------------------------------------

@app.route("/api/report")
def api_report():
    inv = load_inventory()
    usage = load_usage()

    total_value = sum(i["quantity"] * i["price"] for i in inv.values())
    total_items = len(inv)
    low_stock = [i for i in inv.values() if i["quantity"] <= i["low_stock_threshold"]]

    consumed: dict = {}
    restocked: dict = {}
    for e in usage:
        key = e["item_key"]
        if e["amount"] < 0:
            restocked[key] = restocked.get(key, 0) + abs(e["amount"])
        else:
            consumed[key] = consumed.get(key, 0) + e["amount"]

    top_consumed = sorted(
        [{"key": k, "name": inv.get(k, {}).get("name", k),
          "unit": inv.get(k, {}).get("unit", ""), "total": v}
         for k, v in consumed.items()],
        key=lambda x: x["total"], reverse=True
    )[:10]

    top_restocked = sorted(
        [{"key": k, "name": inv.get(k, {}).get("name", k),
          "unit": inv.get(k, {}).get("unit", ""), "total": v}
         for k, v in restocked.items()],
        key=lambda x: x["total"], reverse=True
    )[:10]

    return jsonify({
        "total_value": round(total_value, 2),
        "total_items": total_items,
        "low_stock_count": len(low_stock),
        "low_stock": low_stock,
        "top_consumed": top_consumed,
        "top_restocked": top_restocked,
        "total_usage_events": len(usage),
    })


# ---------------------------------------------------------------------------
# Warehouse catalogue (authoritative list used by UI + seeds)
# ---------------------------------------------------------------------------

WAREHOUSES = {
    "Cheney Brothers": [
        "Riviera Beach, FL",
        "Ocala, FL",
        "Punta Gorda, FL",
    ],
    "US Foods": [
        "Manassas, VA",
        "Zebulon, NC",
        "La Mirada, CA",
        "Chicago, IL",
        "Alcoa, TN",
    ],
    # Chefs Warehouse DCs receive their own POs but DO NOT appear in
    # the Inventory tab -- they're surfaced only on the Pending POs
    # tab via /api/chefs-warehouse/pos. CW operates four FULLY SEPARATE
    # receiving DCs (no consolidation between them; each issues its own
    # PO#s):
    #   - Bronx, NY      a.k.a. "CW NY" (Dairyland)
    #   - Chicago, IL    a.k.a. "CW Midwest"
    #   - Hanover, MD    a.k.a. "CW Mid-Atlantic"
    #   - Opa Locka, FL  a.k.a. "CW Florida"
    "Chefs Warehouse": [
        "Bronx, NY",
        "Chicago, IL",
        "Hanover, MD",
        "Opa Locka, FL",
    ],
}


@app.route("/api/report-status")
def api_report_status():
    """JSON: which distributor warehouses have sent this week\'s inventory & usage report."""
    from integrations import report_status as _rs
    force = request.args.get("refresh") in ("1", "true", "yes", "on")
    return jsonify(_rs.get_status(force=force))


@app.route("/report-status")
def report_status_page():
    """Mobile-friendly weekly-report status page (read-only, no login)."""
    from flask import Response
    from integrations import report_status as _rs
    force = request.args.get("refresh") in ("1", "true", "yes", "on")
    html = _rs.render_html(_rs.get_status(force=force))
    return Response(html, mimetype="text/html")


@app.route("/api/warehouses")
def api_warehouses():
    return jsonify(WAREHOUSES)


# ---------------------------------------------------------------------------
# API – Distributors (unified view across Cheney Brothers and US Foods)
# ---------------------------------------------------------------------------

def _warehouse_last_count(wh_items):
    """Most recent 'count received' timestamp for a warehouse.

    A count is a rep inventory worksheet (stamps item['last_count_at']) or a
    vendor-truth on-hand true-up. Falls back to a vendor-truth last_synced so
    vendor-fed warehouses still show a date before their next worksheet lands.
    Returns an ISO string or None. ISO strings sort chronologically.
    """
    best = None
    for x in wh_items:
        ts = x.get("last_count_at")
        if not ts and x.get("last_synced_from") == "vendor-truth":
            ts = x.get("last_synced")
        if ts and (best is None or ts > best):
            best = ts
    return best


@app.route("/api/distributors")
def api_distributors():
    inv = load_inventory()
    items_enriched = [_enrich_on_order(dict(v)) for v in inv.values()]
    groups: dict[str, list] = {}
    for item in items_enriched:
        dist = item.get("distributor") or "Unassigned"
        groups.setdefault(dist, []).append(item)

    summary = []
    for dist, items in sorted(groups.items()):
        total_qty = sum(i["quantity"] for i in items)
        total_value = sum(i["quantity"] * i["price"] for i in items)
        total_on_order = sum(i.get("on_order_qty", 0) for i in items)
        low = [i for i in items if i["quantity"] <= i["low_stock_threshold"]]

        # Sub-group by warehouse
        wh_groups: dict[str, list] = {}
        for i in items:
            wh_groups.setdefault(i.get("warehouse") or "Unassigned", []).append(i)
        warehouses = []
        for wh_name, wh_items in sorted(wh_groups.items()):
            warehouses.append({
                "warehouse": wh_name,
                "last_count_at": _warehouse_last_count(wh_items),
                "item_count": len(wh_items),
                "total_quantity": round(sum(x["quantity"] for x in wh_items), 2),
                "total_value": round(sum(x["quantity"] * x["price"] for x in wh_items), 2),
                "total_on_order": round(sum(x.get("on_order_qty", 0) for x in wh_items), 2),
                "low_stock_count": sum(1 for x in wh_items if x["quantity"] <= x["low_stock_threshold"]),
                "items": sorted(wh_items, key=lambda x: x["name"]),
            })

        summary.append({
            "distributor": dist,
            "item_count": len(items),
            "total_quantity": round(total_qty, 2),
            "total_value": round(total_value, 2),
            "total_on_order": round(total_on_order, 2),
            "low_stock_count": len(low),
            "warehouses": warehouses,
        })
    return jsonify(summary)


# ---------------------------------------------------------------------------
# API – Sync (pull current on-hand from distributors)
# ---------------------------------------------------------------------------

@app.route("/api/sync", methods=["POST"])
def api_sync():
    from sync_inventory import sync_all

    dry_run = bool((request.json or {}).get("dry_run", False))
    reports = sync_all(dry_run=dry_run)
    return jsonify({"dry_run": dry_run, "reports": reports})


@app.route("/api/seed", methods=["POST"])
def api_seed():
    from seed_bagels import seed

    reset = bool((request.json or {}).get("reset", False))
    try:
        summary = seed(reset=reset)
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "error": str(exc)}), 200
    return jsonify({"ok": True, **summary})


@app.route("/api/migrate-units", methods=["POST"])
def api_migrate_units():
    from inventory_tracker import migrate_units_to_case

    inv = load_inventory()
    try:
        summary = migrate_units_to_case(inv)
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "error": str(exc)}), 200
    save_inventory(inv)
    return jsonify({"ok": True, **summary})


@app.route("/api/email/scan", methods=["POST"])
def api_email_scan():
    # The whole route runs inside one try block so that ANY failure (import,
    # JSON parse, scan_email itself) becomes a structured 200 with status:
    # "error" + a traceback excerpt -- never a generic Flask 500. That's a
    # lot easier to debug from the client side when there are no logs handy.
    #
    # Bound the work so we don't blow past gunicorn's worker timeout. The
    # original default (300 messages * N mailboxes, each pulling full MIME
    # bytes via Graph) regularly killed the worker on Render's starter
    # plan, which surfaced as a generic 500 at the edge regardless of the
    # try/except below. Callers can override `max_messages` in the body or
    # set MS365_FILTER for a wider sweep done out of band.
    import traceback as _tb
    dry_run = False
    try:
        from integrations.email_scanner import EmailInboxClient
        from sync_inventory import _apply_events
        body = request.json or {}
        dry_run = bool(body.get("dry_run", False))
        # Default keeps within gunicorn's 180s budget (60 messages * 2
        # mailboxes is comfortable). The hard cap of 2000 lets ad-hoc deep
        # sweeps go further when paired with a narrowing MS365_FILTER; the
        # 180s worker timeout still binds, so callers requesting more than a
        # few hundred without a filter should expect a 504.
        try:
            max_messages = int(body.get("max_messages") or 60)
        except (TypeError, ValueError):
            max_messages = 60
        max_messages = max(1, min(max_messages, 2000))

        # Optional wide-lookback override. Without this, the scan uses
        # whatever MS365_FILTER is set to on the service (typically empty,
        # which returns the most recent max_messages). Pass `lookback_days`
        # to one-shot a deeper sweep -- useful for backfilling on_order
        # entries whose source emails predate the normal scan window.
        filter_override = None
        try:
            lookback_days = int(body.get("lookback_days") or 0)
        except (TypeError, ValueError):
            lookback_days = 0
        if lookback_days > 0:
            from datetime import datetime, timezone, timedelta
            since = datetime.now(timezone.utc) - timedelta(days=lookback_days)
            # Graph wants ISO 8601 with a Z suffix, no microseconds.
            iso = since.replace(microsecond=0).isoformat().replace("+00:00", "Z")
            # Mirror cowork_graph_scan: pre-filter to attachment-bearing
            # messages so the page budget isn't burned on non-PO mail.
            # Graph rejects hasAttachments + $orderby (InefficientFilter),
            # so _scan_ms365_mailbox drops orderby when the filter
            # contains "hasAttachments".
            filter_override = f"hasAttachments eq true and receivedDateTime ge {iso}"

        client = EmailInboxClient()
        try:
            scan = client.scan(max_messages=max_messages,
                               filter_override=filter_override)
        except Exception as exc:  # noqa: BLE001 — surface NotConfigured + transport
            report = {
                "distributor": "Email Inbox",
                "source": client.source(),
                "status": ("not_configured" if type(exc).__name__ == "NotConfiguredError"
                           else "error"),
                "fetched": 0, "updated": 0, "unchanged": 0,
                "unmatched": [], "changes": [],
                "error": f"{type(exc).__name__}: {exc}",
                "traceback": _tb.format_exc()[-2000:],
                "messages_seen": 0, "messages_parsed": 0,
                "by_event_type": {"on_hand": 0, "restock": 0, "usage": 0},
            }
            return jsonify({"dry_run": dry_run, "reports": [report]})

        report = _apply_events(
            events=list(scan.events),
            messages_seen=scan.messages_seen,
            messages_parsed=scan.messages_parsed,
            errors=list(scan.errors or []),
            dry_run=dry_run,
            source=client.source(),
        )
    except Exception as exc:  # noqa: BLE001
        report = {
            "distributor": "Email Inbox",
            "source": "unknown",
            "status": "error",
            "fetched": 0,
            "updated": 0,
            "unchanged": 0,
            "unmatched": [],
            "changes": [],
            "error": f"{type(exc).__name__}: {exc}",
            "traceback": _tb.format_exc()[-2000:],
            "messages_seen": 0,
            "messages_parsed": 0,
        }
    return jsonify({"dry_run": dry_run, "reports": [report]})


@app.route("/api/email/ingest-events", methods=["POST"])
def api_email_ingest_events():
    """Accept externally-parsed EmailEvents and apply them through the same
    PO revision-replace pipeline as /api/email/scan.

    Used by the Cowork scheduled routine that reads M365 mailboxes via the
    Outlook MCP, parses attachments client-side, and POSTs events here -- so
    the web service never needs outbound Graph credentials.

    Request body:
        {
          "dry_run": false,
          "source": "cowork-routine",        # free-form tag for the report
          "messages_seen": 12,
          "messages_parsed": 3,
          "errors": ["..."],
          "events": [
            {
              "event_type": "restock"|"on_hand"|"usage",
              "item": {
                "quantity": 24.0,
                "distributor": "US Foods",
                "name": "Plain Bagel 4oz [USF - Manassas]",   # optional
                "variety": "Plain",                            # optional
                "warehouse": "Manassas, VA",                   # optional
                "unit": "each",                                # optional
                "price": 0.0,                                  # optional
                "case_cost": 27.0,                             # optional
                "case_size": 168,                              # optional
                "weekly_usage": 0                              # optional
              },
              "source_message_id": "...",
              "source_subject": "...",
              "po_number": "2125123456",   # required for PO revision semantics
              "po_revision": "1",          # numeric string; "" allowed
              "po_order_date": "2026-05-13" # ISO YYYY-MM-DD from the PO PDF;
                                            #   anchors ordered_at + ETA in
                                            #   the on_order entry. Optional.
            }, ...
          ]
        }
    """
    import traceback as _tb
    try:
        from integrations import EmailEvent, SyncItem
        from sync_inventory import _apply_events
    except Exception as exc:  # noqa: BLE001
        return jsonify({
            "dry_run": False,
            "reports": [{
                "distributor": "Email Inbox",
                "source": "external",
                "status": "error",
                "fetched": 0,
                "updated": 0,
                "unchanged": 0,
                "unmatched": [],
                "changes": [],
                "error": f"import failed: {type(exc).__name__}: {exc}",
                "traceback": _tb.format_exc()[-2000:],
                "messages_seen": 0,
                "messages_parsed": 0,
            }],
        })

    payload = request.json or {}
    dry_run = bool(payload.get("dry_run", False))
    source = str(payload.get("source") or "external").strip() or "external"
    messages_seen = int(payload.get("messages_seen") or 0)
    messages_parsed = int(payload.get("messages_parsed") or 0)
    errors = list(payload.get("errors") or [])
    raw_events = payload.get("events") or []

    if not isinstance(raw_events, list):
        return jsonify({"ok": False, "error": "events must be a list"}), 400

    # Filter out events for POs that were canceled by the operator. The
    # source emails may still live in the inbox so the scanner keeps
    # surfacing them — we silently drop them here so the cancel sticks.
    from inventory_tracker import load_canceled_pos
    canceled = load_canceled_pos()
    canceled_skipped = 0
    if canceled:
        kept_events = []
        for ev in raw_events:
            po = str((ev or {}).get("po_number") or "").strip()
            if po and po in canceled:
                canceled_skipped += 1
                continue
            kept_events.append(ev)
        raw_events = kept_events

    built: list[EmailEvent] = []
    build_errors: list[str] = []
    for idx, e in enumerate(raw_events):
        if not isinstance(e, dict):
            build_errors.append(f"events[{idx}]: not an object")
            continue
        try:
            etype = e.get("event_type")
            if etype not in ("on_hand", "restock", "usage"):
                build_errors.append(f"events[{idx}]: bad event_type {etype!r}")
                continue
            raw_item = e.get("item") or {}
            item = SyncItem(
                quantity=float(raw_item.get("quantity") or 0),
                distributor=str(raw_item.get("distributor") or ""),
                name=raw_item.get("name"),
                variety=raw_item.get("variety"),
                warehouse=raw_item.get("warehouse"),
                unit=raw_item.get("unit"),
                price=(float(raw_item["price"])
                       if raw_item.get("price") is not None else None),
                distributor_sku=raw_item.get("distributor_sku"),
                case_cost=(float(raw_item["case_cost"])
                           if raw_item.get("case_cost") is not None else None),
                case_size=(int(raw_item["case_size"])
                           if raw_item.get("case_size") is not None else None),
                weekly_usage=(float(raw_item["weekly_usage"])
                              if raw_item.get("weekly_usage") is not None else None),
            )
            built.append(EmailEvent(
                event_type=etype,
                item=item,
                source_message_id=str(e.get("source_message_id") or ""),
                source_subject=str(e.get("source_subject") or ""),
                po_number=str(e.get("po_number") or ""),
                po_revision=str(e.get("po_revision") or ""),
                po_order_date=str(e.get("po_order_date") or ""),
            ))
        except (TypeError, ValueError, KeyError) as exc:
            build_errors.append(f"events[{idx}]: {exc}")

    try:
        all_errors = errors + build_errors
        if canceled_skipped:
            all_errors.append(f"skipped {canceled_skipped} event(s) for canceled POs")
        report = _apply_events(
            events=built,
            messages_seen=messages_seen,
            messages_parsed=messages_parsed,
            errors=all_errors,
            dry_run=dry_run,
            source=source,
        )
    except Exception as exc:  # noqa: BLE001
        return jsonify({
            "dry_run": dry_run,
            "reports": [{
                "distributor": "Email Inbox",
                "source": source,
                "status": "error",
                "fetched": len(built),
                "updated": 0, "unchanged": 0,
                "unmatched": [], "changes": [],
                "error": str(exc),
                "messages_seen": messages_seen,
                "messages_parsed": messages_parsed,
            }],
        }), 500

    return jsonify({"dry_run": dry_run, "reports": [report]})


# ---------------------------------------------------------------------------
# Microsoft Graph change-notification subscriptions for Daily Production.
#
# These three routes turn the Daily Production pipeline into a push model:
#   - /webhooks/graph/notifications   Graph -> us. New-message pings; we
#                                     fetch + parse + ingest within seconds
#                                     of the email landing in the inbox.
#   - /api/graph/subscriptions        POST = create subs for every mailbox
#                                     in MS365_USER. GET = list local state.
#   - /api/graph/subscriptions/renew  POST = PATCH each sub to push its
#                                     expiration out ~68h. Mail subs cap at
#                                     71h; a daily Render cron calls this.
#
# The webhook deliberately lives OUTSIDE /api/ so the _gate_writes middleware
# doesn't reject it — Graph won't send our INVENTORY_API_TOKEN header.
# Instead, every notification carries a clientState (set via
# GRAPH_SUB_CLIENT_STATE) which the handler verifies.
# ---------------------------------------------------------------------------

@app.route("/webhooks/graph/notifications", methods=["POST"])
def graph_webhook_notifications():
    """Microsoft Graph -> us. Either a validation handshake or a real
    notification batch.

    Graph requires:
      - The validation handshake (a one-time GET-or-POST with
        ``?validationToken=…``) returns the token verbatim as
        ``text/plain``. Anything else fails the subscription create.
      - A regular notification POST returns 2xx within 10 seconds, or
        Graph backs off and eventually disables the subscription.
    """
    # 1. Validation handshake — Graph sends this once when a subscription
    #    is created. The body is empty; the token comes via query string.
    validation_token = request.args.get("validationToken")
    if validation_token:
        resp = make_response(validation_token, 200)
        resp.headers["Content-Type"] = "text/plain"
        return resp

    # 2. Real notification batch.
    try:
        payload = request.get_json(force=True, silent=True) or {}
    except Exception:  # noqa: BLE001
        payload = {}

    try:
        from integrations.graph_subscriptions import handle_notification
        result = handle_notification(payload)
    except Exception as exc:  # noqa: BLE001
        # Don't 500 — that makes Graph back off and eventually disable the
        # sub. Log via the response body and return 202 so Graph stays happy.
        import traceback as _tb
        return jsonify({
            "ok": False,
            "error": f"webhook handler crashed: {exc}",
            "traceback": _tb.format_exc()[-1500:],
        }), 202

    return jsonify(result), 202


@app.route("/api/graph/subscriptions", methods=["GET", "POST"])
def api_graph_subscriptions():
    """GET = list locally-tracked subscriptions.
    POST = create a subscription per mailbox in MS365_USER.

    POST returns the Graph response (id + expirationDateTime per mailbox).
    Run this once after deploying or if you ever delete the sub list.
    """
    import traceback as _tb
    try:
        from integrations.graph_subscriptions import (
            create_subscriptions, list_subscriptions,
        )
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"import failed: {exc}"}), 500

    if request.method == "GET":
        return jsonify({"ok": True, "subscriptions": list_subscriptions()})

    try:
        result = create_subscriptions()
    except Exception as exc:  # noqa: BLE001
        return jsonify({
            "ok": False, "error": str(exc),
            "traceback": _tb.format_exc()[-1500:],
        }), 500
    return jsonify(result)


@app.route("/api/graph/subscriptions/renew", methods=["POST"])
def api_graph_subscriptions_renew():
    """Renew every tracked subscription. Called daily by a Render cron.

    If a subscription is missing on Graph's side (404), this recreates it
    so coverage doesn't silently lapse.
    """
    import traceback as _tb
    try:
        from integrations.graph_subscriptions import renew_subscriptions
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"import failed: {exc}"}), 500
    try:
        result = renew_subscriptions()
    except Exception as exc:  # noqa: BLE001
        return jsonify({
            "ok": False, "error": str(exc),
            "traceback": _tb.format_exc()[-1500:],
        }), 500
    return jsonify(result)


@app.route("/api/export.xlsx")
def api_export_xlsx():
    from openpyxl import Workbook
    from export_bagels_xlsx import _write_summary_sheet, _write_items_sheet

    inv = load_inventory()
    items = list(inv.values())
    cheney = [i for i in items if (i.get("distributor") or "") == "Cheney Brothers"]
    usfoods = [i for i in items if (i.get("distributor") or "") == "US Foods"]

    wb = Workbook()
    _write_summary_sheet(wb.active, inv)
    wb.active.title = "Summary"
    _write_items_sheet(wb.create_sheet("Unified List"), items)
    _write_items_sheet(wb.create_sheet("Cheney Brothers"), cheney)
    _write_items_sheet(wb.create_sheet("US Foods"), usfoods)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="bagel_inventory.xlsx",
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000)
