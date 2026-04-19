#!/usr/bin/env python3
"""Export the unified bagel inventory to an Excel (.xlsx) workbook.

Reads items from the inventory store (populated by seed_bagels.py or the app)
and writes a multi-sheet workbook:

  - Summary         roll-up per distributor AND per warehouse
  - Unified List    every SKU with Distributor + Warehouse columns
  - Cheney Brothers only Cheney SKUs (3 FL warehouses)
  - US Foods        only US Foods SKUs (5 warehouses)

Usage:
    python export_bagels_xlsx.py [output_path]

If output_path is omitted, writes to bagel_inventory.xlsx in the cwd.
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from inventory_tracker import load_inventory


HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="E8EEF9")
SECTION_FONT = Font(bold=True, color="1F2A44")
SUBSECTION_FILL = PatternFill("solid", fgColor="F4F6FB")

HEADERS = [
    "Name", "Distributor", "Warehouse", "Category", "Quantity", "Unit",
    "Price per Unit", "Extended Value", "Case Size", "Case Cost",
    "Weekly Usage", "Weeks Remaining", "Low-Stock Threshold", "Status",
]

# Column indices (1-based) for number-format and total-row formulas.
COL_QTY = 5
COL_PRICE = 7
COL_EXTENDED = 8
COL_CASE_SIZE = 9
COL_CASE_COST = 10
COL_WEEKLY = 11
COL_WEEKS = 12
COL_STATUS = 14

# Any SKU with fewer than this many weeks of supply is flagged.
WEEKS_THRESHOLD = 4


def _weeks_remaining(item: dict):
    weekly = item.get("weekly_usage") or 0
    if weekly <= 0:
        return None
    return round(item["quantity"] / weekly, 1)


def _status(item: dict) -> str:
    if item["quantity"] <= item["low_stock_threshold"]:
        return "LOW"
    weeks = _weeks_remaining(item)
    if weeks is not None and weeks < WEEKS_THRESHOLD:
        return "SHORT (<4wk)"
    return "OK"


def _write_header(ws, row=1):
    for col, label in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = f"A{row + 1}"


def _write_row(ws, row: int, item: dict):
    extended = item["quantity"] * item["price"]
    values = [
        item["name"],
        item.get("distributor") or "Unassigned",
        item.get("warehouse") or "Unassigned",
        item.get("category", ""),
        item["quantity"],
        item["unit"],
        item["price"],
        extended,
        item.get("case_size") or "",
        item.get("case_cost") or 0,
        item.get("weekly_usage") or 0,
        _weeks_remaining(item),
        item["low_stock_threshold"],
        _status(item),
    ]
    for col, v in enumerate(values, start=1):
        ws.cell(row=row, column=col, value=v)
    ws.cell(row=row, column=COL_PRICE).number_format = '"$"#,##0.0000'
    ws.cell(row=row, column=COL_EXTENDED).number_format = '"$"#,##0.00'
    ws.cell(row=row, column=COL_CASE_COST).number_format = '"$"#,##0.00'
    ws.cell(row=row, column=COL_WEEKLY).number_format = '#,##0.0'
    ws.cell(row=row, column=COL_WEEKS).number_format = '#,##0.0'

    weeks = _weeks_remaining(item)
    if weeks is not None and weeks < WEEKS_THRESHOLD:
        ws.cell(row=row, column=COL_WEEKS).font = Font(bold=True, color="B91C1C")
    status = _status(item)
    if status == "LOW":
        ws.cell(row=row, column=COL_STATUS).font = Font(bold=True, color="B91C1C")
    elif status.startswith("SHORT"):
        ws.cell(row=row, column=COL_STATUS).font = Font(bold=True, color="B45309")


def _autosize(ws):
    widths = {
        1: 46, 2: 18, 3: 22, 4: 12, 5: 10, 6: 8, 7: 14, 8: 16,
        9: 10, 10: 12, 11: 14, 12: 15, 13: 18, 14: 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_items_sheet(ws, items: list):
    _write_header(ws, row=1)
    items = sorted(
        items,
        key=lambda x: (x.get("distributor", ""), x.get("warehouse", ""), x["name"]),
    )
    for i, item in enumerate(items, start=2):
        _write_row(ws, i, item)
    if items:
        total_row = len(items) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=COL_QTY,
                value=f"=SUM(E2:E{total_row - 1})").font = Font(bold=True)
        ws.cell(row=total_row, column=COL_EXTENDED,
                value=f"=SUM(H2:H{total_row - 1})").font = Font(bold=True)
        ws.cell(row=total_row, column=COL_EXTENDED).number_format = '"$"#,##0.00'
        ws.cell(row=total_row, column=COL_WEEKLY,
                value=f"=SUM(K2:K{total_row - 1})").font = Font(bold=True)
        ws.cell(row=total_row, column=COL_WEEKLY).number_format = '#,##0.0'
    _autosize(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(ws.max_row, 1)}"


def _write_summary_sheet(ws, inv: dict):
    # Group by distributor -> warehouse
    dist_groups: dict[str, dict[str, list]] = {}
    for item in inv.values():
        dist = item.get("distributor") or "Unassigned"
        wh = item.get("warehouse") or "Unassigned"
        dist_groups.setdefault(dist, {}).setdefault(wh, []).append(item)

    ws.cell(row=1, column=1, value="Unified Bagel Inventory").font = Font(size=16, bold=True)
    ws.cell(row=2, column=1, value="Cheney Brothers + US Foods · per-warehouse roll-up").font = Font(italic=True, color="64748B")

    headers = [
        "Distributor / Warehouse", "SKU Count", "Units on Hand", "Inventory Value",
        "Case Cost", "Weekly Usage", "Low-Stock SKUs", "SKUs <4wk",
    ]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 22
    ws.freeze_panes = "A5"

    def _count_short(items):
        n = 0
        for it in items:
            w = _weeks_remaining(it)
            if w is not None and w < WEEKS_THRESHOLD:
                n += 1
        return n

    row = 5
    grand_sku = grand_qty = grand_val = grand_weekly = grand_low = grand_short = 0
    n_cols = len(headers)

    for dist in sorted(dist_groups):
        warehouses = dist_groups[dist]
        dist_items = [i for wh_items in warehouses.values() for i in wh_items]
        d_sku = len(dist_items)
        d_qty = sum(i["quantity"] for i in dist_items)
        d_val = sum(i["quantity"] * i["price"] for i in dist_items)
        d_weekly = sum(i.get("weekly_usage") or 0 for i in dist_items)
        d_low = sum(1 for i in dist_items if i["quantity"] <= i["low_stock_threshold"])
        d_short = _count_short(dist_items)
        # All SKUs for a distributor share a flat case cost; surface it on the roll-up row.
        d_case_costs = {i.get("case_cost") for i in dist_items if i.get("case_cost")}
        d_case_cost = next(iter(d_case_costs)) if len(d_case_costs) == 1 else None

        cell = ws.cell(row=row, column=1, value=dist)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        for c in range(2, n_cols + 1):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=2, value=d_sku).font = SECTION_FONT
        ws.cell(row=row, column=3, value=d_qty).font = SECTION_FONT
        val_cell = ws.cell(row=row, column=4, value=d_val)
        val_cell.font = SECTION_FONT
        val_cell.number_format = '"$"#,##0.00'
        case_cell = ws.cell(row=row, column=5, value=d_case_cost if d_case_cost is not None else "")
        case_cell.font = SECTION_FONT
        case_cell.number_format = '"$"#,##0.00'
        weekly_cell = ws.cell(row=row, column=6, value=d_weekly)
        weekly_cell.font = SECTION_FONT
        weekly_cell.number_format = '#,##0.0'
        ws.cell(row=row, column=7, value=d_low).font = SECTION_FONT
        short_cell = ws.cell(row=row, column=8, value=d_short)
        short_cell.font = Font(bold=True, color="B91C1C") if d_short else SECTION_FONT
        grand_sku += d_sku
        grand_qty += d_qty
        grand_val += d_val
        grand_weekly += d_weekly
        grand_low += d_low
        grand_short += d_short
        row += 1

        for wh in sorted(warehouses):
            wh_items = warehouses[wh]
            sku = len(wh_items)
            qty = sum(i["quantity"] for i in wh_items)
            val = sum(i["quantity"] * i["price"] for i in wh_items)
            weekly = sum(i.get("weekly_usage") or 0 for i in wh_items)
            low = sum(1 for i in wh_items if i["quantity"] <= i["low_stock_threshold"])
            short = _count_short(wh_items)
            label = ws.cell(row=row, column=1, value=f"    {wh}")
            label.fill = SUBSECTION_FILL
            for c in range(2, n_cols + 1):
                ws.cell(row=row, column=c).fill = SUBSECTION_FILL
            ws.cell(row=row, column=2, value=sku)
            ws.cell(row=row, column=3, value=qty)
            ws.cell(row=row, column=4, value=val).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=5, value="")
            ws.cell(row=row, column=6, value=weekly).number_format = '#,##0.0'
            ws.cell(row=row, column=7, value=low)
            sc = ws.cell(row=row, column=8, value=short)
            if short:
                sc.font = Font(bold=True, color="B91C1C")
            row += 1

    # Grand total
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=grand_sku).font = Font(bold=True)
    ws.cell(row=row, column=3, value=grand_qty).font = Font(bold=True)
    gt = ws.cell(row=row, column=4, value=grand_val)
    gt.font = Font(bold=True)
    gt.number_format = '"$"#,##0.00'
    gw = ws.cell(row=row, column=6, value=grand_weekly)
    gw.font = Font(bold=True)
    gw.number_format = '#,##0.0'
    ws.cell(row=row, column=7, value=grand_low).font = Font(bold=True)
    gs = ws.cell(row=row, column=8, value=grand_short)
    gs.font = Font(bold=True, color="B91C1C") if grand_short else Font(bold=True)

    for col, w in {1: 34, 2: 12, 3: 15, 4: 18, 5: 12, 6: 14, 7: 16, 8: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def export(output: Path):
    inv = load_inventory()
    if not inv:
        print("  Inventory is empty. Run: python seed_bagels.py")
        return

    items = list(inv.values())
    cheney = [i for i in items if (i.get("distributor") or "") == "Cheney Brothers"]
    usfoods = [i for i in items if (i.get("distributor") or "") == "US Foods"]

    wb = Workbook()
    _write_summary_sheet(wb.active, inv)
    wb.active.title = "Summary"

    _write_items_sheet(wb.create_sheet("Unified List"), items)
    _write_items_sheet(wb.create_sheet("Cheney Brothers"), cheney)
    _write_items_sheet(wb.create_sheet("US Foods"), usfoods)

    wb.save(output)
    print(f"  Wrote {len(items)} SKUs to {output}")
    print(f"    Cheney Brothers: {len(cheney)} SKUs across "
          f"{len({i.get('warehouse') for i in cheney})} warehouse(s)")
    print(f"    US Foods:        {len(usfoods)} SKUs across "
          f"{len({i.get('warehouse') for i in usfoods})} warehouse(s)")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("bagel_inventory.xlsx")
    export(out)
