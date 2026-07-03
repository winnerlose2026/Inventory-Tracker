"""Tests for the Cheney per-facility inventory & usage xlsx parser."""
import io
import openpyxl
from integrations.cheney_inventory_report import (
    parse_report_xlsx, warehouse_from_filename)


def _wb(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    b = io.BytesIO()
    wb.save(b)
    return b.getvalue()


def _wb_multi(sheets):
    """sheets: list of (title, rows). First entry reuses the default sheet."""
    wb = openpyxl.Workbook()
    first = True
    for title, rows in sheets:
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        for r in rows:
            ws.append(r)
        first = False
    b = io.BytesIO()
    wb.save(b)
    return b.getvalue()


def test_warehouse_from_filename():
    assert warehouse_from_filename("H&HRVBMay272026.xlsx") == "Riviera Beach, FL"
    assert warehouse_from_filename("H&HOcalaMay272026.xlsx") == "Ocala, FL"
    assert warehouse_from_filename("H&HPuntaGordaMay272026.xlsx") == "Punta Gorda, FL"
    assert warehouse_from_filename("whatever.xlsx") == ""


def test_parse_mfg_format():
    rows = [
        ["H&H Bagels — Riviera Beach Inventory", "", "", "", ""],
        ["Cheney Item #", "Mfg#", "Description", "Cases On Hand", "Weekly Usage"],
        ["10150011", "1150", "BAGEL PLAIN PARBAKED", "48", "6"],
        ["10153019", "1152", "BAGEL POPPY PARBAKED", "12", "3"],
        ["10158022", "1158", "BAGEL EVERYTHING", "30", "5"],
        ["10199999", "9999", "BAGEL MYSTERY FLAVOR", "5", "1"],
        ["", "", "", "", ""],
    ]
    ev, err = parse_report_xlsx(_wb(rows), "H&HRVBMay272026.xlsx")
    by = {e["item"]["variety"]: e["item"] for e in ev}
    assert set(by) == {"Plain", "Poppy Seed", "Everything"}, set(by)
    assert all(it["warehouse"] == "Riviera Beach, FL" for it in by.values())
    assert by["Plain"]["quantity"] == 48 and by["Plain"]["unit"] == "cs"
    assert by["Poppy Seed"]["weekly_usage"] == 3
    assert all(e["event_type"] == "on_hand" for e in ev)
    assert any("9999" in x for x in err), err  # unmapped surfaced, not silent


def test_parse_description_only():
    rows = [
        ["Description", "On Hand", "Weekly Usage"],
        ["Plain", "20", "4"],
        ["BAGEL POPPY PARBAKED", "10", "2"],          # keyword fallback
        ["Whole Wheat Everything", "6", "1"],          # compound before "Everything"
    ]
    ev, err = parse_report_xlsx(_wb(rows), "H&HOcalaMay272026.xlsx")
    v = {e["item"]["variety"] for e in ev}
    assert v == {"Plain", "Poppy Seed", "Whole Wheat Everything"}, v
    assert all(e["item"]["warehouse"] == "Ocala, FL" for e in ev)


def test_unknown_warehouse_filename():
    ev, err = parse_report_xlsx(_wb([["Description", "On Hand"], ["Plain", "5"]]),
                                "mystery.xlsx")
    assert ev == [] and err



def test_parse_stock_inventory_format():
    """Ross's CB Direct on-hand export: Item # / Description / ... / Stock."""
    rows = [
        ["Riviera Beach"],
        ["08564 MICHAEL ROSS"],
        ["Item #", "Description", "Brand", "Pack", "Size", "UOM", "Stock"],
        ["FROZEN GROCERY"],
        ["10153048", "BAGEL EVERYTHING PARBAKED", "H & H", 1, "60CT", "cs", 184],
        ["10153018", "BAGEL PLAIN PARBAKED", "H & H", 1, "60CT", "cs", 181],
        ["10153049", "BAGEL WHOLE WHEAT EVERYTHING P", "H & H", 1, "60CT", "cs", 227],
    ]
    ev, err = parse_report_xlsx(_wb(rows), "HHBagRVB6-8-2026.xlsx")
    by = {e["item"]["variety"]: e["item"] for e in ev}
    assert set(by) == {"Everything", "Plain", "Whole Wheat Everything"}, set(by)
    assert all(e["event_type"] == "on_hand" for e in ev)
    assert by["Everything"]["quantity"] == 184 and by["Everything"]["unit"] == "cs"
    assert all("weekly_usage" not in it for it in by.values())  # stock sheet carries no usage
    assert all(it["warehouse"] == "Riviera Beach, FL" for it in by.values())
    print("OK test_parse_stock_inventory_format")


def test_parse_case_movement_usage():
    """Ross's monthly case-movement export -> usage_rate (monthly -> weekly)."""
    rows = [
        ["Report Creation Date : 6/8/2026"],
        ["Drill Down Reporting : Date Range >= 05/01/2026 AND <= 05/31/2026"],
        ["DSRGroup =MICHAEL ROSS-8564, DC =01 RIVIERA - 01 RIVIERA, Brands =H & H,"],
        ["Products", "Pack", "Dist Item #", "Mfq.Product Code", "GTIN", "Full Cases"],
        ["Sum of All Products Activity", "", "", "", "", 1147],
        ["BAGEL EVERYTHING PARBAKED", "1:60 CT", "10153048", "1158", "", 258],
        ["BAGEL PLAIN PARBAKED", "1:60 CT", "10153018", "1150", "", 250],
        ["BAGEL WHOLE WHEAT PARBAKED", "1:60 CT", "10153042", "1156", "", 50],
    ]
    ev, err = parse_report_xlsx(_wb(rows), "HHBagRVBMay2026.xlsx")
    assert ev and all(e["event_type"] == "usage_rate" for e in ev), [e["event_type"] for e in ev]
    wk = {e["item"]["variety"]: e["item"]["weekly_usage"] for e in ev}
    # 258 * 7 / 31 = 58.26 ; 250 -> 56.45 ; 50 -> 11.29 ; Sum row skipped.
    assert wk == {"Everything": 58.26, "Plain": 56.45, "Whole Wheat": 11.29}, wk
    assert all(e["item"]["warehouse"] == "Riviera Beach, FL" for e in ev)
    assert all(e["item"]["quantity"] == 0.0 for e in ev)  # rate-only; no on-hand
    assert err == [], err  # Sum-of-activity total carries no code -> skipped silently
    print("OK test_parse_case_movement_usage")


def test_combined_usage_and_stock_workbook():
    """Ross's current 'Usage&Stock' export: the usage (case-movement) grid AND
    the on-hand stock grid live in ONE workbook on separate sheets. BOTH must
    parse, and every event must carry the report's END date (6/27) as
    count_date -- not the range start (6/22) or the email/scan time."""
    usage = [
        ["Report Creation Date : 6/28/2026"],
        ["Drill Down Reporting : Date Range >= 06/22/2026 AND <= 06/27/2026"],
        ["DSRGroup =MICHAEL ROSS-8564, DC =01 RIVIERA,"],
        ["Products", "Pack", "Dist Item #", "Mfq.Product Code", "GTIN", "Full Cases"],
        ["Sum of All Products Activity", "", "", "", "", 400],
        ["BAGEL EVERYTHING PARBAKED", "1:60 CT", "10153048", "1158", "", 56],
        ["BAGEL PLAIN PARBAKED", "1:60 CT", "10153018", "1150", "", 63],
    ]
    stock = [
        ["Riviera Beach"],
        ["08564 MICHAEL ROSS"],
        ["Item #", "Description", "Brand", "Pack", "Size", "UOM", "Stock"],
        ["FROZEN GROCERY"],
        ["10153048", "BAGEL EVERYTHING PARBAKED", "H & H", 1, "60CT", "cs", 184],
        ["10153018", "BAGEL PLAIN PARBAKED", "H & H", 1, "60CT", "cs", 181],
    ]
    b = _wb_multi([("Usage", usage), ("Stock", stock)])
    ev, err = parse_report_xlsx(
        b, "HHRivieraBeachUsage&StockJune22-June272026.xlsx")
    types = sorted({e["event_type"] for e in ev})
    assert types == ["on_hand", "usage_rate"], (types, err)
    oh = {e["item"]["variety"]: e["item"]["quantity"]
          for e in ev if e["event_type"] == "on_hand"}
    assert oh == {"Everything": 184, "Plain": 181}, (oh, err)
    ur = {e["item"]["variety"] for e in ev if e["event_type"] == "usage_rate"}
    assert ur == {"Everything", "Plain"}, (ur, err)
    assert all(e.get("count_date") == "2026-06-27" for e in ev), (
        [e.get("count_date") for e in ev])
    assert all(it["warehouse"] == "Riviera Beach, FL"
               for it in (e["item"] for e in ev))
    print("OK test_combined_usage_and_stock_workbook")


def test_combined_usage_and_stock_one_sheet():
    """Ross's 'Usage&Stock' export with the usage grid and the on-hand grid
    STACKED on a single sheet. Both must parse (usage_rate + on_hand) and every
    event must carry the report end date (6/27)."""
    rows = [
        ["Report Creation Date : 6/28/2026"],
        ["Drill Down Reporting : Date Range >= 06/22/2026 AND <= 06/27/2026"],
        ["Products", "Pack", "Dist Item #", "Mfq.Product Code", "GTIN", "Full Cases"],
        ["Sum of All Products Activity", "", "", "", "", 400],
        ["BAGEL EVERYTHING PARBAKED", "1:60 CT", "10153048", "1158", "", 56],
        ["BAGEL PLAIN PARBAKED", "1:60 CT", "10153018", "1150", "", 63],
        ["", "", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["Item #", "Description", "Brand", "Pack", "Size", "UOM", "Stock"],
        ["10153048", "BAGEL EVERYTHING PARBAKED", "H & H", 1, "60CT", "cs", 184],
        ["10153018", "BAGEL PLAIN PARBAKED", "H & H", 1, "60CT", "cs", 181],
    ]
    ev, err = parse_report_xlsx(_wb(rows), "HHRivieraBeachUsage&StockJune22-June272026.xlsx")
    oh = {e["item"]["variety"]: e["item"]["quantity"]
          for e in ev if e["event_type"] == "on_hand"}
    ur = {e["item"]["variety"] for e in ev if e["event_type"] == "usage_rate"}
    assert oh == {"Everything": 184, "Plain": 181}, (oh, err)
    assert ur == {"Everything", "Plain"}, (ur, err)
    assert all(e.get("count_date") == "2026-06-27" for e in ev), (
        [e.get("count_date") for e in ev])
    print("OK test_combined_usage_and_stock_one_sheet")


def test_count_date_from_filename_when_no_range():
    """A stock-only sheet has no Date Range line -> fall back to the filename
    date (here 6-8-2026)."""
    rows = [
        ["Item #", "Description", "Brand", "Pack", "Size", "UOM", "Stock"],
        ["10153018", "BAGEL PLAIN PARBAKED", "H & H", 1, "60CT", "cs", 100],
    ]
    ev, err = parse_report_xlsx(_wb(rows), "HHBagRVB6-8-2026.xlsx")
    assert ev, err
    assert all(e.get("count_date") == "2026-06-08" for e in ev), (
        [e.get("count_date") for e in ev], err)
    print("OK test_count_date_from_filename_when_no_range")


def test_extract_stock_image():
    """A Cheney file with an embedded stock image -> extract_stock_image returns
    that image bytes, the warehouse, and the Date Range end as count_date."""
    import io as _io, tempfile, os
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    from integrations.cheney_inventory_report import extract_stock_image
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "Drill Down Reporting : Date Range >= 06/22/2026 AND <= 06/27/2026"
    ws["A2"] = "Products"; ws["E2"] = "Full Cases"; ws["E3"] = 100
    tmp = os.path.join(tempfile.gettempdir(), "_stockimg.png")
    PILImage.new("RGB", (40, 20), (123, 200, 50)).save(tmp)
    ws.add_image(XLImage(tmp), "H2")
    b = _io.BytesIO(); wb.save(b)
    wh, cd, img, ctype = extract_stock_image(b.getvalue(), "HHOcalaUsage&Stock.xlsx")
    assert wh == "Ocala, FL", wh
    assert cd == "2026-06-27", cd
    assert img and len(img) > 0 and ctype == "image/png", (bool(img), ctype)
    print("OK test_extract_stock_image")


def test_extract_stock_image_none_when_no_image():
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "no image here"
    import io as _io
    from integrations.cheney_inventory_report import extract_stock_image
    b = _io.BytesIO(); wb.save(b)
    wh, cd, img, ctype = extract_stock_image(b.getvalue(), "HHRVBUsage&Stock.xlsx")
    assert wh == "Riviera Beach, FL" and img is None, (wh, bool(img))
    print("OK test_extract_stock_image_none_when_no_image")


if __name__ == "__main__":
    test_warehouse_from_filename()
    test_parse_mfg_format()
    test_parse_description_only()
    test_unknown_warehouse_filename()
    test_parse_stock_inventory_format()
    test_parse_case_movement_usage()
    test_combined_usage_and_stock_workbook()
    test_combined_usage_and_stock_one_sheet()
    test_count_date_from_filename_when_no_range()
    test_extract_stock_image()
    test_extract_stock_image_none_when_no_image()
    print("ALL CHENEY PARSER TESTS PASSED")
