"""Reporting blueprint — sales / Toast ingest + reports, traceability search,
PLH (productivity vs labor-hours), bakery sales, the forecast endpoints
(decrement-daily / backfill-historical / true-up), the summary report +
report-status pages, and the xlsx export. Extracted from app.py (refactor —
see REFACTOR_PLAN.md). Shared helpers come from core/."""
import io
from datetime import datetime, timedelta

from flask import Blueprint, jsonify, request, send_file

from core.errors import _safe_err
from inventory_tracker import (
    load_inventory, load_usage, save_inventory, save_usage,
)

reporting_bp = Blueprint("reporting", __name__)


CASE_PRICE_BY_DISTRIBUTOR = {
    "US Foods":        27.00,
    "Cheney Brothers": 26.50,
}
CASE_PRICE_DEFAULT = 29.50
LABOR_RATE_DEFAULT = 17.00


def _case_price_for(distributor: str) -> float:
    return CASE_PRICE_BY_DISTRIBUTOR.get((distributor or "").strip(),
                                         CASE_PRICE_DEFAULT)


def _plh_bucket_keys(grain: str, offset: int = 0):
    """Return [(key, start_iso, end_iso, label)] for the buckets in scope.

    grain = "week"     -> 4 ISO weeks ending this week (Mon-Sun)
    grain = "month"    -> 4 trailing calendar months ending this month
    grain = "quarter"  -> the 4 quarters of a calendar year

    offset shifts the window backward by:
      week   -> 1 week per step  (4-week window slides one week at a time)
      month  -> 1 month per step (4-month window slides one month at a time)
      quarter-> 1 year per step  (4-quarter window jumps a full year)
    offset = 0 is the current window. Negative offsets are clamped to 0
    (no "future" windows past now).
    """
    from datetime import datetime, timedelta
    today = datetime.now().date()
    if offset < 0:
        offset = 0

    if grain == "month":
        # 4 trailing calendar months ending with the current month. Each
        # prev/next click steps the window back by 1 month (overlapping
        # navigation rather than jumping the whole window at once).
        cur_abs = today.year * 12 + (today.month - 1)
        # The newest month in this window:
        newest_abs = cur_abs - offset
        # Build oldest -> newest so the chart reads left to right naturally.
        out = []
        for i in range(3, -1, -1):
            abs_i = newest_abs - i
            yy = abs_i // 12
            mm = (abs_i % 12) + 1
            start = datetime(yy, mm, 1).date()
            nxt_abs = abs_i + 1
            n_yy = nxt_abs // 12
            n_mm = (nxt_abs % 12) + 1
            end = (datetime(n_yy, n_mm, 1).date() - timedelta(days=1))
            key   = start.strftime("%Y-%m")
            label = start.strftime("%b %Y")
            out.append((key, start.isoformat(), end.isoformat(), label))
        return out

    if grain == "quarter":
        anchor_year = today.year - offset
        out = []
        for q in range(1, 5):
            start_mm = (q - 1) * 3 + 1
            start = datetime(anchor_year, start_mm, 1).date()
            end_mm = start_mm + 3
            end_yy = anchor_year
            if end_mm > 12:
                end_mm = 1
                end_yy = anchor_year + 1
            end = (datetime(end_yy, end_mm, 1).date() - timedelta(days=1))
            key   = f"{anchor_year}-Q{q}"
            label = f"Q{q} {anchor_year}"
            out.append((key, start.isoformat(), end.isoformat(), label))
        return out

    # default: weekly. 4 ISO weeks anchored on the Monday `offset` weeks
    # back from this week (each prev/next click shifts the window by 1 week,
    # so navigation is smooth and overlapping rather than jumping a full
    # 4-week window at a time).
    day = today.weekday()        # Mon=0..Sun=6
    this_monday = today - timedelta(days=day)
    anchor_monday = this_monday - timedelta(days=7 * offset)
    out = []
    for i in range(3, -1, -1):    # 3 windows ago -> latest
        wk_start = anchor_monday - timedelta(days=7 * i)
        wk_end   = wk_start + timedelta(days=6)
        iso = wk_start.isocalendar()
        key   = f"{iso.year:04d}-W{iso.week:02d}"
        label = f"Week of {wk_start.strftime('%b %d')}"
        out.append((key, wk_start.isoformat(), wk_end.isoformat(), label))
    return out


def _plh_window_label(grain: str, buckets: list) -> str:
    """Human label for the chart header — e.g. 'Apr 20 – May 17, 2026',
    'Apr – Jun 2026', or 'Q1 – Q4 2026'."""
    if not buckets:
        return ""
    first_start = buckets[0][1]
    last_end    = buckets[-1][2]
    from datetime import date
    fs = date.fromisoformat(first_start)
    le = date.fromisoformat(last_end)
    if grain == "quarter":
        # Always within one year for this view.
        return f"Q1 – Q4 {fs.year}"
    if grain == "month":
        if fs.year == le.year:
            return f"{fs.strftime('%b')} – {le.strftime('%b')} {fs.year}"
        return f"{fs.strftime('%b %Y')} – {le.strftime('%b %Y')}"
    # week
    if fs.year == le.year:
        return f"{fs.strftime('%b %d')} – {le.strftime('%b %d, %Y')}"
    return f"{fs.strftime('%b %d, %Y')} – {le.strftime('%b %d, %Y')}"


def _date_in_range(d: str, start: str, end: str) -> bool:
    if not d:
        return False
    d = d[:10]
    return start <= d <= end


def _bucket_is_empty(start: str, end: str, prod, labor, bsales) -> bool:
    """Cheap check: does any production / labor / bakery-sales row fall in
    the [start, end] window? Used to skip 'we have not gotten there yet'
    periods when anchoring the chart's default view."""
    for r in prod:
        if _date_in_range(r.get("production_date") or "", start, end):
            return False
    for e in labor:
        if _date_in_range(e.get("date") or "", start, end):
            if (e.get("hours") or 0) > 0 or (e.get("dollars") or 0) > 0:
                return False
    for w in bsales:
        if _date_in_range(w.get("week_start") or "", start, end):
            if (w.get("total") or 0) > 0:
                return False
    return True


def datetime_now_iso() -> str:
    from datetime import datetime
    return datetime.now().isoformat()


@reporting_bp.route("/api/sales/ingest", methods=["POST"])
def api_sales_ingest():
    """Accept Toast product-mix rows pushed by an external aggregator.

    Body: {
      "rows": [
        {"restaurant_guid": "...", "location": "UES",
         "business_date": "2026-05-12",
         "item_guid": "...", "item": "Plain Bagel",
         "menu_group": "Bagels",          # optional
         "qty": 124, "gross": 372.00, "net": 369.50}
      ],
      "replace_dates": false   # if true, wipe existing rows for the
                               # (location, date) pairs in this batch
                               # before appending. Useful to refresh a
                               # day without orphaning items removed
                               # from the catalog.
    }

    Dedupe key: (restaurant_guid, business_date, item_guid). Posting the
    same item for the same day overwrites the prior row.
    """
    from inventory_tracker import load_sales, save_sales
    body = request.json or {}
    rows = body.get("rows") or []
    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "rows must be a list"}), 400
    replace_dates = bool(body.get("replace_dates", False))

    existing = load_sales()
    # Key existing rows.
    by_key = {(r.get("restaurant_guid"), r.get("business_date"), r.get("item_guid")): i
              for i, r in enumerate(existing)}

    if replace_dates:
        seed = set((r.get("restaurant_guid"), r.get("business_date")) for r in rows)
        existing = [r for r in existing
                    if (r.get("restaurant_guid"), r.get("business_date")) not in seed]
        by_key = {(r.get("restaurant_guid"), r.get("business_date"), r.get("item_guid")): i
                  for i, r in enumerate(existing)}

    added, updated = 0, 0
    for r in rows:
        if not r.get("restaurant_guid") or not r.get("business_date") or not r.get("item_guid"):
            continue
        key = (r.get("restaurant_guid"), r.get("business_date"), r.get("item_guid"))
        if key in by_key:
            existing[by_key[key]] = r
            updated += 1
        else:
            existing.append(r)
            by_key[key] = len(existing) - 1
            added += 1
    save_sales(existing)
    return jsonify({"ok": True, "added": added, "updated": updated,
                    "total_rows": len(existing)})


@reporting_bp.route("/api/report/toast-sales")
def api_report_toast_sales():
    """Top-selling items aggregated by week or month, optionally by
    location. Output is the data behind the Report -> Top Consumed
    section.

    Query params:
      period        'week' | 'month' (default 'week')
      buckets       int, default 8 (weeks) / 6 (months)
      location      restaurant_guid OR location string, optional
      top_n         int, default 10
    """
    from inventory_tracker import load_sales
    from datetime import datetime as _dt, timedelta as _td

    args = request.args
    period = (args.get("period") or "week").lower()
    if period not in ("week", "month"):
        return jsonify({"ok": False, "error": "period must be week|month"}), 400
    try:
        buckets = int(args.get("buckets") or (8 if period == "week" else 6))
    except (TypeError, ValueError):
        buckets = 8 if period == "week" else 6
    try:
        top_n = int(args.get("top_n") or 10)
    except (TypeError, ValueError):
        top_n = 10
    location_q = (args.get("location") or "").strip()
    # Anchor date — the Report page's Prev/Next/Today buttons and the
    # date picker all set this to a single YYYY-MM-DD. When present we
    # return the SINGLE week/month bucket containing that date so the
    # user sees the period they picked (matching the Restock card's
    # one-window-at-a-time semantics). When absent we fall back to
    # returning the most recent N buckets that have data.
    end_date_q = (args.get("end_date") or "").strip()
    # When True (default) a cache miss for the picked bucket triggers a
    # live Toast pull. The client can pass live=0 to force cache-only
    # behavior (e.g. for fast retries while a fetch is in flight).
    live_q = (args.get("live") or "1").strip() not in ("0", "false", "no")

    # Bucket each row.
    def _bucket_key(date_iso: str) -> str:
        try:
            d = _dt.strptime(date_iso, "%Y-%m-%d")
        except ValueError:
            return ""
        if period == "week":
            # ISO week, Monday-start. Anchor on the Monday of that week.
            mon = d - _td(days=d.weekday())
            return mon.strftime("%Y-%m-%d")
        return d.strftime("%Y-%m")

    def _bucket_range(anchor_iso: str):
        """Return (start_date_iso, end_date_iso) inclusive for the
        bucket containing anchor_iso. Caps end at today — we never
        ask Toast for future dates."""
        ed = _dt.strptime(anchor_iso, "%Y-%m-%d")
        if period == "week":
            start = ed - _td(days=ed.weekday())
            end   = start + _td(days=6)
        else:
            start = ed.replace(day=1)
            if ed.month == 12:
                end = ed.replace(day=31)
            else:
                end = ed.replace(month=ed.month + 1, day=1) - _td(days=1)
        today = _dt.now().date()
        if end.date() > today:
            end = _dt.combine(today, _dt.min.time())
        return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")

    def _target_locations(loc_q: str):
        """Return [(restaurant_guid, location_name), ...] for the live
        fetch. If a specific location is selected, returns just that
        one. Otherwise returns every active retail location."""
        from inventory_tracker import TOAST_RETAIL_LOCATIONS
        if loc_q:
            if "-" in loc_q or len(loc_q) == 36:
                for L in TOAST_RETAIL_LOCATIONS:
                    if L["restaurant_guid"] == loc_q:
                        return [(L["restaurant_guid"], L["location"])]
                return [(loc_q, "")]
            q = loc_q.lower()
            return [(L["restaurant_guid"], L["location"])
                    for L in TOAST_RETAIL_LOCATIONS
                    if q in (L.get("location") or "").lower()]
        return [(L["restaurant_guid"], L["location"])
                for L in TOAST_RETAIL_LOCATIONS
                if (L.get("status") or "active") == "active"]

    def _ensure_bucket_cached(all_rows: list, anchor_iso: str, loc_q: str):
        """If the bucket containing anchor_iso has no rows for the
        target locations, pull from Toast and append to sales.json.
        Returns (updated_rows, meta dict) where meta carries fetch info
        for the client (rows_fetched, fetch_error, fetched_at)."""
        meta = {"rows_fetched": 0, "fetch_error": None, "fetched_at": None}
        if not live_q:
            return all_rows, meta
        try:
            start_iso, end_iso = _bucket_range(anchor_iso)
        except ValueError:
            return all_rows, meta
        targets = _target_locations(loc_q)
        if not targets:
            return all_rows, meta
        # Which (guid, date) pairs already have at least 1 row?
        cached_keys = set()
        for r in all_rows:
            d = (r.get("business_date") or "").strip()
            if start_iso <= d <= end_iso:
                cached_keys.add((r.get("restaurant_guid") or "", d))
        # Enumerate every (target_location, date) pair in the bucket
        # range; anything not in cached_keys is a fetch candidate.
        missing = []
        s = _dt.strptime(start_iso, "%Y-%m-%d")
        e = _dt.strptime(end_iso,   "%Y-%m-%d")
        d_cursor = s
        while d_cursor <= e:
            d_iso = d_cursor.strftime("%Y-%m-%d")
            for (guid, name) in targets:
                if (guid, d_iso) not in cached_keys:
                    missing.append((guid, d_iso, name))
            d_cursor += _td(days=1)
        if not missing:
            return all_rows, meta
        try:
            from integrations import toast_api
            if not toast_api.is_configured():
                meta["fetch_error"] = "toast_not_configured"
                return all_rows, meta
            new_rows = toast_api.fetch_product_mix_batch(
                missing, max_workers=6, timeout_s=28.0)
        except Exception as ex:
            meta["fetch_error"] = _safe_err(ex, "toast fetch")
            return all_rows, meta
        if not new_rows:
            # No orders for any of the missing (guid, date) pairs is a
            # legitimate response — still record the attempt so the UI
            # knows a fetch ran.
            meta["fetched_at"] = _dt.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
            return all_rows, meta
        from inventory_tracker import save_sales
        # Dedupe on (restaurant_guid, business_date, item_guid) — if a
        # date was partially cached, prefer the fresh row.
        by_key: dict = {}
        for r in all_rows:
            k = (r.get("restaurant_guid"), r.get("business_date"),
                 r.get("item_guid"))
            by_key[k] = r
        for r in new_rows:
            k = (r.get("restaurant_guid"), r.get("business_date"),
                 r.get("item_guid"))
            by_key[k] = r
        merged = list(by_key.values())
        save_sales(merged)
        meta["rows_fetched"] = len(new_rows)
        meta["fetched_at"]   = _dt.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        return merged, meta

    # Load the full row set ONCE. If an anchor date is set we may
    # mutate this with a Toast fetch before bucketing.
    rows = load_sales() or []
    fetch_meta = {"rows_fetched": 0, "fetch_error": None, "fetched_at": None}
    if end_date_q:
        try:
            rows, fetch_meta = _ensure_bucket_cached(rows, end_date_q,
                                                    location_q)
        except Exception as ex:
            fetch_meta["fetch_error"] = _safe_err(ex, "toast fetch")

    # Apply the location filter (after the live fetch so newly-pulled
    # rows are visible).
    if location_q:
        if "-" in location_q or len(location_q) == 36:
            rows = [r for r in rows if r.get("restaurant_guid") == location_q]
        else:
            q = location_q.lower()
            rows = [r for r in rows if q in (r.get("location") or "").lower()]

    if not rows and not end_date_q:
        return jsonify({"ok": True, "period": period, "buckets": [],
                        "total_rows": 0, "fetch": fetch_meta})

    # Aggregate by NORMALIZED display name (+ menu group) rather than
    # item_guid. Toast issues a fresh item_guid every time a menu item
    # is re-published (price change, name tweak, modifier shuffle), so
    # the same user-visible product can appear under several guids in
    # a single week. Keying on guid surfaces these as duplicate rows;
    # keying on name collapses them. The underlying sales.json still
    # keeps the raw guid for traceability.
    def _norm(s: str) -> str:
        return " ".join((s or "").lower().split())

    buckets_map: dict = {}
    for r in rows:
        bk = _bucket_key(r.get("business_date") or "")
        if not bk:
            continue
        slot = buckets_map.setdefault(bk, {})
        name      = (r.get("item") or "").strip()
        menu_grp  = (r.get("menu_group") or "").strip()
        if not name and not r.get("item_guid"):
            continue
        item_key  = (_norm(name), _norm(menu_grp))
        agg = slot.setdefault(item_key, {
            "item":       name,
            "menu_group": menu_grp,
            "qty":        0,
            "gross":      0.0,
            "net":        0.0,
            "guid_count": 0,
        })
        agg["qty"]        += int(r.get("qty") or 0)
        agg["gross"]      += float(r.get("gross") or 0)
        agg["net"]        += float(r.get("net") or 0)
        agg["guid_count"] += 1  # informational; not used in render
        # Prefer the longest/non-empty observed values for display.
        if name and len(name) > len(agg["item"]):
            agg["item"] = name
        if menu_grp and not agg["menu_group"]:
            agg["menu_group"] = menu_grp

    def _label_for(bk: str) -> str:
        if period == "week":
            d = _dt.strptime(bk, "%Y-%m-%d")
            return f"{d.strftime('%b %-d')} \u2013 {(d + _td(days=6)).strftime('%b %-d, %Y')}"
        d = _dt.strptime(bk, "%Y-%m")
        return d.strftime("%B %Y")

    def _serialize(bk: str) -> dict:
        items = list(buckets_map.get(bk, {}).values())
        items.sort(key=lambda x: x["gross"], reverse=True)
        total = sum(x["gross"] for x in items) or 1
        for it in items:
            it["mix_pct"] = round(100 * it["gross"] / total, 2)
            it["gross"]   = round(it["gross"], 2)
            it["net"]     = round(it["net"], 2)
        return {
            "key":         bk,
            "label":       _label_for(bk),
            "total_gross": round(total, 2) if items else 0.0,
            "items":       items[:top_n],
            "item_count":  len(items),
        }

    # When a specific date is picked, return ONLY the bucket containing
    # that date \u2014 even if it has no rows, so the UI can render a clear
    # "no data for week of X" state instead of falling silently empty.
    if end_date_q:
        try:
            ed = _dt.strptime(end_date_q, "%Y-%m-%d")
            if period == "week":
                anchor = (ed - _td(days=ed.weekday())).strftime("%Y-%m-%d")
            else:
                anchor = ed.strftime("%Y-%m")
            return jsonify({"ok": True, "period": period,
                            "buckets":     [_serialize(anchor)],
                            "anchor":      anchor,
                            "anchor_date": end_date_q,
                            "total_rows":  len(rows),
                            "fetch":       fetch_meta})
        except ValueError:
            pass

    # No anchor date \u2014 fall back to the most recent N buckets with data.
    bucket_keys = sorted(buckets_map.keys(), reverse=True)
    out_buckets = [_serialize(bk) for bk in bucket_keys[:buckets]]

    return jsonify({"ok": True, "period": period,
                    "buckets": out_buckets,
                    "total_rows": len(rows)})


@reporting_bp.route("/api/sales/locations")
def api_sales_locations():
    """List every retail Toast location in the registry, with row
    counts and date ranges from the sales store. Locations with no
    ingested data return rows=0 — they still show in the dropdown so
    the user can pick any location for comparison.
    """
    from inventory_tracker import load_sales, TOAST_RETAIL_LOCATIONS
    rows = load_sales() or []
    # Start with the registry — every retail location guaranteed, with
    # state abbreviation so the UI can render "Penn Station, NY" without
    # any per-location lookup.
    by_loc: dict = {}
    for L in TOAST_RETAIL_LOCATIONS:
        by_loc[L["restaurant_guid"]] = {
            "restaurant_guid": L["restaurant_guid"],
            "location":        L["location"],
            "state":           L.get("state") or "",
            "status":          L.get("status") or "active",
            "rows":            0,
            "min_date":        "",
            "max_date":        "",
        }
    # Layer in any ingested data — increment counts + date range.
    for r in rows:
        guid = r.get("restaurant_guid") or ""
        if not guid:
            continue
        slot = by_loc.setdefault(guid, {
            "restaurant_guid": guid,
            "location":        r.get("location") or "",
            "state":           "",
            "rows":            0,
            "min_date":        r.get("business_date") or "",
            "max_date":        r.get("business_date") or "",
        })
        slot["rows"] += 1
        d = r.get("business_date") or ""
        if d:
            if not slot["min_date"] or d < slot["min_date"]:
                slot["min_date"] = d
            if not slot["max_date"] or d > slot["max_date"]:
                slot["max_date"] = d
    return jsonify({"ok": True,
                    "locations": sorted(by_loc.values(),
                                        key=lambda x: x["location"])})


@reporting_bp.route("/api/traceability/search")
def api_traceability_search():
    """Cross-source traceability search.

    Query params:
      ``from``        ISO date — only include records on/after this date.
      ``to``          ISO date — only include records on/before this date.
      ``lot``         Substring match on lot_number (case-insensitive).
      ``mfg_code``    4-digit item code (matches the first 4 digits of
                      any lot OR the variety via HH_MFG_CODE_TO_VARIETY).
      ``warehouse``   Exact warehouse name.
      ``distributor`` Exact distributor name.

    Returns a list of matching production records, each with these
    downstream fields populated for the requesting line:
      ``po_status``      pending | arrived | canceled | unknown
      ``ship_date``      from the on_order entry, if pending.
      ``arrival_date``   from the on_order entry, if pending.
      ``eta``            from the on_order entry, if pending.
      ``lines[].on_hand_now``  current cs on hand for that variety at
                                that warehouse (live inventory).
      ``lines[].usage_total_cs`` total cs moved on usage log for this
                                PO + variety (from usage.json).
      ``lines[].usage_event_count``  count of usage events.
    """
    from inventory_tracker import (
        load_production, load_inventory, load_usage,
        load_canceled_pos,
    )
    try:
        from integrations.hh_mfg_codes import HH_MFG_CODE_TO_VARIETY
    except Exception:
        HH_MFG_CODE_TO_VARIETY = {}

    args = request.args
    q_from = (args.get("from") or "").strip()
    q_to   = (args.get("to") or "").strip()
    q_lot  = (args.get("lot") or "").strip().lower()
    q_code = (args.get("mfg_code") or "").strip()
    q_wh   = (args.get("warehouse") or "").strip()
    q_dist = (args.get("distributor") or "").strip()

    records = load_production()

    # Apply filters.
    out = []
    code_variety = HH_MFG_CODE_TO_VARIETY.get(q_code) if q_code else None
    for r in records:
        if q_from and (r.get("production_date") or "") < q_from:
            continue
        if q_to and (r.get("production_date") or "") > q_to:
            continue
        if q_wh and (r.get("warehouse") or "") != q_wh:
            continue
        if q_dist and (r.get("distributor") or "") != q_dist:
            continue
        # Lot / mfg-code filter only keeps the record if at least one
        # line matches. We do NOT prune the other lines from the record
        # — operators usually want to see the full sheet for context.
        if q_lot or q_code:
            matched_any = False
            for L in r.get("lines", []):
                lot = (L.get("lot_number") or "")
                if q_lot and q_lot in lot.lower():
                    matched_any = True
                    break
                if q_code:
                    # Match either the first 4 digits of the lot, or the
                    # variety mapped from the code.
                    if lot.startswith(q_code):
                        matched_any = True
                        break
                    if code_variety and (L.get("variety") or "") == code_variety:
                        matched_any = True
                        break
            if not matched_any:
                continue
        out.append(r)

    # Build downstream lookups.
    inv = load_inventory()
    # {(warehouse, variety) -> on_hand_qty} for currently-on-hand snapshot
    on_hand = {}
    for item in inv.values():
        wh = item.get("warehouse") or ""
        name = item.get("name") or ""
        variety = name.split(" Bagel")[0] if " Bagel" in name else name
        on_hand[(wh, variety)] = (on_hand.get((wh, variety), 0)
                                  + float(item.get("quantity") or 0))

    # Pending-PO lookup keyed by po_number — first non-empty wins.
    # Inventory on_order is checked first (USF + Cheney); then Chefs
    # Warehouse POs (which live in their own file because they don't
    # touch inventory). A Daily Production sheet stamped with a CW PO#
    # therefore still resolves to "pending" with the right ship/arrival
    # data.
    po_pending = {}
    for item in inv.values():
        for e in item.get("on_order") or []:
            po = (e.get("po_number") or "").strip()
            if not po:
                continue
            slot = po_pending.setdefault(po, {})
            for k in ("ship_date", "arrival_date", "eta"):
                if not slot.get(k) and e.get(k):
                    slot[k] = e[k]
    try:
        from inventory_tracker import load_chefs_warehouse_pos
        for cw in load_chefs_warehouse_pos():
            po = (cw.get("po_number") or "").strip()
            if not po or cw.get("canceled"):
                continue
            slot = po_pending.setdefault(po, {})
            for k in ("ship_date", "arrival_date", "eta"):
                if not slot.get(k) and cw.get(k):
                    slot[k] = cw[k]
    except Exception:  # noqa: BLE001
        # CW POs are auxiliary; failure here must not break traceability.
        pass

    try:
        canceled = load_canceled_pos()
    except Exception:
        canceled = {}

    # Usage rollup keyed by (po_number, variety).
    usage_log = load_usage() or []
    usage_roll: dict = {}  # {(po, variety): {cs, n}}
    for e in usage_log:
        po = (e.get("po_number") or "").strip()
        if not po:
            continue
        name = e.get("item_name") or ""
        variety = name.split(" Bagel")[0] if " Bagel" in name else name
        slot = usage_roll.setdefault((po, variety), {"cs": 0.0, "n": 0})
        slot["cs"] += abs(float(e.get("amount") or 0))
        slot["n"] += 1

    # Enrich each record.
    enriched = []
    for r in out:
        rcopy = dict(r)
        po = (r.get("po_number") or "").strip()
        wh = r.get("warehouse") or ""
        pend = po_pending.get(po) if po else None
        if po and po in canceled:
            rcopy["po_status"] = "canceled"
        elif pend:
            rcopy["po_status"]    = "pending"
            rcopy["ship_date"]    = pend.get("ship_date") or ""
            rcopy["arrival_date"] = pend.get("arrival_date") or ""
            rcopy["eta"]          = pend.get("eta") or ""
        elif po:
            # PO existed but isn't in on_order anymore -> rolled over.
            rcopy["po_status"] = "arrived"
        else:
            rcopy["po_status"] = "unknown"

        enriched_lines = []
        for L in r.get("lines") or []:
            v = L.get("variety") or ""
            lcopy = dict(L)
            lcopy["on_hand_now"] = round(on_hand.get((wh, v), 0), 2)
            ur = usage_roll.get((po, v)) if po else None
            lcopy["usage_total_cs"]    = round(ur["cs"], 2) if ur else 0
            lcopy["usage_event_count"] = ur["n"] if ur else 0
            enriched_lines.append(lcopy)
        rcopy["lines"] = enriched_lines
        enriched.append(rcopy)

    # Sort newest-first by production date for stable grouping.
    enriched.sort(key=lambda r: (r.get("production_date") or "",
                                  r.get("po_number") or ""),
                  reverse=True)
    return jsonify({
        "ok": True,
        "filters": {
            "from": q_from, "to": q_to, "lot": q_lot, "mfg_code": q_code,
            "warehouse": q_wh, "distributor": q_dist,
        },
        "count": len(enriched),
        "records": enriched,
        "mfg_code_map": HH_MFG_CODE_TO_VARIETY,
    })


@reporting_bp.route("/api/report/plh")
def api_report_plh():
    """Production revenue, labor hours, and $PLH per time bucket.

    Query params:
      grain   "week" (default) | "month" | "quarter"
      offset  int >= 0. 0 = current window. Each step shifts back one
              full window: 4 weeks / 1 quarter / 1 year.
    """
    from inventory_tracker import load_production, load_labor, load_bakery_sales
    grain = (request.args.get("grain") or "week").lower()
    try:
        offset = max(0, int(request.args.get("offset") or 0))
    except (TypeError, ValueError):
        offset = 0
    prod  = load_production()
    labor = load_labor()
    bsales = load_bakery_sales()

    # For week/month grain, anchor the view at the most recent POPULATED
    # period instead of "today." If the current week or month hasn't been
    # filled in yet (the bakery model spreadsheet is updated weekly), the
    # auto_shift slides the window back so we still show four populated
    # bars instead of three + a blank.
    auto_shift = 0
    if grain in ("week", "month"):
        # Probe up to 4 periods back; if everything is empty, stop shifting
        # and just render the empties so the user can see the gap.
        for probe in range(0, 5):
            pb = _plh_bucket_keys(grain, offset=probe)
            ls, le = pb[-1][1], pb[-1][2]
            if not _bucket_is_empty(ls, le, prod, labor, bsales):
                auto_shift = probe
                break
    effective_offset = offset + auto_shift
    buckets = _plh_bucket_keys(grain, offset=effective_offset)

    out = []
    for key, start, end, label in buckets:
        bucket = {
            "key": key, "label": label, "start": start, "end": end,
            "total_cs": 0,
            "revenue_dollars": 0.0,            # cs * case price (production-side)
            "by_distributor": {},
            "labor_hours": 0.0,
            "labor_dollars": 0.0,
            "bakery_sales_dollars": 0.0,        # true sales from bakery_sales.json
            "bakery_sales_by_channel": {},
            "bakery_sales_weeks": 0,            # how many weekly rows landed here
            "splh": None,                       # sales per labor hour
            "labor_pct_of_sales": None,         # labor $ / sales $
            "plh": None,                        # legacy: revenue / hours
        }
        for r in prod:
            if not _date_in_range(r.get("production_date") or "", start, end):
                continue
            dist = r.get("distributor") or ""
            price = _case_price_for(dist)
            for line in (r.get("lines") or []):
                cs = int(line.get("cs_count") or 0)
                bucket["total_cs"] += cs
                bucket["revenue_dollars"] += cs * price
                bdist = bucket["by_distributor"].setdefault(
                    dist or "Other", {"cs": 0, "revenue_dollars": 0.0})
                bdist["cs"] += cs
                bdist["revenue_dollars"] += cs * price
        # Track bakery-xlsx labor separately so the sales-vs-labor metrics
        # (SPLH, labor % of sales) reconcile to the workbook's I11/O10
        # numbers exactly. The "all labor" pool still drives total labor
        # hours/cost and the production-side $PLH, since those reflect the
        # full picture of what the bakery actually paid out.
        bx_hours = 0.0
        bx_dollars = 0.0
        for e in labor:
            if not _date_in_range(e.get("date") or "", start, end):
                continue
            hrs = float(e.get("hours") or 0)
            dol = float(e.get("dollars") or 0)
            bucket["labor_hours"]   += hrs
            bucket["labor_dollars"] += dol
            if (e.get("source") or "") == "bakery-xlsx":
                bx_hours   += hrs
                bx_dollars += dol
        bucket["bakery_xlsx_labor_hours"]   = round(bx_hours, 4)
        bucket["bakery_xlsx_labor_dollars"] = round(bx_dollars, 2)
        # Aggregate bakery weekly sales whose week_start falls in this
        # bucket. A week is counted toward whichever bucket its Monday
        # lands in -- accepting the small straddle effect at month/
        # quarter boundaries in exchange for clean, deterministic math.
        for w in bsales:
            ws = (w.get("week_start") or "")
            if not _date_in_range(ws, start, end):
                continue
            bucket["bakery_sales_dollars"] += float(w.get("total") or 0)
            bucket["bakery_sales_weeks"]   += 1
            for ch, amt in (w.get("channels") or {}).items():
                bucket["bakery_sales_by_channel"][ch] = (
                    bucket["bakery_sales_by_channel"].get(ch, 0.0)
                    + float(amt or 0)
                )
        # If we have hours but no dollars on any entry, impute via the default
        # rate so the cost picture is at least directionally right.
        if bucket["labor_hours"] and not bucket["labor_dollars"]:
            bucket["labor_dollars"] = bucket["labor_hours"] * LABOR_RATE_DEFAULT
            bucket["labor_dollars_imputed"] = True
        if bucket["labor_hours"]:
            bucket["plh"] = bucket["revenue_dollars"] / bucket["labor_hours"]
        # Sales-vs-labor metrics ALWAYS use the bakery-xlsx labor pool so
        # the chart matches the workbook the user updates each week.
        if bx_hours and bucket["bakery_sales_dollars"]:
            bucket["splh"] = bucket["bakery_sales_dollars"] / bx_hours
        if bx_dollars and bucket["bakery_sales_dollars"]:
            bucket["labor_pct_of_sales"] = (
                bx_dollars / bucket["bakery_sales_dollars"]
            )
        out.append(bucket)

    return jsonify({
        "grain":            grain,
        "offset":           offset,
        "effective_offset": effective_offset,
        "auto_shift":       auto_shift,
        "window_label":     _plh_window_label(grain, buckets),
        "case_prices": {
            "US Foods": 27.00, "Cheney Brothers": 26.50, "default": 29.50,
        },
        "labor_rate_default": LABOR_RATE_DEFAULT,
        "buckets":          out,
    })


@reporting_bp.route("/api/admin/labor/ingest", methods=["POST"])
def api_admin_labor_ingest():
    """Append or replace bakery labor entries.

    Body:
      entries:  [{date: YYYY-MM-DD, hours: float, dollars?: float, source?: str}]
      replace:  bool — when true, REPLACE all entries from the same `source`
                with the new set; when false (default), append + dedupe by
                (date, source) keeping the new entry.

    No Toast call here — this is the "feed me labor data from any source"
    endpoint. A separate script can pull from Toast and POST. CSV upload
    or manual entry can use this same endpoint.
    """
    from inventory_tracker import load_labor, save_labor
    body = request.json or {}
    incoming = body.get("entries") or []
    replace = bool(body.get("replace", False))
    if not incoming:
        return jsonify({"ok": False, "error": "entries required"}), 400

    existing = load_labor()
    sources_in_incoming = {(e.get("source") or "manual") for e in incoming}
    if replace:
        existing = [e for e in existing if (e.get("source") or "manual")
                    not in sources_in_incoming]

    # Dedupe by (date, source). Newer wins.
    by_key = {}
    for e in existing:
        by_key[(e.get("date"), e.get("source") or "manual")] = e
    for e in incoming:
        d = (e.get("date") or "").strip()
        if not d:
            continue
        rec = {
            "date":    d,
            "hours":   float(e.get("hours") or 0),
            "dollars": float(e.get("dollars") or 0),
            "source":  e.get("source") or "manual",
            "ingested_at": datetime_now_iso(),
        }
        by_key[(rec["date"], rec["source"])] = rec
    merged = list(by_key.values())
    save_labor(merged)
    return jsonify({"ok": True, "total_entries": len(merged),
                    "added_or_updated": len(incoming)})


@reporting_bp.route("/api/admin/bakery-sales/ingest", methods=["POST"])
def api_admin_bakery_sales_ingest():
    """Append or replace bakery weekly sales entries.

    Body:
      entries:  [
        {
          week_start:  "YYYY-MM-DD",   # Monday -- required, used as the key
          week_end:    "YYYY-MM-DD",
          channels:    {channel: dollars, ...},
          total:       float,
          splh:        float | null,
          source:      str,            # default "bakery-xlsx"
        }, ...
      ]
      replace:  bool -- when true, REPLACE all entries that share a source
                with the incoming set; when false (default), dedupe by
                (week_start, source) keeping the new entry.
    """
    from inventory_tracker import load_bakery_sales, save_bakery_sales
    body = request.json or {}
    incoming = body.get("entries") or []
    replace = bool(body.get("replace", False))
    if not incoming:
        return jsonify({"ok": False, "error": "entries required"}), 400

    existing = load_bakery_sales()
    sources_in_incoming = {(e.get("source") or "bakery-xlsx") for e in incoming}
    if replace:
        existing = [e for e in existing
                    if (e.get("source") or "bakery-xlsx")
                    not in sources_in_incoming]

    by_key = {}
    for e in existing:
        by_key[(e.get("week_start"), e.get("source") or "bakery-xlsx")] = e
    added = 0
    for e in incoming:
        ws_ = (e.get("week_start") or "").strip()
        if not ws_:
            continue
        rec = {
            "week_start": ws_,
            "week_end":   (e.get("week_end") or "").strip() or None,
            "location":   e.get("location") or "Bakery",
            "channels":   {k: float(v or 0) for k, v in
                           (e.get("channels") or {}).items()},
            "total":      float(e.get("total") or 0),
            "splh":       (float(e["splh"]) if e.get("splh") not in
                           (None, "", "#DIV/0!") else None),
            "source":     e.get("source") or "bakery-xlsx",
            "ingested_at": datetime_now_iso(),
        }
        by_key[(rec["week_start"], rec["source"])] = rec
        added += 1
    merged = sorted(by_key.values(), key=lambda r: r.get("week_start") or "")
    save_bakery_sales(merged)
    return jsonify({"ok": True, "total_entries": len(merged),
                    "added_or_updated": added})


@reporting_bp.route("/api/report/bakery-sales")
def api_report_bakery_sales():
    """Return bakery weekly sales rows, optionally filtered by date range.

    Query params:
      start  YYYY-MM-DD  (inclusive -- compared against week_start)
      end    YYYY-MM-DD  (inclusive -- compared against week_start)
      limit  int         (default 26 -- most recent N weeks)
    """
    from inventory_tracker import load_bakery_sales
    rows = load_bakery_sales() or []
    args = request.args
    start = (args.get("start") or "").strip()
    end   = (args.get("end") or "").strip()
    if start:
        rows = [r for r in rows if (r.get("week_start") or "") >= start]
    if end:
        rows = [r for r in rows if (r.get("week_start") or "") <= end]
    try:
        limit = int(args.get("limit") or 26)
    except (TypeError, ValueError):
        limit = 26
    rows = sorted(rows, key=lambda r: r.get("week_start") or "")
    if limit and limit > 0:
        rows = rows[-limit:]
    channel_totals: dict[str, float] = {}
    grand_total = 0.0
    for r in rows:
        grand_total += float(r.get("total") or 0)
        for ch, v in (r.get("channels") or {}).items():
            channel_totals[ch] = channel_totals.get(ch, 0.0) + float(v or 0)
    return jsonify({
        "rows": rows,
        "summary": {
            "weeks":          len(rows),
            "total":          grand_total,
            "channel_totals": channel_totals,
        },
    })


@reporting_bp.route("/api/forecast/decrement-daily", methods=["POST"])
def api_forecast_decrement_daily():
    """Apply one day's worth of forecast usage to every SKU.

    Idempotent: for each SKU, if a non-reversed entry already exists today
    with source == "forecast-daily", that SKU is skipped. Safe to re-run
    or to call mid-day after the cron.

    Body (all optional):
      dry_run             bool   compute deltas without writing
      date                str    YYYY-MM-DD override for testing (defaults
                                 to today UTC). The idempotency check keys
                                 off this value.
      warehouse_prefix    str    restrict to SKUs whose `warehouse` field
                                 starts with this string (e.g. "Zebulon").
                                 Default: all SKUs with weekly_usage > 0.

    Returns {ok, date, applied: [...], skipped_existing: [...], dry_run}.
    """
    from datetime import datetime, timezone
    from inventory_tracker import (
        load_inventory, save_inventory, load_usage, save_usage,
        _variety_from_name,
    )

    body = request.json or {}
    dry_run = bool(body.get("dry_run"))
    date_iso = (body.get("date") or "").strip()
    wh_prefix = (body.get("warehouse_prefix") or "").strip()
    if not date_iso:
        date_iso = datetime.now(timezone.utc).date().isoformat()

    inv = load_inventory()
    usage = load_usage()

    # Index existing forecast-daily entries by (item_key, YYYY-MM-DD).
    # Reversed entries don't count — the slot is "open" to re-fill.
    have_today: set[tuple[str, str]] = set()
    for e in usage:
        if e.get("source") != "forecast-daily":
            continue
        if e.get("reversed"):
            continue
        ts = (e.get("timestamp") or "")[:10]
        if not ts:
            continue
        have_today.add((e.get("item_key") or "", ts))

    applied: list[dict] = []
    skipped: list[str] = []
    now_iso = datetime.now(timezone.utc).isoformat()
    for key, item in inv.items():
        wkly = float(item.get("weekly_usage") or 0)
        if wkly <= 0:
            continue
        wh = item.get("warehouse") or ""
        if wh_prefix and not wh.startswith(wh_prefix):
            continue
        if (key, date_iso) in have_today:
            skipped.append(item.get("name") or key)
            continue
        daily_amount = round(wkly / 7.0, 4)
        if daily_amount <= 0:
            continue
        cur_qty = float(item.get("quantity") or 0)
        new_qty = round(cur_qty - daily_amount, 4)
        # Floor at 0 — we don't model negative inventory; the operator
        # will see the SKU pinned at 0 and reorder.
        if new_qty < 0:
            new_qty = 0.0
            daily_amount = round(cur_qty, 4)
            if daily_amount <= 0:
                # Already at zero; nothing to decrement, but record
                # nothing so the SKU stays eligible tomorrow.
                continue
        applied.append({
            "name": item.get("name"),
            "warehouse": wh,
            "weekly_usage": wkly,
            "daily_amount": daily_amount,
            "old_quantity": cur_qty,
            "new_quantity": new_qty,
        })
        if not dry_run:
            entry = {
                "item_key": key,
                "item_name": item.get("name") or "",
                "amount": daily_amount,
                "unit": item.get("unit") or "",
                "note": f"Forecast burn (weekly_usage/7) for {date_iso}",
                "timestamp": now_iso,
                "source": "forecast-daily",
                "forecast_date": date_iso,
            }
            if wh:
                entry["warehouse"] = wh
            v = _variety_from_name(item.get("name") or "")
            if v:
                entry["variety"] = v
            usage.append(entry)
            item["quantity"] = new_qty
            item["updated"] = now_iso

    if not dry_run and applied:
        save_usage(usage)
        save_inventory(inv)

    return jsonify({
        "ok": True,
        "date": date_iso,
        "dry_run": dry_run,
        "applied": applied,
        "skipped_existing": skipped,
    })


@reporting_bp.route("/api/forecast/backfill-historical", methods=["POST"])
def api_forecast_backfill_historical():
    """One-shot opening reconciliation between lot production and on-hand.

    Why this exists:
        Production lots get recorded as they arrive (each PO bakes new
        lots into data/production.json), but FIFO consumption only ever
        knew about *future* burns through the usage ledger. There was no
        equivalent for the "between PO arrival and the first vendor
        snapshot" gap — so the lots-by-pair view sat at cs_consumed=0
        forever, even on SKUs whose on-hand was clearly much smaller
        than total cases received.

    What this does:
        For each (warehouse, variety) pair on an inventory item:

            target_consumed = max(0, sum(lots_produced) - quantity)
            existing_consumed = sum(positive non-reversed usage entries)
            delta = target_consumed - existing_consumed

        If delta > 0, append one positive usage event of `delta` tagged
        `source = "historical-backfill"` so FIFO drains oldest-first
        until the lot remaining sums match the on-hand quantity. The
        endpoint does NOT touch item.quantity — quantity is already
        ground truth from the vendor; this purely seeds the consumption
        side of the ledger.

    Body (all optional):
        warehouse   str    limit to one warehouse (e.g. "Zebulon, NC").
                           Default: every warehouse.
        dry_run     bool   compute deltas without writing.

    Returns {ok, applied: [...], no_op: [...], dry_run}.
    """
    from datetime import datetime, timezone
    from inventory_tracker import (
        load_inventory, load_production, load_usage, save_usage,
        _variety_from_name,
    )

    body = request.json or {}
    warehouse = (body.get("warehouse") or "").strip()
    dry_run = bool(body.get("dry_run"))

    inv = load_inventory()
    prod = load_production()
    usage = load_usage()

    # Sum produced cases per (warehouse, variety).
    produced_by_pair: dict[tuple[str, str], float] = {}
    for r in prod:
        wh = r.get("warehouse") or ""
        for L in r.get("lines") or []:
            v = (L.get("variety") or "").strip()
            if not wh or not v:
                continue
            produced_by_pair[(wh, v)] = (
                produced_by_pair.get((wh, v), 0.0)
                + float(L.get("cs_count") or 0)
            )

    # Existing positive (non-reversed) consumption per (warehouse, variety).
    inv_for_lookup = inv
    consumed_by_pair: dict[tuple[str, str], float] = {}
    for e in usage:
        if e.get("reversed"):
            continue
        if e.get("source") == "reversal":
            continue
        amt = float(e.get("amount") or 0)
        if amt <= 0:
            continue
        e_wh = e.get("warehouse") or ""
        e_var = e.get("variety") or ""
        if not e_wh or not e_var:
            it = inv_for_lookup.get(e.get("item_key") or "") or {}
            e_wh = e_wh or it.get("warehouse", "")
            e_var = e_var or _variety_from_name(
                it.get("name") or e.get("item_name") or "")
        if not e_wh or not e_var:
            continue
        consumed_by_pair[(e_wh, e_var)] = (
            consumed_by_pair.get((e_wh, e_var), 0.0) + amt
        )

    applied: list[dict] = []
    no_op: list[dict] = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for key, item in inv.items():
        wh = item.get("warehouse") or ""
        if warehouse and wh != warehouse:
            continue
        name = item.get("name") or ""
        v = _variety_from_name(name)
        if not wh or not v:
            continue
        produced = produced_by_pair.get((wh, v), 0.0)
        if produced <= 0:
            continue
        cur_qty = float(item.get("quantity") or 0)
        target_consumed = max(0.0, produced - cur_qty)
        existing = consumed_by_pair.get((wh, v), 0.0)
        delta = round(target_consumed - existing, 4)
        if delta <= 0.001:
            no_op.append({
                "name": name, "warehouse": wh,
                "produced": produced, "quantity": cur_qty,
                "target_consumed": target_consumed,
                "existing_consumed": existing,
            })
            continue
        applied.append({
            "name": name, "warehouse": wh,
            "produced": produced, "quantity": cur_qty,
            "target_consumed": target_consumed,
            "existing_consumed": existing,
            "backfill_amount": delta,
        })
        if not dry_run:
            usage.append({
                "item_key": key,
                "item_name": name,
                "amount": delta,
                "unit": item.get("unit") or "cs",
                "note": (f"Opening reconciliation: align FIFO lots "
                         f"with vendor on-hand ({cur_qty} cs)"),
                "timestamp": now_iso,
                "source": "historical-backfill",
                "warehouse": wh,
                "variety": v,
            })

    if not dry_run and applied:
        save_usage(usage)

    return jsonify({
        "ok": True,
        "warehouse": warehouse or "(all)",
        "dry_run": dry_run,
        "applied": applied,
        "no_op": no_op,
    })


@reporting_bp.route("/api/forecast/true-up", methods=["POST"])
def api_forecast_true_up():
    """Reconcile against a vendor on-hand snapshot.

    The vendor's number is ground truth. We reverse every uncovered
    "forecast-daily" entry on each named SKU since the last vendor-truth
    (or since the SKU was added if there has never been one), then post a
    single "vendor-truth" usage entry equal to:

        actual_used = prior_truth_qty + arrivals_since - reported_qty

    where:
        prior_truth_qty   = item.quantity (live) + forecast_burned - arrivals_since
                            (i.e. what on-hand WAS at the last truth date)
        arrivals_since    = sum of negative-amount entries since last truth
        forecast_burned   = sum of positive forecast-daily entries since last
                            truth that are NOT already reversed
        reported_qty      = the new vendor snapshot

    If actual_used <= 0 we skip the positive entry (vendor reports more
    than expected — that's a count correction, not a consumption event).

    Quantity is set to reported_qty in inventory.json so the dashboard
    matches the vendor's number.

    Body:
      warehouse     str    label that matches inventory items' `warehouse`
                           field (e.g. "Zebulon, NC"). Required.
      reported_at   str    ISO date/datetime of the snapshot. Defaults to
                           now (UTC).
      items         list   [{name: str, reported_qty: float}, ...]

    Returns {ok, reported_at, reconciled: [...], skipped: [...]}.
    """
    from datetime import datetime, timezone
    from inventory_tracker import (
        load_inventory, save_inventory, load_usage, save_usage,
        _variety_from_name,
    )

    body = request.json or {}
    warehouse = (body.get("warehouse") or "").strip()
    if not warehouse:
        return jsonify({"ok": False, "error": "warehouse required"}), 400
    items_in = body.get("items") or []
    if not isinstance(items_in, list) or not items_in:
        return jsonify({"ok": False, "error": "items must be a non-empty list"}), 400
    reported_at = (body.get("reported_at") or "").strip()
    if not reported_at:
        reported_at = datetime.now(timezone.utc).isoformat()
    dry_run = bool(body.get("dry_run"))

    inv = load_inventory()
    usage = load_usage()

    # Build a lookup of usage entries by item_key for fast filtering.
    by_key: dict[str, list[dict]] = {}
    for idx, e in enumerate(usage):
        k = e.get("item_key") or ""
        if k:
            by_key.setdefault(k, []).append((idx, e))

    reconciled: list[dict] = []
    skipped: list[dict] = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for payload in items_in:
        name = (payload.get("name") or "").strip()
        if not name:
            skipped.append({"name": "", "reason": "missing name"})
            continue
        try:
            reported_qty = float(payload.get("reported_qty"))
        except (TypeError, ValueError):
            skipped.append({"name": name, "reason": "reported_qty not numeric"})
            continue
        key = name.lower()
        item = inv.get(key)
        if not item:
            skipped.append({"name": name, "reason": "SKU not found"})
            continue
        if (item.get("warehouse") or "") != warehouse:
            skipped.append({"name": name, "reason": f"warehouse mismatch ({item.get('warehouse')})"})
            continue

        # Find last vendor-truth timestamp for this SKU (exclusive lower
        # bound for the reconciliation window). Empty string = no prior
        # truth, so window is "all history".
        prior_truth_ts = ""
        for _idx, e in by_key.get(key, []):
            if e.get("source") != "vendor-truth":
                continue
            if e.get("reversed"):
                continue
            ts = e.get("timestamp") or ""
            if ts > prior_truth_ts:
                prior_truth_ts = ts

        arrivals_since = 0.0       # sum of negative amounts (restocks)
        forecast_burned = 0.0      # positive forecast-daily not reversed
        reverse_indices: list[int] = []
        for idx, e in by_key.get(key, []):
            if e.get("reversed"):
                continue
            ts = e.get("timestamp") or ""
            if prior_truth_ts and ts <= prior_truth_ts:
                continue
            amt = float(e.get("amount") or 0)
            if amt < 0:
                arrivals_since += -amt
            elif amt > 0 and e.get("source") == "forecast-daily":
                forecast_burned += amt
                reverse_indices.append(idx)
            # Positive non-forecast entries (manual /api/use, prior
            # vendor-truth) are left untouched — they represent real
            # consumption the operator already booked.

        cur_qty = float(item.get("quantity") or 0)
        prior_truth_qty = cur_qty + forecast_burned - arrivals_since
        actual_used = prior_truth_qty + arrivals_since - reported_qty
        actual_used = round(actual_used, 4)

        entry_record = {
            "name": name,
            "warehouse": warehouse,
            "prior_truth_ts": prior_truth_ts or None,
            "prior_truth_qty": round(prior_truth_qty, 4),
            "arrivals_since": round(arrivals_since, 4),
            "forecast_burned_reversed": round(forecast_burned, 4),
            "reported_qty": reported_qty,
            "actual_used": actual_used,
            "old_quantity": cur_qty,
            "new_quantity": reported_qty,
        }
        reconciled.append(entry_record)

        if dry_run:
            continue

        # Reverse the forecast-daily entries that fell in this window.
        for idx in reverse_indices:
            usage[idx]["reversed"] = True
            usage[idx]["reversed_at"] = now_iso
            usage[idx]["reversed_by"] = "vendor-truth"

        # Post the vendor-truth event. Skip when actual_used <= 0
        # (vendor reported MORE than expected — a positive count
        # correction we model as the quantity bump only, no consumption).
        if actual_used > 0:
            truth_entry = {
                "item_key": key,
                "item_name": item.get("name") or name,
                "amount": actual_used,
                "unit": item.get("unit") or "",
                "note": f"Vendor on-hand true-up at {reported_at}",
                "timestamp": now_iso,
                "source": "vendor-truth",
                "reported_at": reported_at,
                "warehouse": warehouse,
            }
            v = _variety_from_name(item.get("name") or name)
            if v:
                truth_entry["variety"] = v
            usage.append(truth_entry)

        item["quantity"] = reported_qty
        item["updated"] = now_iso
        item["last_synced"] = now_iso
        item["last_synced_from"] = "vendor-truth"
        item["last_count_at"] = now_iso

    if not dry_run and reconciled:
        save_usage(usage)
        save_inventory(inv)

    return jsonify({
        "ok": True,
        "warehouse": warehouse,
        "reported_at": reported_at,
        "dry_run": dry_run,
        "reconciled": reconciled,
        "skipped": skipped,
    })


@reporting_bp.route("/api/report")
def api_report():
    inv = load_inventory()
    usage = load_usage()

    total_value = sum(i["quantity"] * i["price"] for i in inv.values())
    total_items = len(inv)
    low_stock = [i for i in inv.values() if i["quantity"] <= i["low_stock_threshold"]]

    consumed: dict = {}
    restocked: dict = {}
    for e in usage:
        key = e["item_key"]
        if e["amount"] < 0:
            restocked[key] = restocked.get(key, 0) + abs(e["amount"])
        else:
            consumed[key] = consumed.get(key, 0) + e["amount"]

    top_consumed = sorted(
        [{"key": k, "name": inv.get(k, {}).get("name", k),
          "unit": inv.get(k, {}).get("unit", ""), "total": v}
         for k, v in consumed.items()],
        key=lambda x: x["total"], reverse=True
    )[:10]

    top_restocked = sorted(
        [{"key": k, "name": inv.get(k, {}).get("name", k),
          "unit": inv.get(k, {}).get("unit", ""), "total": v}
         for k, v in restocked.items()],
        key=lambda x: x["total"], reverse=True
    )[:10]

    return jsonify({
        "total_value": round(total_value, 2),
        "total_items": total_items,
        "low_stock_count": len(low_stock),
        "low_stock": low_stock,
        "top_consumed": top_consumed,
        "top_restocked": top_restocked,
        "total_usage_events": len(usage),
    })


@reporting_bp.route("/api/report-status")
def api_report_status():
    """JSON: which distributor warehouses have sent this week\'s inventory & usage report."""
    from integrations import report_status as _rs
    force = request.args.get("refresh") in ("1", "true", "yes", "on")
    return jsonify(_rs.get_status(force=force))


@reporting_bp.route("/report-status")
def report_status_page():
    """Mobile-friendly weekly-report status page (read-only; requires login)."""
    from flask import Response
    from integrations import report_status as _rs
    force = request.args.get("refresh") in ("1", "true", "yes", "on")
    html = _rs.render_html(_rs.get_status(force=force))
    return Response(html, mimetype="text/html")


@reporting_bp.route("/api/export.xlsx")
def api_export_xlsx():
    from openpyxl import Workbook
    from export_bagels_xlsx import _write_summary_sheet, _write_items_sheet

    inv = load_inventory()
    items = list(inv.values())
    cheney = [i for i in items if (i.get("distributor") or "") == "Cheney Brothers"]
    usfoods = [i for i in items if (i.get("distributor") or "") == "US Foods"]

    wb = Workbook()
    _write_summary_sheet(wb.active, inv)
    wb.active.title = "Summary"
    _write_items_sheet(wb.create_sheet("Unified List"), items)
    _write_items_sheet(wb.create_sheet("Cheney Brothers"), cheney)
    _write_items_sheet(wb.create_sheet("US Foods"), usfoods)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="bagel_inventory.xlsx",
    )
