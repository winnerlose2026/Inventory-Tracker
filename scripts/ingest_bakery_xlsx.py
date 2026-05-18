#!/usr/bin/env python3
"""Ingest the "Bakery Model - Sales v. Labor" workbook into the inventory
tracker.

Pulled here because the production bakery isn't wired to Toast yet — JD
updates this spreadsheet weekly and reuploads it. Each sheet is one ISO
week (Mon-Sun) named like ``5.11.26 - 5.17.26`` (or earlier sheets like
``7.19 - 7.25`` with no year). Sheet ordering is the source of truth for
which week a sheet represents — we walk 7 days at a time from
2021-07-19 (sheet 0) and cross-check against any explicit year stamped
in the sheet name.

What we read from each sheet
----------------------------
    B3..H3    daily sales (Mon..Sun)            -- weekly_total / 7
    I3        weekly total sales
    B10..H10  daily labor $ (Mon..Sun)          -- REAL daily values
    B11..H11  daily labor hours (Mon..Sun)      -- REAL daily values
    I10       weekly total labor $
    I11       weekly total labor hours
    I13       Labor $ - Gabo & Taxes (alternate metric -- not stored)
    N3..N8    selling channel names
    O3..O8    selling channel sale $
    O10       SPLH (sales per labor hour)

What we push
------------
    POST /api/admin/labor/ingest          (daily labor entries, source=bakery-xlsx)
    POST /api/admin/bakery-sales/ingest   (weekly sales rows,    source=bakery-xlsx)

Both calls use ``replace=True`` so every run is fully idempotent against
its source tag — old bakery-xlsx rows are wiped and replaced with the
fresh parse.

Usage
-----
    INVENTORY_API_TOKEN=... python scripts/ingest_bakery_xlsx.py \
        --xlsx "../Bakery Model - Sales v. Labor.xlsx" \
        [--api-base https://bagel-inventory.onrender.com] \
        [--dry-run]
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

import openpyxl


WORKBOOK_START = date(2021, 7, 19)
DOW_COLS = ["B", "C", "D", "E", "F", "G", "H"]
CHANNEL_ROWS = list(range(3, 9))  # N3..N8

DEFAULT_API = os.environ.get(
    "INVENTORY_API_BASE",
    "https://bagel-inventory.onrender.com",
)
DEFAULT_XLSX = "Bakery Model - Sales v. Labor.xlsx"
SOURCE_TAG = "bakery-xlsx"


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

_NAME_RE = re.compile(
    r"^\s*(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?\s*-\s*"
    r"(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?\s*$"
)


def _parse_sheet_name(name: str):
    m = _NAME_RE.match(name)
    if not m:
        return None, None
    sm, sd, sy, em, ed, ey = m.groups()
    yr_s = int(sy) if sy else None
    yr_e = int(ey) if ey else None
    if yr_s is not None and yr_s < 100:
        yr_s += 2000
    if yr_e is not None and yr_e < 100:
        yr_e += 2000
    start = end = None
    if yr_s:
        try:
            start = date(yr_s, int(sm), int(sd))
        except ValueError:
            pass
    if yr_e:
        try:
            end = date(yr_e, int(em), int(ed))
        except ValueError:
            pass
    return start, end


def _walk_week(idx: int):
    ws = WORKBOOK_START + timedelta(days=7 * idx)
    we = ws + timedelta(days=6)
    return ws, we


def _resolve_week(idx: int, sheet_name: str):
    """Pick the authoritative (week_start, week_end). Prefer the parsed
    name when it has explicit years AND agrees with the walked date.
    Otherwise trust the walk."""
    walked_s, walked_e = _walk_week(idx)
    parsed_s, _ = _parse_sheet_name(sheet_name)
    if parsed_s and parsed_s == walked_s:
        return walked_s, walked_e, "match"
    if parsed_s and parsed_s != walked_s:
        return walked_s, walked_e, (
            "drift: walked=" + str(walked_s) +
            " parsed=" + str(parsed_s) + " -- trusting walk"
        )
    return walked_s, walked_e, "no-year"


def _num(v) -> float:
    """Best-effort numeric coercion. Returns 0.0 for empty / DIV-error
    sentinels."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, str):
        s = v.strip()
        if s.startswith("#") or s.upper() in {"NA", "N/A"}:
            return 0.0
        try:
            return float(s.replace(",", "").replace("$", ""))
        except ValueError:
            return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _all_equal(xs, tol: float = 0.05) -> bool:
    """True when every value in xs is within tol of the first one. Used
    to detect formula-spread placeholders (weekly_total / 7 written into
    every daily cell) vs. real day-by-day actuals."""
    if not xs:
        return False
    base = xs[0]
    return all(abs(x - base) <= tol for x in xs)


def parse_sheet(ws, week_start):
    raw_dol = [_num(ws[col + "10"].value) for col in DOW_COLS]
    raw_hrs = [_num(ws[col + "11"].value) for col in DOW_COLS]
    is_forecast_dollars = _all_equal(raw_dol) and any(v > 0 for v in raw_dol)
    is_forecast_hours   = _all_equal(raw_hrs) and any(v > 0 for v in raw_hrs)

    labor_daily = []
    if not is_forecast_dollars and not is_forecast_hours:
        for i, col in enumerate(DOW_COLS):
            d = week_start + timedelta(days=i)
            hrs = raw_hrs[i]
            dol = raw_dol[i]
            if hrs <= 0 and dol <= 0:
                continue
            labor_daily.append({
                "date":    d.isoformat(),
                "hours":   round(hrs, 4),
                "dollars": round(dol, 2),
            })

    channels = {}
    for r in CHANNEL_ROWS:
        name = ws["N" + str(r)].value
        if not name or not isinstance(name, str):
            continue
        amt = _num(ws["O" + str(r)].value)
        if amt <= 0:
            continue
        channels[name.strip()] = round(amt, 2)

    total = _num(ws["I3"].value)
    splh  = _num(ws["O10"].value) or None

    return {
        "week_start":  week_start.isoformat(),
        "week_end":    (week_start + timedelta(days=6)).isoformat(),
        "labor_daily": labor_daily,
        "channels":    channels,
        "total":       round(total, 2),
        "splh":        round(splh, 4) if splh else None,
    }


def parse_workbook(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    labor_entries = []
    bakery_weeks = []
    drift_notes = []

    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        week_start, week_end, note = _resolve_week(idx, sheet_name)
        if note != "match" and note != "no-year":
            drift_notes.append("  [" + str(idx) + "] " + repr(sheet_name) + ": " + note)
        rec = parse_sheet(ws, week_start)
        if not rec["labor_daily"] and rec["total"] <= 0:
            continue

        for L in rec["labor_daily"]:
            labor_entries.append({
                "date":    L["date"],
                "hours":   L["hours"],
                "dollars": L["dollars"],
                "source":  SOURCE_TAG,
            })
        bakery_weeks.append({
            "week_start": rec["week_start"],
            "week_end":   rec["week_end"],
            "location":   "Bakery",
            "channels":   rec["channels"],
            "total":      rec["total"],
            "splh":       rec["splh"],
            "source":     SOURCE_TAG,
        })

    if drift_notes:
        print("Sheet-name drift detected (walk wins):", file=sys.stderr)
        for n in drift_notes:
            print(n, file=sys.stderr)

    return labor_entries, bakery_weeks


# ---------------------------------------------------------------------------
# Posting
# ---------------------------------------------------------------------------

def _post(url: str, body: dict, token):
    data = json.dumps(body).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = "Bearer " + token
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        msg = e.read().decode("utf-8", errors="replace")
        raise SystemExit("POST " + url + " -> " + str(e.code) + ": " + msg)


def push(api_base: str, labor, weeks, token):
    if labor:
        out = _post(api_base.rstrip("/") + "/api/admin/labor/ingest",
                    {"entries": labor, "replace": True}, token)
        print("labor:        " + json.dumps(out))
    if weeks:
        out = _post(api_base.rstrip("/") + "/api/admin/bakery-sales/ingest",
                    {"entries": weeks, "replace": True}, token)
        print("bakery-sales: " + json.dumps(out))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    p.add_argument("--xlsx", default=DEFAULT_XLSX,
                   help="Path to the workbook (default: %(default)s)")
    p.add_argument("--api-base", default=DEFAULT_API,
                   help="Base URL of the inventory tracker API")
    p.add_argument("--token", default=os.environ.get("INVENTORY_API_TOKEN"),
                   help="Bearer token (or set INVENTORY_API_TOKEN)")
    p.add_argument("--dry-run", action="store_true",
                   help="Parse and print a summary; don't POST")
    args = p.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print("ERROR: workbook not found at " + str(path), file=sys.stderr)
        return 2

    labor, weeks = parse_workbook(path)
    print("Parsed " + str(len(weeks)) + " non-empty week(s), " +
          str(len(labor)) + " daily labor entries.")
    if weeks:
        first, last = weeks[0]["week_start"], weeks[-1]["week_start"]
        print("Week range: " + first + " .. " + last)
        for w in weeks[-3:]:
            ch = ", ".join(k + "=$" + format(v, ",.0f")
                           for k, v in w["channels"].items())
            print("  " + w["week_start"] +
                  "  total=$" + format(w["total"], ",.2f") +
                  "  [" + ch + "]")

    if args.dry_run:
        print("(dry-run -- no POST)")
        return 0

    if not args.token:
        print("ERROR: --token or INVENTORY_API_TOKEN required for POST",
              file=sys.stderr)
        return 2

    push(args.api_base, labor, weeks, args.token)
    return 0


if __name__ == "__main__":
    sys.exit(main())
